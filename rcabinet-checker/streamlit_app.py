"""
R-Cabinet 管理ツール
- フォルダ画像一覧：R-Cabinetのフォルダ毎に画像を一覧表示
- 画像存在チェック：コミックNoを入力して存在確認
"""

# バージョン（デプロイ確認用）
APP_VERSION = "5.0.0"

import streamlit as st
import requests
import base64
import xml.etree.ElementTree as ET
import pandas as pd
import time
import json
import re
from io import BytesIO
from datetime import datetime, timezone, timedelta

JST = timezone(timedelta(hours=9))

# 重いライブラリは遅延読み込み（起動高速化）
_bs4_module = None
_openpyxl_styles = None
_openpyxl_utils = None
_supabase_module = None
_zipfile_module = None
_random_module = None
_pil_module = None

# Gemini AI（遅延読み込み - 起動高速化のため）
GEMINI_AVAILABLE = None
_genai_module = None


def get_bs4():
    """BeautifulSoupを遅延読み込み"""
    global _bs4_module
    if _bs4_module is None:
        from bs4 import BeautifulSoup
        _bs4_module = BeautifulSoup
    return _bs4_module


def get_openpyxl_styles():
    """openpyxlスタイルを遅延読み込み"""
    global _openpyxl_styles, _openpyxl_utils
    if _openpyxl_styles is None:
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        _openpyxl_styles = {'Font': Font, 'Border': Border, 'Side': Side, 'PatternFill': PatternFill, 'Alignment': Alignment}
        _openpyxl_utils = {'get_column_letter': get_column_letter}
    return _openpyxl_styles, _openpyxl_utils


def get_supabase_module():
    """Supabaseを遅延読み込み"""
    global _supabase_module
    if _supabase_module is None:
        from supabase import create_client
        _supabase_module = create_client
    return _supabase_module


def get_zipfile():
    """zipfileを遅延読み込み"""
    global _zipfile_module
    if _zipfile_module is None:
        import zipfile
        _zipfile_module = zipfile
    return _zipfile_module


def get_random():
    """randomを遅延読み込み"""
    global _random_module
    if _random_module is None:
        import random
        _random_module = random
    return _random_module


def get_pil():
    """PILを遅延読み込み"""
    global _pil_module
    if _pil_module is None:
        from PIL import Image
        _pil_module = Image
    return _pil_module

# ページ設定
st.set_page_config(
    page_title="R-Cabinet 管理ツール",
    page_icon="🖼️",
    layout="wide"
)

# 認証情報（Streamlit Secretsから取得）
APP_PASSWORD = st.secrets.get("password", "")
SERVICE_SECRET = st.secrets.get("RMS_SERVICE_SECRET", "")
LICENSE_KEY = st.secrets.get("RMS_LICENSE_KEY", "")
BASE_URL = "https://api.rms.rakuten.co.jp/es/1.0"

# Supabase接続情報
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")

# GitHub接続情報
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO = "uraraka-axis/tools"
GITHUB_MISSING_CSV_PATH = "comic-lister/data/missing_comics.csv"
GITHUB_MISSING_TANPIN_PATH = "comic-lister/data/missing_tanpin.csv"  # 単品用
GITHUB_MISSING_YOYAKU_PATH = "comic-lister/data/missing_yoyaku.csv"  # 予約用
GITHUB_IS_LIST_PATH = "comic-lister/data/is_list.csv"

# 画像存在チェック対象フォルダ（種別 → R-Cabinetのフォルダパス prefix）
# folder_path が指定prefixで始まるフォルダをすべて対象とする（サブフォルダ含む）
CHECK_TARGET_FOLDERS = {
    "セット品": "/comic/comic-set",      # セット配下（セット1, セット2 等含む）
    "単品": "/comic/comic-tanpin",        # 単品配下
    "予約": "/comic/comic-yoyaku",        # 予約配下
}
GITHUB_COMIC_LIST_PATH = "comic-lister/data/comic_list.csv"
GITHUB_FOLDER_HIERARCHY_PATH = "comic-lister/data/folder_hierarchy.xlsx"

# 楽天RMS画像フォルダ管理シート: path prefix → シート名
FOLDER_MANAGEMENT_SHEETS = [
    ("/comic/comic-set",     "コミック・セット"),
    ("/comic/comic-tanpin",  "コミック・単品"),
    ("/comic/comic-yoyaku",  "コミック・予約"),
    ("/dvdblu",              "DVD・ブルーレイ"),
    ("/toy",                 "おもちゃ"),
    ("/calenda",             "カレンダー"),
    ("/st",                  "文房具"),
    ("/bk",                  "本、雑誌"),
    ("/kagu",                "家具"),
]

# Gemini API設定（セルフヒーリング用）
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY", "")


def get_gemini_model():
    """Gemini AIモデルを遅延読み込みで取得"""
    global GEMINI_AVAILABLE, _genai_module

    if GEMINI_AVAILABLE is None:
        try:
            import google.generativeai as genai
            _genai_module = genai
            GEMINI_AVAILABLE = True
        except ImportError:
            GEMINI_AVAILABLE = False
            return None

    if not GEMINI_AVAILABLE or not GEMINI_API_KEY:
        return None

    if _genai_module:
        _genai_module.configure(api_key=GEMINI_API_KEY)
        return _genai_module.GenerativeModel('gemini-2.0-flash')

    return None


def upload_to_github(content: str, path: str, message: str) -> dict:
    """GitHubにファイルをアップロード（上書き更新）"""
    if not GITHUB_TOKEN:
        return {"success": False, "error": "GITHUB_TOKEN未設定"}

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

    # 既存ファイルのSHAを取得（更新時に必要）
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    sha = None

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            sha = response.json().get("sha")
    except:
        pass

    # ファイルをアップロード
    data = {
        "message": message,
        "content": base64.b64encode(content.encode('utf-8')).decode('utf-8'),
        "branch": "master"
    }
    if sha:
        data["sha"] = sha

    try:
        response = requests.put(url, headers=headers, json=data)
        if response.status_code in [200, 201]:
            return {"success": True, "url": response.json().get("content", {}).get("html_url", "")}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:200]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def upload_binary_to_github(content: bytes, path: str, message: str) -> dict:
    """バイナリファイルをGitHubにアップロード（上書き更新）"""
    if not GITHUB_TOKEN:
        return {"success": False, "error": "GITHUB_TOKEN未設定"}

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

    # 既存ファイルのSHAを取得（更新時に必要）
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    sha = None

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            sha = response.json().get("sha")
    except:
        pass

    # ファイルをアップロード
    data = {
        "message": message,
        "content": base64.b64encode(content).decode('utf-8'),
        "branch": "master"
    }
    if sha:
        data["sha"] = sha

    try:
        response = requests.put(url, headers=headers, json=data)
        if response.status_code in [200, 201]:
            return {"success": True, "url": response.json().get("content", {}).get("html_url", "")}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:200]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def download_from_github(path: str) -> dict:
    """GitHubからファイルをダウンロード"""
    if not GITHUB_TOKEN:
        return {"success": False, "error": "GITHUB_TOKEN未設定"}

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3.raw"
    }

    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return {"success": True, "content": response.content, "path": path}
        elif response.status_code == 404:
            return {"success": False, "error": f"ファイルが見つかりません: {path}"}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@st.cache_data(ttl=60, show_spinner=False)
def fetch_github_csv_bytes(path: str, cache_bust: int = 0) -> bytes:
    """GitHubからCSVを取得してbytesで返す（download_buttonで使う安定データ用）。cache_bustで強制再取得可能"""
    result = download_from_github(path)
    if result.get("success"):
        content = result["content"]
        if isinstance(content, bytes):
            return content
        return content.encode('utf-8')
    return b""


def get_github_file_info(path: str) -> dict:
    """GitHubファイルの情報（更新日時など）を取得"""
    if not GITHUB_TOKEN:
        return {}

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

    url = f"https://api.github.com/repos/{GITHUB_REPO}/commits?path={path}&per_page=1"

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200 and response.json():
            commit = response.json()[0]
            date_str = commit.get("commit", {}).get("committer", {}).get("date", "")
            if date_str:
                # ISO形式をパースして日本時間に変換
                dt_utc = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                dt_jst = dt_utc.astimezone(JST)
                return {"last_updated": dt_jst.strftime("%Y-%m-%d %H:%M"), "exists": True}
        return {"exists": False}
    except:
        return {"exists": False}


def trigger_github_actions(workflow_file: str) -> dict:
    """GitHub Actionsワークフローを手動実行"""
    if not GITHUB_TOKEN:
        return {"success": False, "error": "GITHUB_TOKEN未設定"}

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

    url = f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/{workflow_file}/dispatches"

    try:
        response = requests.post(url, headers=headers, json={"ref": "master"})
        if response.status_code == 204:
            return {"success": True, "message": "ワークフローを開始しました"}
        elif response.status_code == 404:
            return {"success": False, "error": "ワークフローが見つかりません"}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:200]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_workflow_runs(workflow_file: str, limit: int = 3) -> list:
    """GitHub Actionsワークフローの実行履歴を取得"""
    if not GITHUB_TOKEN:
        return []

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

    url = f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/{workflow_file}/runs?per_page={limit}"

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            runs = response.json().get("workflow_runs", [])
            result = []
            for run in runs:
                created = run.get("created_at", "")
                if created:
                    dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                    created = dt.astimezone(JST).strftime("%Y-%m-%d %H:%M")
                result.append({
                    "status": run.get("status"),
                    "conclusion": run.get("conclusion"),
                    "created_at": created,
                    "html_url": run.get("html_url")
                })
            return result
        return []
    except:
        return []


@st.cache_resource
def get_supabase_client():
    """Supabaseクライアントを取得（遅延読み込み）"""
    if SUPABASE_URL and SUPABASE_KEY:
        create_client = get_supabase_module()
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    return None


def fetch_all_from_supabase(supabase, table: str, columns: str = "*", filter_col: str = None, filter_val: str = None) -> list:
    """Supabaseから全件取得（ページネーション対応）"""
    all_data = []
    page_size = 1000
    offset = 0

    while True:
        query = supabase.table(table).select(columns).range(offset, offset + page_size - 1)
        if filter_col and filter_val:
            query = query.ilike(filter_col, f"%{filter_val}%")
        response = query.execute()

        if not response.data:
            break

        all_data.extend(response.data)

        if len(response.data) < page_size:
            break

        offset += page_size

    return all_data


def sync_images_to_db(images: list) -> dict:
    """画像一覧をDBに同期（upsert）"""
    supabase = get_supabase_client()
    if not supabase:
        return {"success": False, "error": "Supabase未設定"}

    try:
        # file_nameごとにグループ化（重複検出）
        # file_urlはJSONで {フォルダ名: URL} 形式で保存（複数フォルダの場合でも正確なURLを保持）
        file_dict = {}
        for img in images:
            file_name = img.get("FileName", "")
            folder_name = img.get("FolderName", "")
            file_url = img.get("FileUrl", "")
            if file_name in file_dict:
                # 重複: folder_namesに追加、URLもフォルダ別に記録
                existing_folders = file_dict[file_name]["folder_names"].split(", ")
                if folder_name not in existing_folders:
                    file_dict[file_name]["folder_names"] += f", {folder_name}"
                    url_dict = json.loads(file_dict[file_name]["file_url"])
                    url_dict[folder_name] = file_url
                    file_dict[file_name]["file_url"] = json.dumps(url_dict, ensure_ascii=False)
            else:
                file_dict[file_name] = {
                    "file_name": file_name,
                    "folder_names": folder_name,
                    "file_url": json.dumps({folder_name: file_url}, ensure_ascii=False),
                    "file_size": img.get("FileSize", 0),
                    "file_timestamp": img.get("TimeStamp", "")
                }

        # 既存データを取得（ページネーション対応）
        existing_data = fetch_all_from_supabase(supabase, "rcabinet_images", "file_name, file_timestamp")
        existing_dict = {row["file_name"]: row["file_timestamp"] for row in existing_data}

        # 差分計算
        new_count = 0
        updated_count = 0
        duplicate_count = 0
        unchanged_count = 0

        records_to_upsert = []
        for file_name, record in file_dict.items():
            # 重複チェック（複数フォルダにある）
            if ", " in record["folder_names"]:
                duplicate_count += 1

            if file_name not in existing_dict:
                new_count += 1
                records_to_upsert.append(record)
            elif existing_dict[file_name] != record["file_timestamp"]:
                updated_count += 1
                records_to_upsert.append(record)
            else:
                unchanged_count += 1

        # 削除済み検出（DBにあるがAPIにない）
        deleted_files = set(existing_dict.keys()) - set(file_dict.keys())
        deleted_count = len(deleted_files)

        # upsert実行（100件ずつ）
        for i in range(0, len(records_to_upsert), 100):
            batch = records_to_upsert[i:i+100]
            supabase.table("rcabinet_images").upsert(
                batch, on_conflict="file_name"
            ).execute()

        # 削除済みファイルをDBから削除
        if deleted_files:
            for file_name in deleted_files:
                supabase.table("rcabinet_images").delete().eq("file_name", file_name).execute()

        return {
            "success": True,
            "new": new_count,
            "updated": updated_count,
            "duplicate": duplicate_count,
            "unchanged": unchanged_count,
            "deleted": deleted_count,
            "total": len(file_dict)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def load_images_from_db() -> tuple[list, str]:
    """DBから画像一覧を読み込み（ページネーション対応）"""
    supabase = get_supabase_client()
    if not supabase:
        return [], "Supabase未設定"

    try:
        all_data = fetch_all_from_supabase(supabase, "rcabinet_images", "*")
        images = []
        for row in all_data:
            folder_names_str = row.get("folder_names", "")
            file_url_raw = row.get("file_url", "")
            file_name = row.get("file_name", "")
            file_size = row.get("file_size", 0)
            file_timestamp = row.get("file_timestamp", "")
            # folder_path をパース（新形式: {フォルダ名: パス} の辞書 / 未設定の場合は None）
            folder_path_raw = row.get("folder_path", "")
            try:
                path_map = json.loads(folder_path_raw) if folder_path_raw else {}
                if not isinstance(path_map, dict):
                    path_map = {}
            except (json.JSONDecodeError, TypeError):
                path_map = {}
            # file_urlをJSONとして解析（新形式: {フォルダ名: URL} の辞書）
            try:
                url_data = json.loads(file_url_raw)
                if not isinstance(url_data, dict):
                    url_data = {}
            except (json.JSONDecodeError, TypeError):
                url_data = {}

            # folder_names は必ず分割して展開（file_urlの形式に依存しない）
            folder_list = [f.strip() for f in folder_names_str.split(", ") if f.strip()]
            if not folder_list:
                folder_list = [folder_names_str] if folder_names_str else [""]

            for folder in folder_list:
                if url_data:
                    url = url_data.get(folder, next(iter(url_data.values()), ""))
                else:
                    url = file_url_raw
                images.append({
                    "FolderName": folder,
                    "FolderPath": path_map.get(folder, ""),
                    "FileName": file_name,
                    "FileUrl": url,
                    "FileSize": file_size,
                    "TimeStamp": file_timestamp
                })
        return images, f"{len(images)}件を読み込みました"
    except Exception as e:
        return [], str(e)


def get_db_stats() -> dict:
    """DBの統計情報を取得（カウントクエリで高速化）"""
    supabase = get_supabase_client()
    if not supabase:
        return {}

    try:
        # 全件数をカウント（全件フェッチせず高速）
        total_resp = supabase.table("rcabinet_images").select("*", count="exact").limit(1).execute()
        total = total_resp.count or 0

        # 重複ファイル数（folder_namesにカンマを含む件数）
        dup_resp = supabase.table("rcabinet_images").select("*", count="exact").ilike("folder_names", "%, %").limit(1).execute()
        duplicates = dup_resp.count or 0

        return {"total": total, "duplicates": duplicates, "last_updated": None}
    except Exception:
        return {}


def load_images_from_db_by_folder(folder_name: str) -> list:
    """DBから特定フォルダの画像を読み込み（ページネーション対応）"""
    supabase = get_supabase_client()
    if not supabase:
        return []

    try:
        all_data = fetch_all_from_supabase(supabase, "rcabinet_images", "*", "folder_names", folder_name)
        images = []
        for row in all_data:
            folder_names_str = row.get("folder_names", "")
            file_url_raw = row.get("file_url", "")
            file_name = row.get("file_name", "")
            file_size = row.get("file_size", 0)
            file_timestamp = row.get("file_timestamp", "")
            try:
                url_data = json.loads(file_url_raw)
                if isinstance(url_data, dict) and url_data:
                    # 新形式: フォルダ別に行を展開
                    for folder in folder_names_str.split(", "):
                        folder = folder.strip()
                        if folder:
                            url = url_data.get(folder, next(iter(url_data.values()), ""))
                            images.append({
                                "FolderName": folder,
                                "FileName": file_name,
                                "FileUrl": url,
                                "FileSize": file_size,
                                "TimeStamp": file_timestamp
                            })
                else:
                    images.append({
                        "FolderName": folder_names_str,
                        "FileName": file_name,
                        "FileUrl": file_url_raw,
                        "FileSize": file_size,
                        "TimeStamp": file_timestamp
                    })
            except (json.JSONDecodeError, TypeError):
                images.append({
                    "FolderName": folder_names_str,
                    "FileName": file_name,
                    "FileUrl": file_url_raw,
                    "FileSize": file_size,
                    "TimeStamp": file_timestamp
                })
        return images
    except Exception:
        return []


def build_folder_management_xlsx(folders: list, files: list) -> bytes:
    """楽天RMS画像フォルダ管理シート形式のExcelを生成"""
    import openpyxl
    styles, utils = get_openpyxl_styles()
    Font = styles['Font']
    Alignment = styles['Alignment']
    PatternFill = styles['PatternFill']
    get_column_letter = utils['get_column_letter']

    path_to_name = {
        (f.get('FolderPath') or ''): (f.get('FolderName') or '')
        for f in folders if f.get('FolderPath')
    }

    def classify_sheet(folder_path: str):
        if not folder_path:
            return None
        for prefix, sheet in FOLDER_MANAGEMENT_SHEETS:
            if folder_path == prefix or folder_path.startswith(prefix + '/'):
                return sheet
        return None

    def split_path(folder_path: str):
        parts = [p for p in folder_path.strip('/').split('/') if p]
        d1 = '/' + parts[0] if len(parts) >= 1 else None
        d2 = '/' + '/'.join(parts[:2]) if len(parts) >= 2 else None
        d3 = '/' + '/'.join(parts[:3]) if len(parts) >= 3 else None
        return d1, d2, d3

    wb = openpyxl.Workbook()

    # シート1: フォルダ一覧
    ws = wb.active
    ws.title = "フォルダ一覧"
    ws.append(["No.", "フォルダ名", "ディレクトリパス", "フォルダID"])
    for i, f in enumerate(folders, start=1):
        ws.append([i, f.get('FolderName', ''), f.get('FolderPath', ''), f.get('FolderId', '')])

    # カテゴリ別シート（空でも作成）
    sheet_rows = {sheet: [] for _, sheet in FOLDER_MANAGEMENT_SHEETS}
    for f in files:
        sheet_name = classify_sheet(f.get('FolderPath', ''))
        if not sheet_name:
            continue
        d1, d2, d3 = split_path(f.get('FolderPath', ''))
        c1 = path_to_name.get(d1) if d1 else None
        c2 = path_to_name.get(d2) if d2 else None
        c3 = path_to_name.get(d3) if d3 else None
        sheet_rows[sheet_name].append([
            f.get('FileName', ''), c1, c2, c3, d1, d2, d3
        ])

    def natural_key(value):
        """'セット10' > 'セット2' になるよう、数字部分を抽出した自然ソートキー"""
        s = value or ""
        m = re.search(r'(\d+)', s)
        num = int(m.group(1)) if m else 0
        prefix = re.sub(r'\d+', '', s)
        return (prefix, num)

    def sort_key(row):
        # row = [file_name, c1, c2, c3, d1, d2, d3]
        return (natural_key(row[1]), natural_key(row[2]), natural_key(row[3]), row[0] or "")

    for _, sheet_name in FOLDER_MANAGEMENT_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.append(["No.", "ファイル名", "カテゴリ１", "カテゴリ２", "カテゴリ３",
                   "ディレクトリ１", "ディレクトリ２", "ディレクトリ３"])
        sorted_rows = sorted(sheet_rows[sheet_name], key=sort_key)
        for i, row in enumerate(sorted_rows, start=1):
            ws.append([i] + row)

    # スタイル（Meiryo UI、ヘッダ太字・塗り、列幅オート）
    header_fill = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(name="Meiryo UI", size=10, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Meiryo UI", size=10)
        # 列幅オート
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws[get_column_letter(col_idx)]:
                v = cell.value
                if v is None:
                    continue
                length = sum(2 if ord(ch) > 255 else 1 for ch in str(v))
                if length > max_len:
                    max_len = length
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def check_password():
    """パスワード認証"""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    password_input = st.text_input("パスワードを入力してください", type="password")

    if password_input:
        if password_input == APP_PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが正しくありません")

    return False


# パスワード認証
if not check_password():
    st.stop()


def get_auth_header():
    """ESA認証ヘッダーを生成"""
    credentials = f"{SERVICE_SECRET}:{LICENSE_KEY}"
    encoded = base64.b64encode(credentials.encode()).decode()
    return {"Authorization": f"ESA {encoded}"}


def safe_int(value, default=0):
    """安全にintに変換"""
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default


def style_excel(ws, num_columns=4, url_column=None):
    """Excelワークシートにスタイルを適用"""
    styles, utils = get_openpyxl_styles()
    Font = styles['Font']
    Border = styles['Border']
    Side = styles['Side']
    PatternFill = styles['PatternFill']
    Alignment = styles['Alignment']
    get_column_letter = utils['get_column_letter']

    # フォント設定
    meiryo_font = Font(name='Meiryo UI')
    header_font = Font(name='Meiryo UI', bold=True, color='FFFFFF')
    # 罫線設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # ヘッダー背景色（濃い青）
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # 全セルにフォントと罫線を適用
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=num_columns):
        for cell in row:
            cell.font = meiryo_font
            cell.border = thin_border

    # ヘッダー行のスタイル（1行目）
    for cell in ws[1]:
        if cell.column <= num_columns:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列幅を自動調整
    for col_idx in range(1, num_columns + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        # URL列は固定幅、それ以外は自動調整
        if url_column and col_idx == url_column:
            ws.column_dimensions[column_letter].width = 70
        else:
            ws.column_dimensions[column_letter].width = min(max_length * 1.5 + 2, 40)


def merge_csv_data(is_df, cl_df):
    """IS検索とCL検索の結果をマージ"""
    # comic_list.csvから辞書を作成（N列=CNO, S列=出版社, Y列=シリーズ）
    cl_dict = {}
    for i in range(1, len(cl_df)):
        try:
            cno = str(cl_df.iloc[i, 13]).strip() if len(cl_df.columns) > 13 else ''  # N列
            publisher = str(cl_df.iloc[i, 18]).strip() if len(cl_df.columns) > 18 else ''  # S列
            series = str(cl_df.iloc[i, 24]).strip() if len(cl_df.columns) > 24 else ''  # Y列

            if cno and cno != 'nan':
                cl_dict[cno] = {
                    'publisher': publisher if publisher != 'nan' else '',
                    'series': series if series != 'nan' else ''
                }
        except Exception:
            continue

    # is_list.csvの出版社とシリーズを置換
    for i in range(1, len(is_df)):
        try:
            cno = str(is_df.iloc[i, 6]).strip() if len(is_df.columns) > 6 else ''  # G列（コミックNo）
            if cno in cl_dict:
                if cl_dict[cno]['publisher'] and len(is_df.columns) > 11:
                    is_df.iloc[i, 11] = cl_dict[cno]['publisher']  # L列
                if cl_dict[cno]['series'] and len(is_df.columns) > 13:
                    is_df.iloc[i, 13] = cl_dict[cno]['series']  # N列
        except Exception:
            continue

    return is_df


def normalize_jan_code(value):
    """JANコードを正規化（数値の.0除去、nan除去）"""
    if pd.isna(value):
        return ''
    jan_str = str(value).strip()
    # '.0' を除去（pandasで数値として読み込まれた場合）
    if jan_str.endswith('.0'):
        jan_str = jan_str[:-2]
    # 'nan' は空文字に
    if jan_str.lower() == 'nan':
        return ''
    return jan_str


def extract_first_volumes(merged_df):
    """1巻のみを抽出して整形"""
    first_vol_dict = {}
    latest_vol_dict = {}
    comic_info_dict = {}  # comic_noごとの情報を保持

    # パス1: 全行を処理して first_vol_dict と latest_vol_dict を構築
    for i in range(1, len(merged_df)):
        try:
            comic_no = normalize_jan_code(merged_df.iloc[i, 6]) if len(merged_df.columns) > 6 else ''  # G列
            if not comic_no:
                continue

            # JAN情報（正規化）
            jan_code = normalize_jan_code(merged_df.iloc[i, 5]) if len(merged_df.columns) > 5 else ''  # F列
            if jan_code:
                latest_vol_dict[comic_no] = jan_code

            # 1巻チェック（J列）
            volume = str(merged_df.iloc[i, 9]).strip() if len(merged_df.columns) > 9 else ''
            if volume == '1' or volume == '1.0':
                if comic_no not in first_vol_dict and jan_code:
                    first_vol_dict[comic_no] = jan_code

            # comic_noの最初の出現行の情報を保持
            if comic_no not in comic_info_dict:
                comic_info_dict[comic_no] = {
                    'kaikatsu_narabi': str(merged_df.iloc[i, 3]).strip() if len(merged_df.columns) > 3 else '',
                    'first_isbn': str(merged_df.iloc[i, 4]).strip() if len(merged_df.columns) > 4 else '',
                    'comic_no': comic_no,
                    'genre': str(merged_df.iloc[i, 7]).strip() if len(merged_df.columns) > 7 else '',
                    'title': str(merged_df.iloc[i, 8]).strip() if len(merged_df.columns) > 8 else '',
                    'publisher': str(merged_df.iloc[i, 11]).strip() if len(merged_df.columns) > 11 else '',
                    'author': str(merged_df.iloc[i, 12]).strip() if len(merged_df.columns) > 12 else '',
                    'series': str(merged_df.iloc[i, 13]).strip() if len(merged_df.columns) > 13 else '',
                }
        except Exception:
            continue

    # パス2: result_dataを構築（全行処理後にfirst_janを設定）
    result_data = []
    for comic_no, info in comic_info_dict.items():
        # 1巻のJAN > 最新巻のJAN > 空 の優先順位
        first_jan = first_vol_dict.get(comic_no, latest_vol_dict.get(comic_no, ''))
        info['first_jan'] = first_jan
        result_data.append(info)

    # 快活並びでソート
    result_data.sort(key=lambda x: int(float(x['kaikatsu_narabi'])) if x['kaikatsu_narabi'] and x['kaikatsu_narabi'] != 'nan' else 999999)
    return result_data


def add_folder_hierarchy_info(result_data, hierarchy_df):
    """フォルダ階層情報を付与"""
    hierarchy_list = []
    for i in range(1, len(hierarchy_df)):
        try:
            row = hierarchy_df.iloc[i]
            hierarchy_list.append({
                'genre': str(row[0]).strip() if pd.notna(row[0]) else '',
                'publisher': str(row[1]).strip() if pd.notna(row[1]) else '',
                'series': str(row[2]).strip() if len(row) > 2 and pd.notna(row[2]) else '',
                'main_folder': str(row[3]).strip() if len(row) > 3 and pd.notna(row[3]) else '',
                'sub_folder': str(row[4]).strip() if len(row) > 4 and pd.notna(row[4]) else ''
            })
        except Exception:
            continue

    for data in result_data:
        matched = False
        for h in hierarchy_list:
            if data['genre'] == h['genre'] and data['publisher'] == h['publisher']:
                if data['series'] and h['series']:
                    if data['series'] == h['series']:
                        data['main_folder'] = h['main_folder']
                        data['sub_folder'] = h['sub_folder']
                        matched = True
                        break
                elif not h['series']:
                    data['main_folder'] = h['main_folder']
                    data['sub_folder'] = h['sub_folder']
                    matched = True
                    break
        if not matched:
            data['main_folder'] = ''
            data['sub_folder'] = ''

    return result_data


def get_bookoff_image(jan_code, session):
    """ブックオフから画像URL取得"""
    url = f"https://shopping.bookoff.co.jp/search/keyword/{jan_code}"
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    NO_IMAGE_PATTERNS = ['item_ll', 'no_image', 'noimage', 'no-image', 'dummy', 'blank', 'spacer', 'placeholder']
    BeautifulSoup = get_bs4()

    try:
        response = session.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')
        img_tag = soup.select_one('.productItem__image img, .js-gridImg')

        if img_tag and img_tag.get('src'):
            image_url = img_tag['src']
            if any(no_img in image_url.lower() for no_img in NO_IMAGE_PATTERNS):
                return None
            return image_url
        return None
    except Exception:
        return None


# Amazon用 User-Agent プール（ローテーションでBot判定回避）
AMAZON_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
]

AMAZON_BOT_MARKERS = [
    'api-services-support@amazon.com',
    'to discuss automated access to amazon',
    "sorry, we just need to make sure you're not a robot",
    'type the characters you see in this image',
    'enter the characters you see below',
    '/errors/validatecaptcha',
    'captcha',
]


def _amazon_headers(referer='https://www.amazon.co.jp/'):
    """Amazon用のブラウザ風ヘッダを生成（User-Agentはランダム選択）"""
    random = get_random()
    headers = {
        'User-Agent': random.choice(AMAZON_USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-User': '?1',
    }
    if referer:
        headers['Referer'] = referer
        headers['Sec-Fetch-Site'] = 'same-origin'
    else:
        headers['Sec-Fetch-Site'] = 'none'
    return headers


def _is_amazon_bot_page(response):
    """AmazonのBot検出ページ（CAPTCHA等）かどうか判定"""
    if response is None:
        return False
    if response.status_code in (503, 429):
        return True
    try:
        text_lower = (response.text or '')[:8000].lower()
    except Exception:
        return False
    return any(marker in text_lower for marker in AMAZON_BOT_MARKERS)


def _warmup_amazon_session(session):
    """初回アクセス時にAmazonトップページを訪問してCookieを獲得"""
    if getattr(session, '_amazon_warmed_up', False):
        return
    try:
        session.get(
            'https://www.amazon.co.jp/',
            headers=_amazon_headers(referer=None),
            timeout=15,
        )
    except Exception:
        pass
    session._amazon_warmed_up = True


def get_amazon_image(jan_code, session):
    """Amazonから画像URL取得（UAローテーション・Cookieウォームアップ・CAPTCHA検出・1回リトライ）"""
    import re
    search_url = f"https://www.amazon.co.jp/s?k={jan_code}&i=stripbooks"
    random = get_random()

    _warmup_amazon_session(session)

    SELECTORS = [
        '.s-image',
        'img[data-image-latency]',
        '.s-product-image img',
        '[data-component-type="s-product-image"] img',
        '.s-result-item img[src*="images-na"]',
        '.s-result-item img[src*="m.media-amazon"]',
    ]
    BeautifulSoup = get_bs4()

    for attempt in range(2):  # 初回 + リトライ1回
        try:
            response = session.get(search_url, headers=_amazon_headers(), timeout=15)

            # Bot判定ページ検出 → リトライ or 諦め
            if _is_amazon_bot_page(response):
                if attempt == 0:
                    time.sleep(random.uniform(4.0, 7.0))
                    continue
                return None

            if response.status_code != 200:
                return None

            soup = BeautifulSoup(response.content, 'html.parser')

            # 複数のセレクタを順番に試す
            for selector in SELECTORS:
                img_tags = soup.select(selector)
                for img_tag in img_tags:
                    src = img_tag.get('src') or img_tag.get('data-src')
                    if src and ('images-na' in src or 'm.media-amazon' in src or 'images-amazon' in src):
                        if 'no-img' not in src.lower() and 'no_image' not in src.lower():
                            if '_AC_' in src:
                                src = src.split('._AC_')[0] + '._SY466_.jpg'
                            elif '_SX' in src or '_SY' in src:
                                src = re.sub(r'\._S[XY]\d+_', '._SY466_', src)
                            return src

            # フォールバック: 正規表現でAmazon画像URLを探す
            amazon_img_pattern = r'(https?://[^"\']+(?:images-na\.ssl-images-amazon|m\.media-amazon|images-amazon)[^"\'\s]+\.(?:jpg|jpeg|png))'
            matches = re.findall(amazon_img_pattern, response.text)
            for match in matches:
                if 'no-img' not in match.lower() and 'no_image' not in match.lower() and 'sprite' not in match.lower():
                    if '_AC_' in match:
                        match = match.split('._AC_')[0] + '._SY466_.jpg'
                    return match

            return None
        except Exception:
            if attempt == 0:
                time.sleep(random.uniform(2.0, 4.0))
                continue
            return None

    return None


def get_rakuten_image(jan_code, session):
    """楽天ブックスから画像URL取得（Amazonのフォールバック）"""
    search_url = f"https://books.rakuten.co.jp/search?g=001&isbn={jan_code}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    }
    BeautifulSoup = get_bs4()

    try:
        response = session.get(search_url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')

        # 楽天ブックスの画像セレクタ
        selectors = [
            '.rbcomp__item-list__item__image img',
            '.item-image img',
            'img[src*="thumbnail.image.rakuten"]',
        ]

        for selector in selectors:
            img_tag = soup.select_one(selector)
            if img_tag:
                src = img_tag.get('src') or img_tag.get('data-src')
                if src and 'noimage' not in src.lower():
                    # 大きいサイズに変換
                    src = src.replace('_ex=64x64', '_ex=200x200').replace('_ex=100x100', '_ex=200x200')
                    return src

        return None
    except Exception:
        return None


def get_image_with_gemini_ai(jan_code, session, source_name="amazon"):
    """Gemini AIを使って画像URLを抽出（セルフヒーリング機能）"""
    # Geminiモデルを遅延読み込み
    model = get_gemini_model()
    if not model:
        return None

    # ソース別のURL設定
    if source_name == "amazon":
        search_url = f"https://www.amazon.co.jp/s?k={jan_code}&i=stripbooks"
    elif source_name == "rakuten":
        search_url = f"https://books.rakuten.co.jp/search?g=001&isbn={jan_code}"
    elif source_name == "bookoff":
        search_url = f"https://shopping.bookoff.co.jp/search/keyword/{jan_code}"
    else:
        return None

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }
    BeautifulSoup = get_bs4()

    try:
        response = session.get(search_url, headers=headers, timeout=15)
        if response.status_code != 200:
            return None

        # HTMLの重要部分だけを抽出（トークン節約）
        soup = BeautifulSoup(response.content, 'html.parser')

        # スクリプトとスタイルを削除
        for tag in soup(['script', 'style', 'noscript', 'header', 'footer', 'nav']):
            tag.decompose()

        # 商品画像が含まれそうな部分を抽出
        main_content = soup.find('main') or soup.find('div', {'id': 'search'}) or soup.find('body')
        if main_content:
            html_snippet = str(main_content)[:8000]  # 最大8000文字に制限
        else:
            html_snippet = str(soup)[:8000]

        prompt = f"""以下のHTMLから、JANコード「{jan_code}」の本の表紙画像URLを1つだけ抽出してください。

条件:
- 画像URLのみを返してください（説明不要）
- NO IMAGE、noimage、placeholder等のダミー画像は除外
- https://で始まる完全なURLで返してください
- 見つからない場合は「NOT_FOUND」とだけ返してください

HTML:
{html_snippet}"""

        response = model.generate_content(prompt)
        result = response.text.strip()

        # 結果を検証
        if result and result != "NOT_FOUND" and result.startswith("http"):
            # NO IMAGE系を最終チェック
            no_image_patterns = ['no_image', 'noimage', 'no-image', 'dummy', 'blank', 'spacer', 'placeholder']
            if not any(p in result.lower() for p in no_image_patterns):
                return result

        return None

    except Exception as e:
        # エラーログ（デバッグ用）
        print(f"Gemini AI error: {e}")
        return None


def download_image(image_url, session):
    """画像をダウンロードしてバイトデータを返す（NO IMAGE検出付き）"""
    try:
        response = session.get(image_url, timeout=10)
        response.raise_for_status()
        content = response.content

        # 画像サイズが小さすぎる場合はNO IMAGEの可能性が高い（5KB未満）
        if len(content) < 5000:
            return None

        # 特定のパターンをURLで再チェック
        no_image_patterns = ['no_image', 'noimage', 'no-image', 'dummy', 'blank', 'spacer', 'placeholder']
        if any(pattern in image_url.lower() for pattern in no_image_patterns):
            return None

        return content
    except Exception:
        return None


def resize_to_square(image_data: bytes, size: int = 600, center: bool = False):
    """画像を正方形にリサイズ（アスペクト比維持、白背景パディング、表紙を大きく表示）

    - center=False: バッジ領域を想定して左寄せ（セット・予約用）
    - center=True: バッジなしで中央配置（単品用）
    """
    Image = get_pil()
    img = Image.open(BytesIO(image_data)).convert("RGB")

    if center:
        # バッジなし：中央配置、表示領域をキャンバスの93%に（画像の粗さ対策で少し余白を取る）
        display_w = int(size * 0.93)
        display_h = int(size * 0.93)
    else:
        # バッジ領域を考慮した表示領域（右側にバッジがあるため左寄せ）
        display_w = int(size * 0.75)  # 横は75%まで使用
        display_h = int(size * 0.90)  # 縦は90%まで使用

    # アスペクト比を維持して表示領域にフィット（拡大も許可）
    ratio = min(display_w / img.width, display_h / img.height)
    new_w = int(img.width * ratio)
    new_h = int(img.height * ratio)
    img = img.resize((new_w, new_h), Image.LANCZOS)

    # 白背景の正方形キャンバス
    canvas = Image.new("RGB", (size, size), (255, 255, 255))
    if center:
        # 中央配置
        x = (size - new_w) // 2
        y = (size - new_h) // 2
    else:
        # 左寄せ（少し余白）・上下中央
        x = (display_w - new_w) // 2 + int(size * 0.02)
        y = (size - new_h) // 2
    canvas.paste(img, (x, y))

    return canvas


def add_shipping_badge(base_image, badge_path: str):
    """送料無料バッジを合成（白背景を透明化してオーバーレイ）"""
    Image = get_pil()
    import numpy as np

    base = base_image.convert("RGBA")
    badge = Image.open(badge_path).convert("RGBA")

    # バッジを元画像サイズにリサイズ
    badge = badge.resize(base.size, Image.LANCZOS)

    # 白背景を透明化（RGB各チャンネルが240以上を透明に）
    badge_data = np.array(badge)
    white_mask = (badge_data[:, :, 0] > 240) & (badge_data[:, :, 1] > 240) & (badge_data[:, :, 2] > 240)
    badge_data[white_mask, 3] = 0
    badge = Image.fromarray(badge_data, "RGBA")

    # 合成
    result = Image.alpha_composite(base, badge)
    return result.convert("RGB")


def image_to_bytes(image, quality: int = 95) -> bytes:
    """PIL ImageをJPEGバイトデータに変換"""
    buf = BytesIO()
    image.save(buf, format="JPEG", quality=quality)
    return buf.getvalue()


def _workflow_prepare_target_data(missing_comics: list, is_list_content: str, comic_list_content: str, missing_yoyaku: list = None):
    """画像取得対象(target_data)を構築する（CSV読み込み・マージ・単品/予約JAN引き当て）

    - セット: comic_no（'_'なし）→ is_list.csvの1巻JAN（extract_first_volumes）
    - 単品: comic_no（'_'あり：base_no_vol）→ is_list.csvの該当巻JAN
    - 予約: comic_no（'_'なし）→ is_list.csvの最新巻（最大vol）JAN
    """
    if missing_yoyaku is None:
        missing_yoyaku = []

    try:
        is_df = pd.read_csv(BytesIO(is_list_content.encode('utf-8')), header=None)
    except:
        is_df = pd.read_csv(BytesIO(is_list_content.encode('cp932')), header=None)

    try:
        cl_df = pd.read_csv(BytesIO(comic_list_content.encode('utf-8')), header=None)
    except:
        cl_df = pd.read_csv(BytesIO(comic_list_content.encode('cp932')), header=None)

    merged_df = merge_csv_data(is_df, cl_df)
    result_data = extract_first_volumes(merged_df)

    # is_list からのJAN引き当て辞書を構築
    is_jan_lookup = {}                 # (comic_no, vol_str) → jan
    latest_vol_lookup = {}             # comic_no → (max_vol_int, jan)
    for i in range(1, len(is_df)):
        try:
            cno = str(is_df.iloc[i, 6]).strip() if pd.notna(is_df.iloc[i, 6]) else ''
            cno = cno.replace('.0', '')
            vol_s = str(is_df.iloc[i, 9]).strip() if pd.notna(is_df.iloc[i, 9]) else ''
            vol_s = vol_s.replace('.0', '')
            jan = normalize_jan_code(is_df.iloc[i, 5])
            if not cno or not jan:
                continue
            is_jan_lookup[(cno, vol_s)] = jan
            try:
                vol_n = int(vol_s)
                prev = latest_vol_lookup.get(cno)
                if prev is None or vol_n > prev[0]:
                    latest_vol_lookup[cno] = (vol_n, jan)
            except:
                pass
        except:
            continue

    result_data_dict = {str(d.get('comic_no', '')).strip(): d for d in result_data}

    # 予約 comic_no のセット（'_'なしのセット品と区別するため除外判定に使う）
    yoyaku_set = set(normalize_jan_code(c) for c in missing_yoyaku if normalize_jan_code(c))

    # セット（'_'なし かつ 予約セットに含まれない）
    missing_norm = set(normalize_jan_code(c) for c in missing_comics if normalize_jan_code(c))
    missing_set_only = set(c for c in missing_norm if '_' not in c and c not in yoyaku_set)
    target_data = []
    for d in result_data:
        cno = str(d.get('comic_no', '')).strip()
        if cno in missing_set_only:
            entry = dict(d)
            entry['type'] = 'set'
            entry['is_tanpin'] = False
            target_data.append(entry)

    # 単品（'_'あり）
    tanpin_comics = [c for c in missing_comics if '_' in str(c)]
    for tc in tanpin_comics:
        parts = str(tc).split('_')
        base_no = parts[0]
        vol_num = int(parts[1]) if len(parts) > 1 else 1
        jan_code = is_jan_lookup.get((base_no, str(vol_num)), '')
        base_info = result_data_dict.get(base_no, {})
        if jan_code:
            target_data.append({
                'comic_no': tc,
                'first_jan': jan_code,
                'type': 'tanpin',
                'is_tanpin': True,
                'genre': base_info.get('genre', ''),
                'publisher': base_info.get('publisher', ''),
                'series': base_info.get('series', ''),
                'title': base_info.get('title', ''),
            })

    # 予約（'_'なし、最新巻のJANを引き当て）
    yoyaku_unresolved = []  # is_list.csvに該当なしでJAN引き当て失敗した予約コミックNo
    for yc in missing_yoyaku:
        cno = normalize_jan_code(yc)
        if not cno:
            continue
        latest = latest_vol_lookup.get(cno)
        jan_code = latest[1] if latest else ''
        base_info = result_data_dict.get(cno, {})
        if jan_code:
            target_data.append({
                'comic_no': cno,
                'first_jan': jan_code,
                'type': 'yoyaku',
                'is_tanpin': False,
                'genre': base_info.get('genre', ''),
                'publisher': base_info.get('publisher', ''),
                'series': base_info.get('series', ''),
                'title': base_info.get('title', ''),
            })
        else:
            yoyaku_unresolved.append(cno)

    if yoyaku_unresolved:
        try:
            st.warning(
                "予約コミックNoがis_list.csvに存在せずJANを引き当てできませんでした: "
                + ", ".join(yoyaku_unresolved)
                + "\n（Step ② のJAN取得ワークフローを実行してis_list.csvを更新してください）"
            )
        except Exception:
            pass

    # 型未設定の保険
    for d in target_data:
        if 'type' not in d:
            d['type'] = 'tanpin' if d.get('is_tanpin') or '_' in str(d.get('comic_no', '')) else 'set'
        if 'is_tanpin' not in d:
            d['is_tanpin'] = (d.get('type') == 'tanpin')

    return target_data


def _workflow_process_one_image(data: dict, session, badge_path: str):
    """1件分の画像取得＋加工。結果dict(success/comic_no/jan_code/log/source/image)を返す"""
    import os
    random = get_random()
    comic_no = str(data.get('comic_no', '')).strip()
    jan_code = normalize_jan_code(data.get('first_jan', ''))

    if not jan_code:
        return {'success': False, 'comic_no': comic_no, 'jan_code': '',
                'log': f"⚠️ {comic_no}: JANコードなし - スキップ", 'source': None, 'image': None}

    image_url = get_bookoff_image(jan_code, session)
    source = 'bookoff'

    if not image_url:
        time.sleep(random.uniform(1.5, 3.0))
        image_url = get_amazon_image(jan_code, session)
        source = 'amazon'

    if not image_url:
        time.sleep(random.uniform(0.3, 0.6))
        image_url = get_rakuten_image(jan_code, session)
        source = 'rakuten'

    if not image_url and GEMINI_API_KEY:
        time.sleep(random.uniform(0.5, 1.0))
        ai_result = get_image_with_gemini_ai(jan_code, session, "amazon")
        if ai_result:
            image_url = ai_result
            source = 'gemini_ai'

    if not image_url:
        return {'success': False, 'comic_no': comic_no, 'jan_code': jan_code,
                'log': f"❌ {comic_no} (JAN: {jan_code}): 画像が見つかりません", 'source': None, 'image': None}

    image_data = download_image(image_url, session)
    if not image_data:
        return {'success': False, 'comic_no': comic_no, 'jan_code': jan_code,
                'log': f"❌ {comic_no}: ダウンロード失敗 ({source})", 'source': source, 'image': None}

    ctype = data.get('type') or ('tanpin' if data.get('is_tanpin') else 'set')
    is_tanpin = (ctype == 'tanpin')
    need_badge = ctype in ('set', 'yoyaku')

    if is_tanpin:
        # 単品: 600×600px・中央配置（バッジなし）
        resized = resize_to_square(image_data, 600, center=True)
        final_image = resized
        badge_status = "リサイズのみ（中央）"
    else:
        # セット・予約: 600×600px・左寄せ＋送料無料バッジ
        resized = resize_to_square(image_data, 600)
        if os.path.exists(badge_path):
            final_image = add_shipping_badge(resized, badge_path)
            badge_status = "バッジ付き"
        else:
            final_image = resized
            badge_status = "バッジ画像なし"
    final_bytes = image_to_bytes(final_image)

    image_dict = {
        'comic_no': comic_no,
        'jan_code': jan_code,
        'image_data': final_bytes,
        'source': source,
        'type': ctype,
        'is_tanpin': is_tanpin,
        'badge': need_badge,
        'genre': data.get('genre', ''),
        'publisher': data.get('publisher', ''),
        'series': data.get('series', ''),
        'title': data.get('title', ''),
    }
    return {'success': True, 'comic_no': comic_no, 'jan_code': jan_code,
            'log': f"✅ {comic_no} (JAN: {jan_code}): {source} - {badge_status}",
            'source': source, 'image': image_dict}


def process_workflow_images(missing_comics: list, is_list_content: str, comic_list_content: str, badge_path: str, progress_bar=None, status_text=None, log_container=None):
    """ワークフロー用：不足画像を取得してバッジ合成まで行う（一括実行版）"""
    target_data = _workflow_prepare_target_data(missing_comics, is_list_content, comic_list_content)
    if not target_data:
        return {'success': False, 'error': '処理対象がありません', 'images': [], 'stats': {}}

    session = requests.Session()
    random = get_random()
    downloaded_images = []
    stats = {'total': len(target_data), 'success': 0, 'failed': 0, 'bookoff': 0, 'amazon': 0, 'rakuten': 0, 'gemini_ai': 0}
    logs = []

    for i, data in enumerate(target_data):
        if progress_bar:
            progress_bar.progress((i + 1) / len(target_data))
        if status_text:
            status_text.text(f"処理中: {data.get('comic_no', '')} ({i + 1}/{len(target_data)})")

        result = _workflow_process_one_image(data, session, badge_path)
        logs.append(result['log'])
        if result['success']:
            downloaded_images.append(result['image'])
            stats['success'] += 1
            if result['source']:
                stats[result['source']] = stats.get(result['source'], 0) + 1
        else:
            stats['failed'] += 1

        time.sleep(random.uniform(0.3, 0.8))

    return {'success': True, 'images': downloaded_images, 'stats': stats, 'logs': logs}


def prepare_yahoo_zips(images: list, excel_set_df, excel_tanpin_df, additional_dir: str) -> dict:
    """ヤフー用にリネーム＋追加画像＋ZIP分割生成"""
    import os
    zipfile = get_zipfile()
    Image = get_pil()

    MAX_ZIP_SIZE = 25 * 1024 * 1024  # 25MB

    # コミックNo → 商品コード のマッピング構築
    comic_to_product = {}

    # セット品シート: A列=商品コード, D列=コミックNo
    if excel_set_df is not None:
        for i in range(len(excel_set_df)):
            try:
                product_code = str(excel_set_df.iloc[i, 0]).strip()
                comic_no = str(excel_set_df.iloc[i, 3]).strip().replace('.0', '')
                if product_code and comic_no and product_code != 'nan' and comic_no != 'nan':
                    comic_to_product[comic_no] = {'code': product_code, 'type': 'set'}
            except:
                continue

    # 単品シート: A列=商品コード, E列=コミックNo
    if excel_tanpin_df is not None:
        for i in range(len(excel_tanpin_df)):
            try:
                product_code = str(excel_tanpin_df.iloc[i, 0]).strip()
                comic_no = str(excel_tanpin_df.iloc[i, 4]).strip().replace('.0', '')
                if product_code and comic_no and product_code != 'nan' and comic_no != 'nan':
                    comic_to_product[comic_no] = {'code': product_code, 'type': 'tanpin'}
            except:
                continue

    # 追加画像読み込み
    additional_1_path = os.path.join(additional_dir, "additional_1.jpg")
    additional_2_path = os.path.join(additional_dir, "additional_2.jpg")
    additional_1_data = None
    additional_2_data = None
    if os.path.exists(additional_1_path):
        with open(additional_1_path, 'rb') as f:
            additional_1_data = f.read()
    if os.path.exists(additional_2_path):
        with open(additional_2_path, 'rb') as f:
            additional_2_data = f.read()

    # ファイルリスト構築
    file_entries = []  # [(filename, bytes)]
    mapped_count = 0
    unmapped = []
    logs = []

    for img in images:
        comic_no = str(img['comic_no']).strip()
        mapping = comic_to_product.get(comic_no)

        if not mapping:
            unmapped.append(comic_no)
            logs.append(f"⚠️ {comic_no}: 商品コードが見つかりません")
            continue

        product_code = mapping['code']
        is_set = mapping['type'] == 'set'

        # メイン画像
        file_entries.append((f"{product_code}.jpg", img['image_data']))

        # セット品のみ追加画像
        if is_set:
            if additional_1_data:
                file_entries.append((f"{product_code}_1.jpg", additional_1_data))
            if additional_2_data:
                file_entries.append((f"{product_code}_2.jpg", additional_2_data))

        mapped_count += 1
        logs.append(f"✅ {comic_no} → {product_code} {'(セット品+追加画像)' if is_set else '(単品)'}")

    # ZIP分割生成
    zip_buffers = []
    current_files = []
    current_size = 0

    for filename, data in file_entries:
        file_size = len(data)
        if current_size + file_size > MAX_ZIP_SIZE and current_files:
            # 現在のZIPを保存
            buf = BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fn, fd in current_files:
                    zf.writestr(fn, fd)
            zip_buffers.append(buf.getvalue())
            current_files = []
            current_size = 0

        current_files.append((filename, data))
        current_size += file_size

    # 残りを保存
    if current_files:
        buf = BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fn, fd in current_files:
                zf.writestr(fn, fd)
        zip_buffers.append(buf.getvalue())

    return {
        'zips': zip_buffers,
        'mapped': mapped_count,
        'unmapped': unmapped,
        'total_files': len(file_entries),
        'logs': logs
    }


def prepare_rakuten_queue(images: list, hierarchy_df, folders: list) -> dict:
    """楽天用にフォルダマッピング＋アップロードキュー生成"""
    # フォルダパス → FolderId の辞書
    folder_map = {}
    for f in folders:
        folder_map[f['FolderName']] = f['FolderId']

    # 階層情報パース
    hierarchy_list = []
    if hierarchy_df is not None:
        for i in range(1, len(hierarchy_df)):
            try:
                row = hierarchy_df.iloc[i]
                hierarchy_list.append({
                    'genre': str(row[0]).strip() if pd.notna(row[0]) else '',
                    'publisher': str(row[1]).strip() if pd.notna(row[1]) else '',
                    'series': str(row[2]).strip() if len(row) > 2 and pd.notna(row[2]) else '',
                    'main_folder': str(row[3]).strip() if len(row) > 3 and pd.notna(row[3]) else '',
                    'sub_folder': str(row[4]).strip() if len(row) > 4 and pd.notna(row[4]) else ''
                })
            except:
                continue

    # 画像ごとにフォルダを決定
    queue = []
    unmapped = []
    logs = []

    # is_listのresult_dataが必要だが、workflow_dataから取得する形
    # ここではcomic_noベースでシンプルにマッチング
    for img in images:
        comic_no = str(img['comic_no']).strip()
        file_name = f"{comic_no}.jpg"

        # sub_folderをフォルダ名として使用
        matched_folder = None
        matched_main_folder = None
        matched_folder_id = None

        # result_dataにgenre/publisher情報がある場合はマッチング
        genre = img.get('genre', '')
        publisher = img.get('publisher', '')
        series = img.get('series', '')

        # 画像データに既にmain_folder/sub_folderが付与済みの場合はそちらを優先
        if img.get('main_folder') and img.get('sub_folder'):
            matched_main_folder = img['main_folder']
            matched_folder = img['sub_folder'] or img['main_folder']
        else:
            for h in hierarchy_list:
                if genre == h['genre'] and publisher == h['publisher']:
                    if series and h['series'] and series == h['series']:
                        matched_folder = h['sub_folder'] or h['main_folder']
                        matched_main_folder = h['main_folder']
                        break
                    elif not h['series']:
                        matched_folder = h['sub_folder'] or h['main_folder']
                        matched_main_folder = h['main_folder']
                        break

        if matched_folder and matched_folder in folder_map:
            matched_folder_id = folder_map[matched_folder]
            queue.append({
                'file_bytes': img['image_data'],
                'folder_id': matched_folder_id,
                'folder_name': matched_folder,
                'main_folder': matched_main_folder,
                'file_name': file_name,
                'comic_no': comic_no
            })
            logs.append(f"✅ {comic_no} → {matched_main_folder}/{matched_folder} (ID: {matched_folder_id})")
        else:
            unmapped.append(comic_no)
            if matched_folder:
                logs.append(f"⚠️ {comic_no}: フォルダ '{matched_folder}' がR-Cabinetに見つかりません")
            else:
                logs.append(f"⚠️ {comic_no}: フォルダ階層にマッチしません")

    return {
        'queue': queue,
        'mapped': len(queue),
        'unmapped': unmapped,
        'logs': logs
    }


@st.cache_data(ttl=600, show_spinner=False)
def get_all_folders():
    """R-Cabinetの全フォルダ一覧を取得"""
    url = f"{BASE_URL}/cabinet/folders/get"
    headers = get_auth_header()

    all_folders = []
    offset = 1  # 1始まり（ページ番号）
    limit = 100  # APIの上限は100件

    while True:
        params = {"offset": offset, "limit": limit}

        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)
        except requests.exceptions.RequestException as e:
            return None, f"接続エラー: {str(e)}"

        if response.status_code != 200:
            return None, f"エラー: {response.status_code} - {response.text[:200]}"

        try:
            root = ET.fromstring(response.text)
        except ET.ParseError as e:
            return None, f"XMLパースエラー: {str(e)}"

        # エラーチェック
        system_status = root.findtext('.//systemStatus', '')
        if system_status != 'OK':
            message = root.findtext('.//message', 'Unknown error')
            return None, f"APIエラー: {message}"

        folders = root.findall('.//folder')

        for folder in folders:
            all_folders.append({
                'FolderId': folder.findtext('FolderId', ''),
                'FolderName': folder.findtext('FolderName', ''),
                'FolderPath': folder.findtext('FolderPath', ''),
                'FileCount': safe_int(folder.findtext('FileCount', '0')),
            })

        # 取得件数がlimit未満なら終了（最終ページ）
        if len(folders) < limit:
            break
        offset += 1  # 次のページへ
        time.sleep(0.3)

    return all_folders, None


def create_folder(folder_name, directory_name=None, upper_folder_id=None):
    """R-Cabinetにフォルダを1件作成（cabinet.folder.insert）"""
    url = f"{BASE_URL}/cabinet/folder/insert"
    headers = get_auth_header()
    headers["Content-Type"] = "text/xml;charset=UTF-8"

    # XMLリクエストボディを構築
    folder_elements = f"<folderName>{folder_name}</folderName>"
    if directory_name:
        folder_elements += f"<directoryName>{directory_name}</directoryName>"
    if upper_folder_id:
        folder_elements += f"<upperFolderId>{upper_folder_id}</upperFolderId>"

    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<request><folderInsertRequest><folder>"
        f"{folder_elements}"
        "</folder></folderInsertRequest></request>"
    )

    try:
        response = requests.post(url, headers=headers, data=xml_body.encode('utf-8'), timeout=30)
    except requests.exceptions.RequestException as e:
        return {"success": False, "error": f"接続エラー: {str(e)}"}

    if response.status_code != 200:
        return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:500]}"}

    try:
        root = ET.fromstring(response.text)
    except ET.ParseError as e:
        return {"success": False, "error": f"XMLパースエラー: {str(e)}"}

    system_status = root.findtext('.//systemStatus', '')
    if system_status != 'OK':
        message = root.findtext('.//message', 'Unknown error')
        return {"success": False, "error": f"APIエラー: {message}"}

    folder_id = root.findtext('.//FolderId', '')
    return {"success": True, "folder_id": folder_id}


def upload_image(file_data, file_name, folder_id, file_path_name=None, overwrite=False):
    """R-Cabinetに画像を1枚アップロード（cabinet.file.insert）"""
    url = f"{BASE_URL}/cabinet/file/insert"
    headers = get_auth_header()
    # Content-Typeはrequestsが自動設定（multipart/form-data + boundary）

    # XMLパラメータ部分を構築
    file_elements = f"<fileName>{file_name}</fileName>"
    file_elements += f"<folderId>{folder_id}</folderId>"
    if file_path_name:
        file_elements += f"<filePath>{file_path_name}</filePath>"
    if overwrite:
        file_elements += "<overWrite>true</overWrite>"

    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<request><fileInsertRequest><file>"
        f"{file_elements}"
        "</file></fileInsertRequest></request>"
    )

    try:
        response = requests.post(
            url,
            headers=headers,
            files=[
                ("xml", (None, xml_body.encode("utf-8"), "text/xml;charset=UTF-8")),
                ("file", (file_name, file_data, "application/octet-stream")),
            ],
            timeout=60,
        )
    except requests.exceptions.RequestException as e:
        return {"success": False, "error": f"接続エラー: {str(e)}"}

    if response.status_code != 200:
        return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:500]}"}

    try:
        root = ET.fromstring(response.text)
    except ET.ParseError as e:
        return {"success": False, "error": f"XMLパースエラー: {str(e)}"}

    system_status = root.findtext('.//systemStatus', '')
    if system_status != 'OK':
        message = root.findtext('.//message', 'Unknown error')
        return {"success": False, "error": f"APIエラー: {message}"}

    file_url = root.findtext('.//FileUrl', '')
    file_id = root.findtext('.//FileId', '')
    return {"success": True, "file_url": file_url, "file_id": file_id}


@st.cache_data(ttl=300, show_spinner=False)
def get_folder_files(folder_id: int, max_retries: int = 3):
    """指定フォルダ内の画像一覧を取得（リトライ機能付き）"""
    url = f"{BASE_URL}/cabinet/folder/files/get"
    headers = get_auth_header()

    all_files = []
    offset = 1  # 1始まり（ページ番号）
    limit = 100  # APIの上限は100件

    while True:
        params = {"folderId": folder_id, "offset": offset, "limit": limit}

        # リトライ処理
        for retry in range(max_retries):
            try:
                response = requests.get(url, headers=headers, params=params, timeout=30)
            except requests.exceptions.RequestException as e:
                if retry < max_retries - 1:
                    time.sleep(2)  # 2秒待ってリトライ
                    continue
                return None, f"接続エラー: {str(e)}"

            if response.status_code == 200:
                break  # 成功
            elif response.status_code == 403 and retry < max_retries - 1:
                time.sleep(3)  # 403の場合は3秒待ってリトライ
                continue
            else:
                if retry == max_retries - 1:
                    return None, f"エラー: {response.status_code}"

        try:
            root = ET.fromstring(response.text)
        except ET.ParseError as e:
            return None, f"XMLパースエラー: {str(e)}"

        system_status = root.findtext('.//systemStatus', '')
        if system_status != 'OK':
            message = root.findtext('.//message', 'Unknown error')
            return None, f"APIエラー: {message}"

        files = root.findall('.//file')

        for f in files:
            all_files.append({
                'FileId': f.findtext('FileId', ''),
                'FileName': f.findtext('FileName', ''),
                'FileUrl': f.findtext('FileUrl', ''),
                'FilePath': f.findtext('FilePath', ''),
                'FileSize': f.findtext('FileSize', ''),
                'TimeStamp': f.findtext('TimeStamp', ''),
            })

        # 取得件数がlimit未満なら終了（最終ページ）
        if len(files) < limit:
            break
        offset += 1  # 次のページへ
        time.sleep(0.3)

    return all_files, None


def search_image_by_name(file_name: str):
    """画像名で検索"""
    url = f"{BASE_URL}/cabinet/files/search"
    headers = get_auth_header()
    params = {"fileName": file_name}

    response = requests.get(url, headers=headers, params=params)

    if response.status_code == 200:
        root = ET.fromstring(response.text)
        files = root.findall('.//file')

        results = []
        for f in files:
            results.append({
                'FileId': f.findtext('FileId', ''),
                'FileName': f.findtext('FileName', ''),
                'FileUrl': f.findtext('FileUrl', ''),
                'FolderName': f.findtext('FolderName', ''),
                'FolderPath': f.findtext('FolderPath', ''),
            })
        return results
    return []


def is_exact_match(file_name: str, comic_no: str) -> bool:
    """ファイル名がコミックNoと完全一致するかチェック（拡張子除く）"""
    # 拡張子を除去
    name_without_ext = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
    # 完全一致のみ
    return name_without_ext == comic_no


def check_comic_images(comic_numbers: list, progress_bar=None, status_text=None, typed_comics: dict = None):
    """コミックNoリストの画像存在チェック（DB参照版 - 高速）

    検索対象は CHECK_TARGET_FOLDERS の各パスprefix配下のみ
    （例: /comic/comic-set 配下なら、セット本体・セット1・セット2などすべて含む）。
    全く別のフォルダ（例: /other）は対象外。

    Args:
        comic_numbers: フラットなコミックNoリスト（typed_comics 未指定時に使用）。
        typed_comics: {種別ラベル: [コミックNo, ...]} - 指定時は種別ごとに対応パスで検索。
                      種別ラベルは CHECK_TARGET_FOLDERS のキー（"セット品" / "単品" / "予約"）。
    """
    results = []

    # DBから全画像データを取得（1回だけ）
    if status_text:
        status_text.text("DBからデータを読み込み中...")
    if progress_bar:
        progress_bar.progress(0.1)

    all_images, _ = load_images_from_db()

    if not all_images:
        return None

    if progress_bar:
        progress_bar.progress(0.3)

    if status_text:
        status_text.text("検索インデックスを作成中...")

    # 種別別インデックス: {type_label: {comic_no(ext除く): [img, ...]}}
    # folder_path のprefixで種別を判定し、対象外画像は一切インデックス化しない
    def classify(folder_path: str) -> str | None:
        """folder_pathから種別ラベルを判定。対象外なら None。"""
        if not folder_path:
            return None
        for label, prefix in CHECK_TARGET_FOLDERS.items():
            # 完全一致 or prefix配下（サブフォルダ: /comic/comic-set/sub1）
            if folder_path == prefix or folder_path.startswith(prefix + '/'):
                return label
        return None

    index_by_type: dict = {label: {} for label in CHECK_TARGET_FOLDERS.keys()}
    for img in all_images:
        folder_path = (img.get('FolderPath') or '').strip()
        type_label = classify(folder_path)
        if not type_label:
            continue
        file_name = img.get('FileName', '')
        name_without_ext = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
        index_by_type[type_label].setdefault(name_without_ext, []).append(img)

    if progress_bar:
        progress_bar.progress(0.5)

    if status_text:
        status_text.text("チェック中...")

    # 検索タスクを「(種別, コミックNo)」の形に正規化
    tasks = []
    if typed_comics:
        for type_label, comics in typed_comics.items():
            if type_label not in index_by_type:
                continue
            for cno in comics:
                tasks.append((type_label, cno))
    else:
        # 種別指定なし: 全対象種別を横断検索
        for cno in comic_numbers:
            tasks.append((None, cno))

    total = len(tasks) or 1

    for i, (type_label, comic_no) in enumerate(tasks):
        comic_no_str = str(comic_no).strip()
        matched_imgs = []

        if type_label:
            # 指定種別のみを検索
            type_index = index_by_type.get(type_label, {})
            if comic_no_str in type_index:
                matched_imgs = type_index[comic_no_str]
        else:
            # 全対象種別を横断検索
            for label in index_by_type:
                if comic_no_str in index_by_type[label]:
                    matched_imgs.extend(index_by_type[label][comic_no_str])

        if matched_imgs:
            for img in matched_imgs:
                results.append({
                    'コミックNo': comic_no,
                    '種別': type_label or '-',
                    '存在': '✅ あり',
                    'ファイル名': img.get('FileName', ''),
                    'フォルダ': img.get('FolderName', ''),
                    'URL': img.get('FileUrl', ''),
                })
        else:
            results.append({
                'コミックNo': comic_no,
                '種別': type_label or '-',
                '存在': '❌ なし',
                'ファイル名': '-',
                'フォルダ': '-',
                'URL': '-',
            })

        if progress_bar and (i + 1) % 100 == 0:
            progress_bar.progress(0.5 + 0.5 * (i + 1) / total)

    if progress_bar:
        progress_bar.progress(1.0)

    return results


# 認証情報チェック
if not SERVICE_SECRET or not LICENSE_KEY:
    st.error("⚠️ RMS API認証情報が設定されていません。Streamlit Secretsに設定してください。")
    st.stop()


# サイドバー：モード切替
with st.sidebar:
    st.title("🖼️ R-Cabinet")
    st.caption(f"v{APP_VERSION}")

    st.markdown("<br>", unsafe_allow_html=True)

    mode = st.radio(
        "機能を選択",
        ["🎨 クリエイティブスタジオ", "🛰️ R-Cabi構成把握", "🏗️ R-Cabiフォルダ制作", "☁️ コピー：Local⇒R-Cabi", "💾 コピー：R-Cabi⇒Local", "🔁 コピー：R-Cabi⇒R-Cabi"],
        label_visibility="collapsed"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.divider()
    st.markdown("<br>", unsafe_allow_html=True)


# ============================================================
# 画像ワークフロー（統合モード）のカスタムCSS
# ============================================================
WORKFLOW_CSS = """
<style>
/* ステップナビゲーション */
.workflow-nav {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 1.5rem 2rem;
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    border-radius: 16px;
    margin-bottom: 2rem;
    box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
    position: relative;
    overflow: hidden;
}

.workflow-nav::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 3px;
    background: linear-gradient(90deg, #667eea, #764ba2, #f093fb);
}

.step-item {
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 0.5rem;
    flex: 1;
    position: relative;
    z-index: 1;
}

.step-circle {
    width: 48px;
    height: 48px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 1.1rem;
    transition: all 0.3s ease;
}

.step-circle.completed {
    background: linear-gradient(135deg, #00d9a5 0%, #00b894 100%);
    color: white;
    box-shadow: 0 4px 15px rgba(0, 217, 165, 0.4);
}

.step-circle.active {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    box-shadow: 0 4px 20px rgba(102, 126, 234, 0.5);
    transform: scale(1.1);
    animation: pulse 2s infinite;
}

.step-circle.pending {
    background: rgba(255, 255, 255, 0.1);
    color: rgba(255, 255, 255, 0.5);
    border: 2px dashed rgba(255, 255, 255, 0.2);
}

@keyframes pulse {
    0%, 100% { box-shadow: 0 4px 20px rgba(102, 126, 234, 0.5); }
    50% { box-shadow: 0 4px 30px rgba(102, 126, 234, 0.8); }
}

.step-label {
    font-size: 0.75rem;
    color: rgba(255, 255, 255, 0.7);
    text-align: center;
    max-width: 80px;
}

.step-label.active {
    color: white;
    font-weight: 600;
}

.step-connector {
    flex: 0.5;
    height: 3px;
    background: rgba(255, 255, 255, 0.1);
    position: relative;
    margin-top: -24px;
}

.step-connector.completed {
    background: linear-gradient(90deg, #00d9a5, #00b894);
}

/* ステップコンテンツカード */
.step-card {
    background: linear-gradient(145deg, #ffffff 0%, #f8f9fa 100%);
    border-radius: 16px;
    padding: 2rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
    border: 1px solid rgba(0, 0, 0, 0.05);
}

.step-card-header {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin-bottom: 1.5rem;
    padding-bottom: 1rem;
    border-bottom: 2px solid #f0f0f0;
}

.step-card-icon {
    width: 56px;
    height: 56px;
    border-radius: 14px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.5rem;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
}

.step-card-title {
    font-size: 1.4rem;
    font-weight: 700;
    color: #1a1a2e;
    margin: 0;
}

.step-card-desc {
    font-size: 0.9rem;
    color: #666;
    margin: 0;
}

/* アクションボタン */
.action-btn-primary {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    border: none;
    padding: 0.8rem 2rem;
    border-radius: 12px;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.3s ease;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
}

.action-btn-primary:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
}

/* 結果サマリーカード */
.result-card {
    background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
    border-radius: 12px;
    padding: 1.2rem;
    text-align: center;
    border: 1px solid rgba(0, 0, 0, 0.05);
}

.result-card.success {
    background: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%);
    border-color: rgba(40, 167, 69, 0.2);
}

.result-card.warning {
    background: linear-gradient(135deg, #fff3cd 0%, #ffeeba 100%);
    border-color: rgba(255, 193, 7, 0.2);
}

.result-card.error {
    background: linear-gradient(135deg, #f8d7da 0%, #f5c6cb 100%);
    border-color: rgba(220, 53, 69, 0.2);
}

.result-value {
    font-size: 2rem;
    font-weight: 700;
    color: #1a1a2e;
}

.result-label {
    font-size: 0.8rem;
    color: #666;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* プログレスバー */
.progress-container {
    background: rgba(0, 0, 0, 0.05);
    border-radius: 10px;
    height: 12px;
    overflow: hidden;
    margin: 1rem 0;
}

.progress-bar {
    height: 100%;
    border-radius: 10px;
    background: linear-gradient(90deg, #667eea, #764ba2);
    transition: width 0.3s ease;
}

/* ログエリア */
.log-area {
    background: #1a1a2e;
    border-radius: 12px;
    padding: 1rem;
    max-height: 300px;
    overflow-y: auto;
    font-family: 'JetBrains Mono', 'Fira Code', monospace;
    font-size: 0.85rem;
}

.log-entry {
    padding: 0.3rem 0;
    border-bottom: 1px solid rgba(255, 255, 255, 0.05);
}

.log-entry.success { color: #00d9a5; }
.log-entry.error { color: #ff6b6b; }
.log-entry.info { color: #74b9ff; }

</style>
"""


def render_workflow_step_nav(current_step: int, completed_steps: list):
    """ワークフローのステップナビゲーションを描画"""
    steps = [
        ("①", "不足特定"),
        ("②", "JAN取得"),
        ("③", "画像取得"),
        ("④", "準備"),
        ("⑤", "アップロード")
    ]

    html = '<div class="workflow-nav">'

    for i, (num, label) in enumerate(steps):
        step_num = i + 1
        if step_num in completed_steps:
            status = "completed"
            icon = "✓"
        elif step_num == current_step:
            status = "active"
            icon = num
        else:
            status = "pending"
            icon = num

        label_class = "active" if step_num == current_step else ""

        html += f'''
        <div class="step-item">
            <div class="step-circle {status}">{icon}</div>
            <div class="step-label {label_class}">{label}</div>
        </div>
        '''

        # コネクター（最後以外）
        if i < len(steps) - 1:
            conn_status = "completed" if step_num in completed_steps else ""
            html += f'<div class="step-connector {conn_status}"></div>'

    html += '</div>'
    return html


# メインコンテンツ
if mode == "🎨 クリエイティブスタジオ":
    st.markdown(WORKFLOW_CSS, unsafe_allow_html=True)

    st.title("🎨 クリエイティブスタジオ")
    st.markdown("不足画像の特定から楽天・ヤフーへのアップロードまで、一気通貫で処理します。")

    # セッション状態の初期化
    if "workflow_step" not in st.session_state:
        st.session_state.workflow_step = 1
    if "workflow_completed" not in st.session_state:
        st.session_state.workflow_completed = []
    if "workflow_data" not in st.session_state:
        st.session_state.workflow_data = {}

    current_step = st.session_state.workflow_step
    completed_steps = st.session_state.workflow_completed

    # ステップナビゲーション
    st.markdown(render_workflow_step_nav(current_step, completed_steps), unsafe_allow_html=True)

    # ステップ選択（手動でジャンプ可能）
    with st.expander("ステップを直接選択", expanded=False):
        step_options = {
            "① 不足特定": 1,
            "② JAN取得": 2,
            "③ 画像取得": 3,
            "④ アップロード準備": 4,
            "⑤ アップロード": 5
        }
        selected = st.selectbox(
            "移動先",
            list(step_options.keys()),
            index=current_step - 1,
            label_visibility="collapsed"
        )
        if st.button("移動"):
            st.session_state.workflow_step = step_options[selected]
            st.rerun()

    st.divider()

    # ============================================================
    # Step 1: 不足特定
    # ============================================================
    if current_step == 1:
        st.markdown("""
        <div class="step-card">
            <div class="step-card-header">
                <div class="step-card-icon">🔍</div>
                <div>
                    <p class="step-card-title">Step ① 不足特定</p>
                    <p class="step-card-desc">R-Cabinetに存在しない画像を特定します</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # 入力方法の選択
        input_method = st.radio(
            "入力方法",
            ["出品シートExcel", "テキスト入力"],
            horizontal=True
        )

        comic_numbers = []

        # 種別ごとのコミックNoリスト（ファイル間で集約、重複は後で除去）
        typed_comics_raw = {"セット品": [], "単品": [], "予約": []}

        if input_method == "出品シートExcel":
            st.info(
                "**出品シートExcelをアップロード（複数可）**\n\n"
                "- セット品 → A列: 商品コード ／ D列: コミックNo\n"
                "- 単品 → A列: 商品コード ／ E列: コミックNo\n"
                "- 予約 → A列: 商品コード ／ D列: コミックNo\n\n"
                "各ファイルの **Sheet1** が読み込まれます。Step④のヤフーマッピングにも自動連携されます。\n\n"
                "🔍 存在チェック対象フォルダ: **セット** / **単品** / **予約** のみ（他フォルダ・サブフォルダは対象外）"
            )
            excel_files = st.file_uploader("出品シートExcel", type=['xlsx', 'xls'], key="step1_excel", accept_multiple_files=True)

            if excel_files:
                yahoo_files_data = []  # ヤフー連携用
                # ラベル → yahoo type の変換
                yahoo_type_map = {"セット品": "set", "単品": "tanpin", "予約": "yoyaku"}

                for idx, excel_file in enumerate(excel_files):
                    file_type = st.radio(
                        f"📄 {excel_file.name}",
                        ["セット品", "単品", "予約"],
                        horizontal=True,
                        key=f"step1_ftype_{idx}"
                    )

                    try:
                        df = pd.read_excel(excel_file, sheet_name=0, header=None)
                        excel_file.seek(0)
                        file_bytes = excel_file.read()

                        # 単品のみE列、それ以外（セット品・予約）はD列
                        col_idx = 4 if file_type == "単品" else 3
                        for i in range(len(df)):
                            try:
                                cno = str(df.iloc[i, col_idx]).strip().replace('.0', '')
                                if cno and cno != 'nan' and cno.replace('_', '').isdigit():
                                    typed_comics_raw[file_type].append(cno)
                            except:
                                continue
                        yahoo_files_data.append({'bytes': file_bytes, 'type': yahoo_type_map[file_type]})

                        st.caption(f"→ {len(df)}行 / コミックNo抽出済み")

                    except Exception as e:
                        st.error(f"{excel_file.name} 読み込みエラー: {e}")

                # 種別ごとにユニーク化（順序保持）
                typed_comics_unique = {
                    t: list(dict.fromkeys(v)) for t, v in typed_comics_raw.items()
                }
                total_unique = sum(len(v) for v in typed_comics_unique.values())

                if total_unique:
                    st.info(
                        f"合計: {total_unique}件 "
                        f"（セット: {len(typed_comics_unique['セット品'])}件, "
                        f"単品: {len(typed_comics_unique['単品'])}件, "
                        f"予約: {len(typed_comics_unique['予約'])}件）"
                    )
                    # 元行数と差があれば注意表示
                    raw_total = sum(len(v) for v in typed_comics_raw.values())
                    if raw_total != total_unique:
                        st.caption(f"※ 元の抽出行数 {raw_total}件 から重複 {raw_total - total_unique}件 を除去")

                # 後続処理に渡す
                comic_numbers = [c for v in typed_comics_unique.values() for c in v]
                st.session_state.workflow_data['typed_comics'] = typed_comics_unique
                st.session_state.workflow_data['yahoo_excel_files'] = yahoo_files_data
        else:
            # テキスト入力（種別ごとに独立したテキストエリア。空欄の種別はスキップ）
            st.caption(
                "🔍 存在チェック対象フォルダ: **セット** / **単品** / **予約** のみ（他フォルダ・サブフォルダは対象外）／ "
                "各エリアに改行区切りでコミックNoを入力。使わない種別は空欄のままでOK。"
            )
            text_cols = st.columns(3)
            text_inputs = {}
            placeholders = {
                "セット品": "123456\n234567",
                "単品": "19763_003\n19763_004",
                "予約": "345678\n456789",
            }
            for col, type_label in zip(text_cols, ["セット品", "単品", "予約"]):
                with col:
                    text_inputs[type_label] = st.text_area(
                        f"{type_label}（改行区切り）",
                        height=180,
                        placeholder=placeholders[type_label],
                        key=f"step1_text_{type_label}",
                    )

            # 種別ごとにユニーク化（順序保持）
            typed_comics_text = {"セット品": [], "単品": [], "予約": []}
            for type_label, raw in text_inputs.items():
                if not raw:
                    continue
                nums = [line.strip() for line in raw.split('\n') if line.strip()]
                typed_comics_text[type_label] = list(dict.fromkeys(nums))

            total_unique = sum(len(v) for v in typed_comics_text.values())
            if total_unique:
                st.info(
                    f"合計: {total_unique}件 "
                    f"（セット: {len(typed_comics_text['セット品'])}件, "
                    f"単品: {len(typed_comics_text['単品'])}件, "
                    f"予約: {len(typed_comics_text['予約'])}件）"
                )
                comic_numbers = [c for v in typed_comics_text.values() for c in v]
                st.session_state.workflow_data['typed_comics'] = typed_comics_text

        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("🔍 チェック実行", type="primary", disabled=not comic_numbers):
                progress = st.progress(0)
                status = st.empty()

                typed_comics_for_check = st.session_state.workflow_data.get('typed_comics')
                results = check_comic_images(
                    comic_numbers,
                    progress,
                    status,
                    typed_comics=typed_comics_for_check,
                )

                progress.empty()
                status.empty()

                if results:
                    st.session_state.workflow_data['check_results'] = results
                    # 不足リストが更新されたので、下流のキャッシュ/アップロード済フラグを無効化
                    st.session_state.workflow_data['missing_uploaded'] = False
                    st.session_state.workflow_data.pop('missing_from_github', None)
                    # CSV取得キャッシュも無効化（次回取得時に最新を取りに行く）
                    try:
                        fetch_github_csv_bytes.clear()
                    except Exception:
                        pass
                    st.session_state['csv_cache_bust'] = st.session_state.get('csv_cache_bust', 0) + 1
                    exists_count = len([r for r in results if r['存在'] == '✅ あり'])
                    missing_count = len([r for r in results if r['存在'] == '❌ なし'])

                    cols = st.columns(3)
                    cols[0].metric("総数", len(results))
                    cols[1].metric("存在あり", exists_count)
                    cols[2].metric("存在なし", missing_count)

                    if missing_count > 0:
                        st.success(f"不足画像 {missing_count}件 を特定しました")
                else:
                    st.error("DBにデータがありません")

        # 結果がある場合
        if 'check_results' in st.session_state.workflow_data:
            results = st.session_state.workflow_data['check_results']
            exists_items = [r for r in results if r['存在'] == '✅ あり']
            # RECフォルダを除外した画像
            exists_items_no_rec = [r for r in exists_items if 'REC' not in (r.get('フォルダ', '') or '').upper()]
            missing = [r for r in results if r['存在'] == '❌ なし']

            # 存在あり画像のダウンロード
            if exists_items:
                rec_count = len(exists_items) - len(exists_items_no_rec)
                if exists_items_no_rec:
                    if rec_count > 0:
                        expander_label = f"📦 存在あり画像をダウンロード（{len(exists_items_no_rec)}件、REC {rec_count}件除外）"
                    else:
                        expander_label = f"📦 存在あり画像をダウンロード（{len(exists_items_no_rec)}件）"
                else:
                    expander_label = f"📦 存在あり画像（{len(exists_items)}件すべてRECフォルダ）"
                with st.expander(expander_label):
                    if 'wf_rcab_dl_result' not in st.session_state:
                        st.session_state.wf_rcab_dl_result = None

                    if not exists_items_no_rec:
                        st.info("REC除外後の対象が0件のためダウンロードできません")

                    if st.button("🖼️ R-Cabinetから画像を取得", type="primary", key="wf_rcab_dl_btn", disabled=not exists_items_no_rec):
                        _zipfile = get_zipfile()
                        progress = st.progress(0)
                        status = st.empty()
                        session = requests.Session()

                        downloaded = []
                        failed = []
                        for i, item in enumerate(exists_items_no_rec):
                            comic_no = str(item['コミックNo'])
                            url = item.get('URL', '')
                            folder = item.get('フォルダ', '')
                            file_name = item.get('ファイル名', '') or f"{comic_no}.jpg"
                            if '.' not in file_name:
                                file_name = f"{file_name}.jpg"

                            status.text(f"ダウンロード中: {comic_no} ({i+1}/{len(exists_items_no_rec)})")
                            progress.progress((i + 1) / len(exists_items_no_rec))

                            if not url or url == '-':
                                failed.append(comic_no)
                                continue
                            try:
                                resp = session.get(url, timeout=15)
                                if resp.status_code == 200 and len(resp.content) > 100:
                                    downloaded.append({
                                        'comic_no': comic_no,
                                        'file_name': file_name,
                                        'folder': folder,
                                        'data': resp.content,
                                    })
                                else:
                                    failed.append(comic_no)
                            except Exception:
                                failed.append(comic_no)

                        progress.empty()
                        status.empty()
                        st.session_state.wf_rcab_dl_result = {'downloaded': downloaded, 'failed': failed}
                        st.rerun()

                    if st.session_state.wf_rcab_dl_result:
                        dl_result = st.session_state.wf_rcab_dl_result
                        downloaded = dl_result['downloaded']
                        failed = dl_result['failed']

                        st.success(f"取得完了: {len(downloaded)}件成功" + (f", {len(failed)}件失敗" if failed else ""))

                        if downloaded:
                            _zipfile = get_zipfile()
                            dl_cols = st.columns(2)

                            with dl_cols[0]:
                                buf_flat = BytesIO()
                                with _zipfile.ZipFile(buf_flat, 'w', _zipfile.ZIP_DEFLATED) as zf:
                                    for img in downloaded:
                                        zf.writestr(img['file_name'], img['data'])
                                buf_flat.seek(0)
                                st.download_button(
                                    label=f"📦 フラットZIP（{len(downloaded)}件）",
                                    data=buf_flat,
                                    file_name="rcabinet_images_flat.zip",
                                    mime="application/zip",
                                    key="wf_rcab_dl_flat"
                                )
                                st.caption("全画像を直下に配置")

                            with dl_cols[1]:
                                buf_folder = BytesIO()
                                with _zipfile.ZipFile(buf_folder, 'w', _zipfile.ZIP_DEFLATED) as zf:
                                    for img in downloaded:
                                        folder = img['folder'] if img['folder'] and img['folder'] != '-' else 'その他'
                                        zf.writestr(f"{folder}/{img['file_name']}", img['data'])
                                buf_folder.seek(0)
                                st.download_button(
                                    label=f"📂 フォルダ付きZIP（{len(downloaded)}件）",
                                    data=buf_folder,
                                    file_name="rcabinet_images_by_folder.zip",
                                    mime="application/zip",
                                    key="wf_rcab_dl_folder"
                                )
                                st.caption("R-Cabinetのフォルダ構成を保持")

            if missing:
                st.divider()
                st.markdown("### 不足画像一覧")

                df_missing = pd.DataFrame(missing)
                st.dataframe(df_missing, use_container_width=True, height=200)

                col1, col2 = st.columns([1, 1])
                with col1:
                    if st.button("📤 GitHubにアップロード", type="secondary"):
                        # 種別（セット品/単品/予約）ごとに分離
                        set_comics = [r['コミックNo'] for r in missing if r.get('種別') == 'セット品']
                        tanpin_comics = [r['コミックNo'] for r in missing if r.get('種別') == '単品']
                        yoyaku_comics = [r['コミックNo'] for r in missing if r.get('種別') == '予約']

                        today = datetime.now(JST).strftime('%Y-%m-%d %H:%M')
                        upload_results = []

                        if set_comics:
                            content = '\n'.join([str(c) for c in set_comics])
                            result = upload_to_github(content, GITHUB_MISSING_CSV_PATH, f"Update missing_comics.csv ({len(set_comics)}件) - {today}")
                            if result.get("success"):
                                upload_results.append(f"セット品: {len(set_comics)}件 ✅")

                        if tanpin_comics:
                            content = '\n'.join([str(c) for c in tanpin_comics])
                            result = upload_to_github(content, GITHUB_MISSING_TANPIN_PATH, f"Update missing_tanpin.csv ({len(tanpin_comics)}件) - {today}")
                            if result.get("success"):
                                upload_results.append(f"単品: {len(tanpin_comics)}件 ✅")

                        if yoyaku_comics:
                            content = '\n'.join([str(c) for c in yoyaku_comics])
                            result = upload_to_github(content, GITHUB_MISSING_YOYAKU_PATH, f"Update missing_yoyaku.csv ({len(yoyaku_comics)}件) - {today}")
                            if result.get("success"):
                                upload_results.append(f"予約: {len(yoyaku_comics)}件 ✅")

                        if upload_results:
                            st.success(", ".join(upload_results))
                            st.session_state.workflow_data['missing_uploaded'] = True

                with col2:
                    if st.button("次へ進む →", type="primary"):
                        if 1 not in st.session_state.workflow_completed:
                            st.session_state.workflow_completed.append(1)
                        st.session_state.workflow_step = 2
                        st.rerun()

    # ============================================================
    # Step 2: JAN取得
    # ============================================================
    elif current_step == 2:
        st.markdown("""
        <div class="step-card">
            <div class="step-card-header">
                <div class="step-card-icon">📊</div>
                <div>
                    <p class="step-card-title">Step ② JAN取得</p>
                    <p class="step-card-desc">コミックリスターでJAN情報を取得します（GitHub Actions）</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # 最新の実行履歴を表示
        runs = get_workflow_runs("weekly-comic-lister.yml", limit=3)
        if runs:
            st.markdown("### 実行履歴")
            for run in runs:
                status_icon = "🟢" if run["conclusion"] == "success" else "🔴" if run["conclusion"] == "failure" else "🟡"
                st.write(f"{status_icon} {run['created_at']} - {run['conclusion'] or '実行中'}")

        btn_cols = st.columns([1, 1])
        with btn_cols[0]:
            run_clicked = st.button("📊 CSV生成・取得", type="primary", use_container_width=True)
        with btn_cols[1]:
            reuse_clicked = st.button(
                "♻️ 前回の情報を取得",
                help="GitHub Actionsを起動せず、前回生成済みの is_list.csv / comic_list.csv をそのまま取得します（テスト・再実行用）",
                use_container_width=True,
            )

        if reuse_clicked:
            status_area = st.empty()
            status_area.info("📥 前回のCSVを取得中...")
            is_result = download_from_github(GITHUB_IS_LIST_PATH)
            cl_result = download_from_github(GITHUB_COMIC_LIST_PATH)
            if is_result.get("success") and cl_result.get("success"):
                is_content = is_result["content"]
                if isinstance(is_content, bytes):
                    is_content = is_content.decode('utf-8', errors='replace')
                cl_content = cl_result["content"]
                if isinstance(cl_content, bytes):
                    cl_content = cl_content.decode('utf-8', errors='replace')
                st.session_state.workflow_data['is_list'] = is_content
                st.session_state.workflow_data['comic_list'] = cl_content
                status_area.success("✅ 前回のCSVを取得しました。Step ③に進めます")
            else:
                err_msgs = []
                if not is_result.get("success"):
                    err_msgs.append(f"is_list.csv: {is_result.get('error')}")
                if not cl_result.get("success"):
                    err_msgs.append(f"comic_list.csv: {cl_result.get('error')}")
                status_area.error("CSVのダウンロードに失敗しました / " + " / ".join(err_msgs))

        if run_clicked:
            status_area = st.empty()
            progress_area = st.empty()

            # 1. 不足リストを常にGitHubへアップロード（現在のStep①結果で上書き）
            #    - ステールな前回データがGitHub側に残らないよう、各種別とも空でも上書きする
            check_results = st.session_state.workflow_data.get('check_results', [])
            missing = [r for r in check_results if r.get('存在') == '❌ なし']
            set_comics = [r['コミックNo'] for r in missing if r.get('種別') == 'セット品']
            tanpin_comics = [r['コミックNo'] for r in missing if r.get('種別') == '単品']
            yoyaku_comics = [r['コミックNo'] for r in missing if r.get('種別') == '予約']
            today = datetime.now(JST).strftime('%Y-%m-%d %H:%M')
            status_area.info(
                f"📤 不足リストをGitHubにアップロード中... "
                f"(セット品:{len(set_comics)}件 / 単品:{len(tanpin_comics)}件 / 予約:{len(yoyaku_comics)}件)"
            )
            # 完全な空はGitHub APIで失敗する可能性があるため、空時は改行1文字（下流の split('\n') + strip でスキップされる）
            set_content = '\n'.join([str(c) for c in set_comics]) if set_comics else '\n'
            tanpin_content = '\n'.join([str(c) for c in tanpin_comics]) if tanpin_comics else '\n'
            yoyaku_content = '\n'.join([str(c) for c in yoyaku_comics]) if yoyaku_comics else '\n'
            upload_to_github(set_content, GITHUB_MISSING_CSV_PATH, f"Update missing_comics.csv ({len(set_comics)}件) - {today}")
            upload_to_github(tanpin_content, GITHUB_MISSING_TANPIN_PATH, f"Update missing_tanpin.csv ({len(tanpin_comics)}件) - {today}")
            upload_to_github(yoyaku_content, GITHUB_MISSING_YOYAKU_PATH, f"Update missing_yoyaku.csv ({len(yoyaku_comics)}件) - {today}")
            st.session_state.workflow_data['missing_uploaded'] = True

            # 2. GitHub Actionsを起動
            status_area.info("🚀 GitHub Actionsを起動中...")
            result = trigger_github_actions("weekly-comic-lister.yml")
            if not result.get("success"):
                status_area.error(f"エラー: {result.get('error')}")
            else:
                # 3. ワークフロー完了をポーリング（最大5分）
                import time as _time
                max_wait = 300
                poll_interval = 15
                elapsed = 0
                completed = False

                status_area.info("⏳ ワークフロー実行中... 完了まで2〜3分かかります")

                # トリガー直後は少し待つ（新しいrunが作られるまで）
                _time.sleep(5)
                elapsed += 5

                while elapsed < max_wait:
                    _time.sleep(poll_interval)
                    elapsed += poll_interval
                    minutes = elapsed // 60
                    seconds = elapsed % 60
                    time_str = f"{minutes}分{seconds}秒" if minutes > 0 else f"{seconds}秒"
                    progress_area.progress(min(elapsed / max_wait, 0.95), text=f"⏳ ワークフロー実行中... ({time_str}経過)")

                    runs = get_workflow_runs("weekly-comic-lister.yml", limit=1)
                    if not runs:
                        continue
                    run = runs[0]
                    # まだ実行中なら待機を続ける
                    if run.get("status") in ("queued", "in_progress"):
                        continue
                    # 完了した場合
                    if run.get("conclusion") == "success":
                        completed = True
                    break

                progress_area.empty()

                if completed:
                    # 4. 完了後にCSVを自動ダウンロード
                    status_area.info("📥 ワークフロー完了。CSVをダウンロード中...")
                    is_result = download_from_github(GITHUB_IS_LIST_PATH)
                    cl_result = download_from_github(GITHUB_COMIC_LIST_PATH)

                    if is_result.get("success") and cl_result.get("success"):
                        is_content = is_result["content"]
                        if isinstance(is_content, bytes):
                            is_content = is_content.decode('utf-8', errors='replace')
                        cl_content = cl_result["content"]
                        if isinstance(cl_content, bytes):
                            cl_content = cl_content.decode('utf-8', errors='replace')
                        st.session_state.workflow_data['is_list'] = is_content
                        st.session_state.workflow_data['comic_list'] = cl_content
                        status_area.success("✅ CSV生成・取得が完了しました。Step ③に進めます")
                        st.balloons()
                    else:
                        status_area.error("CSVのダウンロードに失敗しました")
                elif elapsed >= max_wait:
                    status_area.warning("⏱️ タイムアウト（5分）しました。完了後に再度ボタンを押してください")
                else:
                    status_area.error("❌ ワークフローが失敗しました")

        # ファイル状態表示
        st.divider()
        st.markdown("### ファイル状態")
        col1, col2 = st.columns(2)
        with col1:
            if st.session_state.workflow_data.get('is_list'):
                st.success("✅ is_list.csv 取得済み")
            else:
                st.warning("⬜ is_list.csv 未取得")
        with col2:
            if st.session_state.workflow_data.get('comic_list'):
                st.success("✅ comic_list.csv 取得済み")
            else:
                st.warning("⬜ comic_list.csv 未取得")

        # 次へ進むボタン
        st.divider()
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("← 戻る"):
                st.session_state.workflow_step = 1
                st.rerun()
        with col2:
            files_ready = st.session_state.workflow_data.get('is_list') and st.session_state.workflow_data.get('comic_list')
            if st.button("次へ進む →", type="primary", disabled=not files_ready):
                if 2 not in st.session_state.workflow_completed:
                    st.session_state.workflow_completed.append(2)
                st.session_state.workflow_step = 3
                st.rerun()

    # ============================================================
    # Step 3: 画像取得
    # ============================================================
    elif current_step == 3:
        import os
        st.markdown("""
        <div class="step-card">
            <div class="step-card-header">
                <div class="step-card-icon">🖼️</div>
                <div>
                    <p class="step-card-title">Step ③ 画像取得＋加工</p>
                    <p class="step-card-desc">セット: JANコードで1巻目の画像を取得し、送料無料バッジを合成します（600×600px）</p>
                    <p class="step-card-desc">単品: JANコードで対象巻の画像を取得します。中央配置（600×600px）</p>
                    <p class="step-card-desc">予約: JANコードで最新巻の画像を取得し、送料無料バッジを合成します（600×600px）</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # --- 入力データの確認 ---
        st.markdown("### 入力データ")

        # 不足リスト（Step ①の結果 or GitHubから取得済み）
        # missing_comics: セット+単品の従来リスト（下流互換用）
        # missing_yoyaku: 予約のcomic_noリスト（新規）
        missing_comics = []
        missing_yoyaku = []
        if 'check_results' in st.session_state.workflow_data:
            missing = [r for r in st.session_state.workflow_data['check_results'] if r['存在'] == '❌ なし']
            for r in missing:
                cno = normalize_jan_code(r['コミックNo'])
                if not cno:
                    continue
                t = r.get('種別', '')
                if t == '予約':
                    missing_yoyaku.append(cno)
                else:
                    missing_comics.append(cno)
        elif st.session_state.workflow_data.get('missing_from_github'):
            gh = st.session_state.workflow_data['missing_from_github']
            if isinstance(gh, dict):
                missing_comics = list(gh.get('set', [])) + list(gh.get('tanpin', []))
                missing_yoyaku = list(gh.get('yoyaku', []))
            else:
                missing_comics = list(gh)

        total_missing = len(missing_comics) + len(missing_yoyaku)
        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1:
            if total_missing:
                st.success(f"✅ 不足リスト: {total_missing}件")
            else:
                st.warning("⬜ 不足リスト: なし")
                st.caption("Step ①を実行するか、GitHubから取得してください")
        with col_s2:
            if st.session_state.workflow_data.get('is_list'):
                st.success("✅ is_list.csv 取得済み")
            else:
                st.warning("⬜ is_list.csv 未取得")
        with col_s3:
            if st.session_state.workflow_data.get('comic_list'):
                st.success("✅ comic_list.csv 取得済み")
            else:
                st.warning("⬜ comic_list.csv 未取得")

        # GitHubから取得ボタン群
        need_fetch = (not missing_comics and not missing_yoyaku) or not st.session_state.workflow_data.get('is_list') or not st.session_state.workflow_data.get('comic_list')
        if need_fetch:
            st.divider()
            fetch_cols = st.columns(3)
            with fetch_cols[0]:
                if (not missing_comics and not missing_yoyaku) and st.button("📥 不足リスト取得"):
                    with st.spinner("GitHubから取得中..."):
                        parsed_set = []
                        parsed_tanpin = []
                        parsed_yoyaku = []

                        def _parse_csv_lines(content_bytes, is_set_csv=False):
                            """CSVの各行から comic_no を抽出"""
                            out = []
                            if isinstance(content_bytes, bytes):
                                content_str = content_bytes.decode('utf-8', errors='replace')
                            else:
                                content_str = str(content_bytes)
                            for l in content_str.strip().split('\n'):
                                l = l.strip()
                                if not l:
                                    continue
                                if is_set_csv and ',' in l:
                                    # セット: カンマ区切りの複数フィールドから数字っぽいものを拾う
                                    for f in (x.strip() for x in l.split(',') if x.strip()):
                                        if f.replace('_', '').replace('.0', '').isdigit() and len(f.replace('.0', '')) > 1:
                                            out.append(f.replace('.0', ''))
                                            break
                                else:
                                    out.append(l)
                            return out

                        # セット品（missing_comics.csv → 表示上は missing_set.csv）
                        r_set = download_from_github(GITHUB_MISSING_CSV_PATH)
                        if r_set.get("success"):
                            parsed_set = _parse_csv_lines(r_set.get("content", b""), is_set_csv=True)

                        # 単品（missing_tanpin.csv）
                        r_tanpin = download_from_github(GITHUB_MISSING_TANPIN_PATH)
                        if r_tanpin.get("success"):
                            parsed_tanpin = _parse_csv_lines(r_tanpin.get("content", b""))

                        # 予約（missing_yoyaku.csv）
                        r_yoyaku = download_from_github(GITHUB_MISSING_YOYAKU_PATH)
                        if r_yoyaku.get("success"):
                            parsed_yoyaku = _parse_csv_lines(r_yoyaku.get("content", b""))

                    total = len(parsed_set) + len(parsed_tanpin) + len(parsed_yoyaku)
                    if total:
                        missing_comics = parsed_set + parsed_tanpin
                        missing_yoyaku = parsed_yoyaku
                        st.session_state.workflow_data['missing_from_github'] = {
                            'set': parsed_set,
                            'tanpin': parsed_tanpin,
                            'yoyaku': parsed_yoyaku,
                        }
                        st.success(f"{total}件取得（セット: {len(parsed_set)}件 / 単品: {len(parsed_tanpin)}件 / 予約: {len(parsed_yoyaku)}件）")
                        st.rerun()
                    else:
                        st.error("取得失敗またはデータなし")
            with fetch_cols[1]:
                if not st.session_state.workflow_data.get('is_list') and st.button("📥 is_list.csv取得"):
                    with st.spinner("取得中..."):
                        result = download_from_github(GITHUB_IS_LIST_PATH)
                    if result.get("success"):
                        content = result["content"]
                        if isinstance(content, bytes):
                            content = content.decode('utf-8', errors='replace')
                        st.session_state.workflow_data['is_list'] = content
                        st.success("取得完了")
                        st.rerun()
                    else:
                        st.error("取得失敗")
            with fetch_cols[2]:
                if not st.session_state.workflow_data.get('comic_list') and st.button("📥 comic_list.csv取得"):
                    with st.spinner("取得中..."):
                        result = download_from_github(GITHUB_COMIC_LIST_PATH)
                    if result.get("success"):
                        content = result["content"]
                        if isinstance(content, bytes):
                            content = content.decode('utf-8', errors='replace')
                        st.session_state.workflow_data['comic_list'] = content
                        st.success("取得完了")
                        st.rerun()
                    else:
                        st.error("取得失敗")

        # GitHubから取得した不足リストも反映
        if not missing_comics and not missing_yoyaku and st.session_state.workflow_data.get('missing_from_github'):
            gh = st.session_state.workflow_data['missing_from_github']
            if isinstance(gh, dict):
                missing_comics = list(gh.get('set', [])) + list(gh.get('tanpin', []))
                missing_yoyaku = list(gh.get('yoyaku', []))
            else:
                missing_comics = list(gh)

        # --- CSVダウンロード ---
        # (表示ファイル名, GitHubパス) — 表示名はユーザー希望に合わせて missing_set.csv を使用
        csv_files = [
            ("missing_tanpin.csv", GITHUB_MISSING_TANPIN_PATH),
            ("missing_set.csv", GITHUB_MISSING_CSV_PATH),
            ("missing_yoyaku.csv", GITHUB_MISSING_YOYAKU_PATH),
            ("is_list.csv", GITHUB_IS_LIST_PATH),
            ("comic_list.csv", GITHUB_COMIC_LIST_PATH),
        ]
        if 'csv_cache_bust' not in st.session_state:
            st.session_state.csv_cache_bust = 0

        dl_header_cols = st.columns([6, 1])
        with dl_header_cols[0]:
            st.markdown("**📄 CSVファイルをダウンロード**")
        with dl_header_cols[1]:
            if st.button("🔄 再取得", key="csv_refresh_btn", help="GitHubからCSVを再取得", use_container_width=True):
                st.session_state.csv_cache_bust += 1
                fetch_github_csv_bytes.clear()
                st.rerun()

        cache_bust = st.session_state.csv_cache_bust
        dl_cols = st.columns(len(csv_files))
        for idx, (fname, gpath) in enumerate(csv_files):
            with dl_cols[idx]:
                data = fetch_github_csv_bytes(gpath, cache_bust=cache_bust)
                if data:
                    st.download_button(
                        label=f"💾 {fname}",
                        data=data,
                        file_name=fname,
                        mime="text/csv",
                        key=f"dl_{fname}_{cache_bust}",
                        use_container_width=True,
                    )
                else:
                    st.button(
                        f"❌ {fname}",
                        key=f"dl_err_{fname}_{cache_bust}",
                        disabled=True,
                        help="取得失敗（右上の再取得ボタンでリトライ）",
                        use_container_width=True,
                    )

        # --- 実行セクション ---
        st.divider()
        has_any = bool(missing_comics) or bool(missing_yoyaku)
        all_ready = has_any and st.session_state.workflow_data.get('is_list') and st.session_state.workflow_data.get('comic_list')

        # セット品・単品・予約の内訳
        if has_any:
            set_count = len([c for c in missing_comics if '_' not in str(c)])
            tanpin_count = len([c for c in missing_comics if '_' in str(c)])
            yoyaku_count = len(missing_yoyaku)
            total_count = set_count + tanpin_count + yoyaku_count
            st.markdown(f"**対象: {total_count}件**（セット品: {set_count}件 / 単品: {tanpin_count}件 / 予約: {yoyaku_count}件）")

        # --- 画像取得：1件ずつrerunで進めるインクリメンタル方式 ---
        if 'wf_img_processing' not in st.session_state:
            st.session_state.wf_img_processing = False
        if 'wf_img_paused' not in st.session_state:
            st.session_state.wf_img_paused = False
        if 'wf_img_target_data' not in st.session_state:
            st.session_state.wf_img_target_data = []
        if 'wf_img_index' not in st.session_state:
            st.session_state.wf_img_index = 0
        if 'wf_img_downloaded' not in st.session_state:
            st.session_state.wf_img_downloaded = []
        if 'wf_img_stats' not in st.session_state:
            st.session_state.wf_img_stats = {}
        if 'wf_img_logs' not in st.session_state:
            st.session_state.wf_img_logs = []
        if 'wf_img_badge_path' not in st.session_state:
            st.session_state.wf_img_badge_path = ''

        is_processing = st.session_state.wf_img_processing
        is_paused = st.session_state.wf_img_paused

        ctrl_cols = st.columns([3, 2, 5])
        with ctrl_cols[0]:
            start_clicked = st.button(
                "🖼️ 画像取得開始",
                type="primary",
                disabled=not all_ready or is_processing,
                key="wf_img_start_btn",
                use_container_width=True,
            )
        with ctrl_cols[1]:
            pause_clicked = False
            resume_clicked = False
            if is_processing and is_paused:
                resume_clicked = st.button("▶️ 再開", key="wf_img_resume_btn", use_container_width=True)
            elif is_processing:
                pause_clicked = st.button("⏸️ 一時停止", key="wf_img_pause_btn", use_container_width=True)

        if start_clicked:
            badge_path = os.path.join(os.path.dirname(__file__), "images", "badge_free_shipping.jpg")
            target_data = _workflow_prepare_target_data(
                missing_comics,
                st.session_state.workflow_data['is_list'],
                st.session_state.workflow_data['comic_list'],
                missing_yoyaku=missing_yoyaku,
            )
            if not target_data:
                st.error("処理対象がありません")
            else:
                st.session_state.wf_img_target_data = target_data
                st.session_state.wf_img_index = 0
                st.session_state.wf_img_downloaded = []
                st.session_state.wf_img_stats = {
                    'total': len(target_data), 'success': 0, 'failed': 0,
                    'bookoff': 0, 'amazon': 0, 'rakuten': 0, 'gemini_ai': 0,
                }
                st.session_state.wf_img_logs = []
                st.session_state.wf_img_badge_path = badge_path
                st.session_state.wf_img_processing = True
                st.session_state.wf_img_paused = False
                st.session_state.wf_img_session = requests.Session()
                # 既存結果をクリア
                for k in ('downloaded_images', 'image_stats', 'image_logs'):
                    st.session_state.workflow_data.pop(k, None)
                st.rerun()

        if pause_clicked:
            st.session_state.wf_img_paused = True
            # そこまでの結果を保存（ダウンロード可能にする）
            st.session_state.workflow_data['downloaded_images'] = list(st.session_state.wf_img_downloaded)
            st.session_state.workflow_data['image_stats'] = dict(st.session_state.wf_img_stats)
            st.session_state.workflow_data['image_logs'] = list(st.session_state.wf_img_logs)
            st.rerun()

        if resume_clicked:
            st.session_state.wf_img_paused = False
            st.rerun()

        # 進捗表示
        if is_processing:
            total = len(st.session_state.wf_img_target_data)
            idx = st.session_state.wf_img_index
            if total:
                st.progress(min(idx / total, 1.0))
            if is_paused:
                st.warning(f"⏸️ 一時停止中（{idx}/{total}件処理済み）。下部の「取得結果」から現時点の画像をダウンロードできます。")
            else:
                cur_no = st.session_state.wf_img_target_data[idx].get('comic_no', '') if idx < total else ''
                st.info(f"🔄 処理中: {idx + 1}/{total} {cur_no}")

        # 1件処理 → rerun
        if is_processing and not is_paused:
            total = len(st.session_state.wf_img_target_data)
            idx = st.session_state.wf_img_index
            if idx < total:
                data = st.session_state.wf_img_target_data[idx]
                session = st.session_state.get('wf_img_session') or requests.Session()
                result = _workflow_process_one_image(data, session, st.session_state.wf_img_badge_path)
                st.session_state.wf_img_logs.append(result['log'])
                if result['success']:
                    st.session_state.wf_img_downloaded.append(result['image'])
                    st.session_state.wf_img_stats['success'] = st.session_state.wf_img_stats.get('success', 0) + 1
                    if result['source']:
                        st.session_state.wf_img_stats[result['source']] = st.session_state.wf_img_stats.get(result['source'], 0) + 1
                else:
                    st.session_state.wf_img_stats['failed'] = st.session_state.wf_img_stats.get('failed', 0) + 1
                st.session_state.wf_img_index += 1
                # 毎回ワークフロー側にも反映（途中停止でも最新状態を保持）
                st.session_state.workflow_data['downloaded_images'] = list(st.session_state.wf_img_downloaded)
                st.session_state.workflow_data['image_stats'] = dict(st.session_state.wf_img_stats)
                st.session_state.workflow_data['image_logs'] = list(st.session_state.wf_img_logs)
                # レート制限
                time.sleep(get_random().uniform(0.3, 0.8))
                st.rerun()
            else:
                # 完了
                st.session_state.wf_img_processing = False
                st.session_state.workflow_data['downloaded_images'] = list(st.session_state.wf_img_downloaded)
                st.session_state.workflow_data['image_stats'] = dict(st.session_state.wf_img_stats)
                st.session_state.workflow_data['image_logs'] = list(st.session_state.wf_img_logs)
                st.rerun()

        # --- 結果表示セクション ---
        if 'downloaded_images' in st.session_state.workflow_data:
            images = st.session_state.workflow_data['downloaded_images']
            stats = st.session_state.workflow_data.get('image_stats', {})
            logs = st.session_state.workflow_data.get('image_logs', [])

            st.divider()
            st.markdown("### 取得結果")

            # 統計
            stat_cols = st.columns(6)
            stat_cols[0].metric("対象", stats.get('total', 0))
            stat_cols[1].metric("成功", stats.get('success', 0))
            stat_cols[2].metric("失敗", stats.get('failed', 0))
            stat_cols[3].metric("ブックオフ", stats.get('bookoff', 0))
            stat_cols[4].metric("Amazon", stats.get('amazon', 0))
            stat_cols[5].metric("楽天/AI", f"{stats.get('rakuten', 0)}/{stats.get('gemini_ai', 0)}")

            # 画像プレビュー（3列グリッド、最大6件）
            if images:
                preview_count = min(len(images), 6)
                st.markdown(f"### プレビュー（{preview_count}/{len(images)}件）")
                for row_start in range(0, preview_count, 3):
                    cols = st.columns(3)
                    for j, col in enumerate(cols):
                        idx = row_start + j
                        if idx < len(images):
                            img = images[idx]
                            with col:
                                st.image(img['image_data'], width=200)
                                badge_label = "🏷️ バッジ付き" if img['badge'] else "📷 バッジなし"
                                st.caption(f"{img['comic_no']} ({img['source']}) {badge_label}")

                # ZIPダウンロード
                st.divider()
                zipfile = get_zipfile()
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for img in images:
                        filename = f"{img['comic_no']}.jpg"
                        zf.writestr(filename, img['image_data'])

                st.download_button(
                    label=f"📦 ZIPダウンロード（{len(images)}件）",
                    data=zip_buffer.getvalue(),
                    file_name="workflow_images.zip",
                    mime="application/zip"
                )

            # ログ
            with st.expander("ログ", expanded=False):
                for log in logs:
                    st.text(log)

        # ナビゲーション
        st.divider()
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("← 戻る"):
                st.session_state.workflow_step = 2
                st.rerun()
        with col2:
            has_images = 'downloaded_images' in st.session_state.workflow_data and st.session_state.workflow_data['downloaded_images']
            if st.button("次へ進む →", type="primary", disabled=not has_images):
                if 3 not in st.session_state.workflow_completed:
                    st.session_state.workflow_completed.append(3)
                st.session_state.workflow_step = 4
                st.rerun()

    # ============================================================
    # Step 4: アップロード準備
    # ============================================================
    elif current_step == 4:
        import os as _os
        st.markdown("""
        <div class="step-card">
            <div class="step-card-header">
                <div class="step-card-icon">📦</div>
                <div>
                    <p class="step-card-title">Step ④ アップロード準備</p>
                    <p class="step-card-desc">楽天・ヤフー向けにファイルを整形します</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Step ③の画像確認
        images = st.session_state.workflow_data.get('downloaded_images', [])
        if not images:
            st.warning("Step ③で画像を取得してください。")
            if st.button("← Step ③に戻る"):
                st.session_state.workflow_step = 3
                st.rerun()
        else:
            st.info(f"対象画像: {len(images)}件（セット品: {len([i for i in images if not i.get('is_tanpin')])}件 / 単品: {len([i for i in images if i.get('is_tanpin')])}件）")

            tab_yahoo, tab_rakuten = st.tabs(["🛒 ヤフー準備", "🏪 楽天準備"])

            # ============================================
            # ヤフータブ
            # ============================================
            with tab_yahoo:
                st.markdown("### データフォーマットExcel")

                excel_set_df = None
                excel_tanpin_df = None

                # Step ①で出品シートExcelが読み込み済みか確認
                yahoo_from_step1 = st.session_state.workflow_data.get('yahoo_excel_files')

                if yahoo_from_step1:
                    st.success(f"Step ①で出品シートExcel {len(yahoo_from_step1)}件 を読み込み済み")
                    try:
                        set_dfs = []
                        tanpin_dfs = []
                        for fd in yahoo_from_step1:
                            df = pd.read_excel(BytesIO(fd['bytes']), sheet_name=0, header=None)
                            if fd['type'] == 'set':
                                set_dfs.append(df)
                            else:
                                tanpin_dfs.append(df)
                        if set_dfs:
                            excel_set_df = pd.concat(set_dfs, ignore_index=True)
                        if tanpin_dfs:
                            excel_tanpin_df = pd.concat(tanpin_dfs, ignore_index=True)
                        st.caption(f"セット品: {len(set_dfs)}ファイル / 単品: {len(tanpin_dfs)}ファイル")
                    except Exception as e:
                        st.error(f"Excel読み込みエラー: {e}")
                        yahoo_from_step1 = None

                if not yahoo_from_step1:
                    st.info(
                        "**出品シートExcelをアップロード（複数可）**\n\n"
                        "- セット品 → A列: 商品コード ／ D列: コミックNo\n"
                        "- 単品 → A列: 商品コード ／ E列: コミックNo\n\n"
                        "各ファイルの **Sheet1** が読み込まれます。"
                    )
                    yahoo_excels = st.file_uploader("Excelファイル", type=['xlsx', 'xls'], key="yahoo_excel", accept_multiple_files=True)

                    if yahoo_excels:
                        set_dfs = []
                        tanpin_dfs = []
                        for idx, yf in enumerate(yahoo_excels):
                            file_type = st.radio(
                                f"📄 {yf.name}",
                                ["セット品", "単品"],
                                horizontal=True,
                                key=f"yahoo_ftype_{idx}"
                            )
                            try:
                                df = pd.read_excel(yf, sheet_name=0, header=None)
                                if file_type == "セット品":
                                    set_dfs.append(df)
                                else:
                                    tanpin_dfs.append(df)
                            except Exception as e:
                                st.error(f"{yf.name} 読み込みエラー: {e}")

                        if set_dfs:
                            excel_set_df = pd.concat(set_dfs, ignore_index=True)
                        if tanpin_dfs:
                            excel_tanpin_df = pd.concat(tanpin_dfs, ignore_index=True)

                # プレビュー（どちらの読み込み方法でも共通）
                if excel_set_df is not None:
                    with st.expander("マッピングプレビュー", expanded=True):
                        set_mappings = []
                        for i in range(len(excel_set_df)):
                            try:
                                code = str(excel_set_df.iloc[i, 0]).strip()
                                cno = str(excel_set_df.iloc[i, 3]).strip().replace('.0', '')
                                if code != 'nan' and cno != 'nan' and code and cno and cno.replace('_', '').isdigit():
                                    matched = any(str(img['comic_no']) == cno for img in images)
                                    set_mappings.append({'商品コード': code, 'コミックNo': cno, '対象画像': '✅' if matched else '❌'})
                            except:
                                continue

                        tanpin_mappings = []
                        if excel_tanpin_df is not None:
                            for i in range(len(excel_tanpin_df)):
                                try:
                                    code = str(excel_tanpin_df.iloc[i, 0]).strip()
                                    cno = str(excel_tanpin_df.iloc[i, 4]).strip().replace('.0', '')
                                    if code != 'nan' and cno != 'nan' and code and cno and cno.replace('_', '').isdigit():
                                        matched = any(str(img['comic_no']) == cno for img in images)
                                        tanpin_mappings.append({'商品コード': code, 'コミックNo': cno, '対象画像': '✅' if matched else '❌'})
                                except:
                                    continue

                        if set_mappings:
                            st.markdown(f"**セット品: {len(set_mappings)}件**")
                            st.dataframe(pd.DataFrame(set_mappings), use_container_width=True, height=150)
                        if tanpin_mappings:
                            st.markdown(f"**単品: {len(tanpin_mappings)}件**")
                            st.dataframe(pd.DataFrame(tanpin_mappings), use_container_width=True, height=150)

                # ZIP生成ボタン
                st.divider()
                can_generate = excel_set_df is not None
                if st.button("📦 ZIP生成", type="primary", disabled=not can_generate, key="yahoo_zip_btn"):
                    additional_dir = _os.path.join(_os.path.dirname(__file__), "images")
                    result = prepare_yahoo_zips(images, excel_set_df, excel_tanpin_df, additional_dir)

                    st.session_state.workflow_data['yahoo_zips'] = result
                    st.rerun()

                # ZIP結果表示
                if 'yahoo_zips' in st.session_state.workflow_data:
                    result = st.session_state.workflow_data['yahoo_zips']
                    zips = result['zips']

                    col_m1, col_m2, col_m3 = st.columns(3)
                    col_m1.metric("マッピング成功", result['mapped'])
                    col_m2.metric("未マッチ", len(result['unmapped']))
                    col_m3.metric("ZIP数", len(zips))

                    for i, zip_data in enumerate(zips):
                        size_mb = len(zip_data) / (1024 * 1024)
                        st.download_button(
                            label=f"📥 yahoo_upload_{i+1:03d}.zip ({size_mb:.1f}MB)",
                            data=zip_data,
                            file_name=f"yahoo_upload_{i+1:03d}.zip",
                            mime="application/zip",
                            key=f"yahoo_zip_dl_{i}"
                        )

                    with st.expander("ログ", expanded=False):
                        for log in result['logs']:
                            st.text(log)

            # ============================================
            # 楽天タブ
            # ============================================
            with tab_rakuten:
                st.markdown("### フォルダマッピング")

                # フォルダ階層リスト取得
                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    if st.button("📥 フォルダ階層リスト取得", key="rakuten_hierarchy_btn"):
                        with st.spinner("GitHubから取得中..."):
                            result = download_from_github(GITHUB_FOLDER_HIERARCHY_PATH)
                        if result.get("success"):
                            content = result["content"]
                            if isinstance(content, bytes):
                                st.session_state.workflow_data['hierarchy_xlsx'] = content
                            st.success("フォルダ階層リスト取得完了")
                            st.rerun()
                        else:
                            st.error(f"取得失敗: {result.get('error')}")

                with col_r2:
                    if st.button("📂 楽天フォルダ一覧取得", key="rakuten_folders_btn"):
                        with st.spinner("R-Cabinet APIから取得中..."):
                            folders, error = get_all_folders()
                        if error:
                            st.error(error)
                        elif folders:
                            st.session_state.workflow_data['rakuten_folders'] = folders
                            st.success(f"{len(folders)}フォルダ取得")
                            st.rerun()

                # 状態表示
                col_s1, col_s2 = st.columns(2)
                with col_s1:
                    if st.session_state.workflow_data.get('hierarchy_xlsx'):
                        st.success("✅ フォルダ階層リスト取得済み")
                    else:
                        st.warning("⬜ フォルダ階層リスト未取得")
                with col_s2:
                    folders = st.session_state.workflow_data.get('rakuten_folders')
                    if folders:
                        st.success(f"✅ 楽天フォルダ: {len(folders)}件")
                    else:
                        st.warning("⬜ 楽天フォルダ未取得")

                # データ確認エリア
                with st.expander("データ確認（中身を見る）", expanded=False):
                    # 1. 画像に紐づく情報（merge_csv_data結果）
                    st.markdown("**画像の出版社・シリーズ情報**")
                    img_info_df = pd.DataFrame([
                        {'コミックNo': img['comic_no'], 'ジャンル': img.get('genre', ''), '出版社': img.get('publisher', ''), 'シリーズ': img.get('series', ''), 'タイトル': img.get('title', '')}
                        for img in images
                    ])
                    st.dataframe(img_info_df, use_container_width=True, height=150)

                    # 2. フォルダ階層リスト
                    if st.session_state.workflow_data.get('hierarchy_xlsx'):
                        st.markdown("**フォルダ階層リスト**")
                        try:
                            h_df = pd.read_excel(BytesIO(st.session_state.workflow_data['hierarchy_xlsx']), sheet_name="フォルダ階層リスト", header=None)
                            col_names = ['ジャンル', '出版社', 'シリーズ', 'メインフォルダ', 'サブフォルダ'] + [f'列{i}' for i in range(6, 20)]
                            h_df.columns = col_names[:len(h_df.columns)]
                            st.dataframe(h_df, use_container_width=True, height=200)
                        except Exception as e:
                            st.error(f"表示エラー: {e}")

                    # 3. 楽天フォルダ一覧
                    if st.session_state.workflow_data.get('rakuten_folders'):
                        st.markdown("**楽天フォルダ一覧（R-Cabinet API）**")
                        f_df = pd.DataFrame(st.session_state.workflow_data['rakuten_folders'])
                        st.dataframe(f_df, use_container_width=True, height=200)

                # キュー生成
                st.divider()
                hierarchy_ready = st.session_state.workflow_data.get('hierarchy_xlsx')
                folders_ready = st.session_state.workflow_data.get('rakuten_folders')

                if st.button("🏪 アップロードキュー生成", type="primary", disabled=not (hierarchy_ready and folders_ready), key="rakuten_queue_btn"):
                    hierarchy_bytes = st.session_state.workflow_data['hierarchy_xlsx']
                    hierarchy_df = pd.read_excel(BytesIO(hierarchy_bytes), sheet_name="フォルダ階層リスト", header=None)
                    folders = st.session_state.workflow_data['rakuten_folders']

                    result = prepare_rakuten_queue(images, hierarchy_df, folders)
                    st.session_state.workflow_data['rakuten_queue'] = result
                    st.rerun()

                # キュー結果表示
                if 'rakuten_queue' in st.session_state.workflow_data:
                    result = st.session_state.workflow_data['rakuten_queue']

                    col_m1, col_m2 = st.columns(2)
                    col_m1.metric("マッピング成功", result['mapped'])
                    col_m2.metric("未マッチ", len(result['unmapped']))

                    if result['queue']:
                        queue_df = pd.DataFrame([
                            {'コミックNo': q['comic_no'], 'メインフォルダ': q.get('main_folder', ''), 'サブフォルダ': q['folder_name'], 'FolderId': q['folder_id']}
                            for q in result['queue']
                        ])
                        st.dataframe(queue_df, use_container_width=True, height=200)

                    with st.expander("ログ", expanded=False):
                        for log in result['logs']:
                            st.text(log)

            # ナビゲーション
            st.divider()
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("← 戻る"):
                    st.session_state.workflow_step = 3
                    st.rerun()
            with col2:
                yahoo_ready = 'yahoo_zips' in st.session_state.workflow_data
                rakuten_ready = 'rakuten_queue' in st.session_state.workflow_data
                if st.button("次へ進む →", type="primary", disabled=not (yahoo_ready or rakuten_ready)):
                    if 4 not in st.session_state.workflow_completed:
                        st.session_state.workflow_completed.append(4)
                    st.session_state.workflow_step = 5
                    st.rerun()

    # ============================================================
    # Step 5: アップロード
    # ============================================================
    elif current_step == 5:
        st.markdown("""
        <div class="step-card">
            <div class="step-card-header">
                <div class="step-card-icon">🚀</div>
                <div>
                    <p class="step-card-title">Step ⑤ アップロード</p>
                    <p class="step-card-desc">楽天・ヤフーに画像をアップロードします（手動 / API）</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        tab_manual, tab_api = st.tabs(["📁 手動アップロード", "🤖 APIアップロード"])

        with tab_manual:
            st.markdown("### 手動アップロード用ZIP")
            st.caption("フォルダ構成どおりに画像を振り分けたZIPをダウンロードし、R-Cabinetに手動でアップロードします。")

            rakuten_queue = st.session_state.workflow_data.get('rakuten_queue')
            yahoo_zips = st.session_state.workflow_data.get('yahoo_zips')

            # 楽天ZIP生成
            st.markdown("#### 楽天（R-Cabinet）")
            if rakuten_queue and rakuten_queue.get('queue'):
                queue = rakuten_queue['queue']
                st.info(f"対象: {len(queue)}件（{len(rakuten_queue.get('unmapped', []))}件未マッチ）")

                if st.button("📦 楽天用ZIP生成", type="primary", key="rakuten_manual_zip"):
                    import zipfile
                    import openpyxl
                    zip_buffer = BytesIO()
                    mapping_rows = []
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for q in queue:
                            main_folder = q.get('main_folder', '')
                            sub_folder = q.get('folder_name', '')
                            file_name = q['file_name']
                            # sub_folderとmain_folderが同じ場合はサブフォルダなし
                            if sub_folder and sub_folder != main_folder:
                                zip_path = f"{main_folder}/{sub_folder}/{file_name}"
                            elif main_folder:
                                zip_path = f"{main_folder}/{file_name}"
                            else:
                                zip_path = f"{sub_folder}/{file_name}"
                            zf.writestr(zip_path, q['file_bytes'])
                            mapping_rows.append({
                                'コミックNo': q['comic_no'],
                                'メインフォルダ': main_folder,
                                'サブフォルダ': sub_folder if sub_folder != main_folder else '',
                                'ファイル名': file_name,
                                'ZIPパス': zip_path
                            })

                        # マッピング一覧Excelを生成してZIP直下に配置
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "振り分け一覧"
                        font = openpyxl.styles.Font(name='Meiryo UI')
                        headers = ['No', 'コミックNo', 'メインフォルダ', 'サブフォルダ', 'ファイル名', 'ZIPパス']
                        ws.append(headers)
                        for idx, row in enumerate(mapping_rows, 1):
                            ws.append([idx, row['コミックNo'], row['メインフォルダ'], row['サブフォルダ'], row['ファイル名'], row['ZIPパス']])
                        # フォント適用・列幅調整（全角文字は幅2として計算）
                        def calc_display_width(text):
                            return sum(2 if ord(c) > 0x7F else 1 for c in str(text))

                        for col_idx, header in enumerate(headers, 1):
                            max_w = calc_display_width(header)
                            for r in range(1, ws.max_row + 1):
                                cell = ws.cell(row=r, column=col_idx)
                                cell.font = font
                                max_w = max(max_w, calc_display_width(cell.value or ''))
                            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_w + 2, 80)
                        xlsx_buffer = BytesIO()
                        wb.save(xlsx_buffer)
                        zf.writestr("振り分け一覧.xlsx", xlsx_buffer.getvalue())

                    st.session_state.workflow_data['rakuten_manual_zip'] = zip_buffer.getvalue()
                    st.rerun()

                if st.session_state.workflow_data.get('rakuten_manual_zip'):
                    zip_data = st.session_state.workflow_data['rakuten_manual_zip']
                    size_mb = len(zip_data) / (1024 * 1024)
                    st.download_button(
                        label=f"📥 rakuten_upload.zip ({size_mb:.1f}MB)",
                        data=zip_data,
                        file_name="rakuten_upload.zip",
                        mime="application/zip",
                        key="rakuten_manual_zip_dl"
                    )
            else:
                st.warning("Step ④でアップロードキューを生成してください。")

            # ヤフーZIP
            st.divider()
            st.markdown("#### ヤフー")
            if yahoo_zips and yahoo_zips.get('zips'):
                zips = yahoo_zips['zips']
                st.info(f"マッピング成功: {yahoo_zips['mapped']}件 / ZIP数: {len(zips)}")
                for i, zip_data in enumerate(zips):
                    size_mb = len(zip_data) / (1024 * 1024)
                    st.download_button(
                        label=f"📥 yahoo_upload_{i+1:03d}.zip ({size_mb:.1f}MB)",
                        data=zip_data,
                        file_name=f"yahoo_upload_{i+1:03d}.zip",
                        mime="application/zip",
                        key=f"yahoo_manual_zip_dl_{i}"
                    )
            else:
                st.warning("Step ④でヤフー用ZIPを生成してください。")

        with tab_api:
            st.markdown("### APIアップロード")
            st.info("🚧 この機能は現在実装中です。")

            tab_api_yahoo, tab_api_rakuten = st.tabs(["🛒 ヤフー", "🏪 楽天"])

            with tab_api_yahoo:
                st.warning("Yahoo認証トークンが未設定です。設定後にアップロードが可能になります。")

            with tab_api_rakuten:
                st.success("楽天API認証は設定済みです。実装予定。")

        st.divider()
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("← 戻る"):
                st.session_state.workflow_step = 4
                st.rerun()
        with col2:
            if st.button("ワークフローをリセット", type="secondary"):
                st.session_state.workflow_step = 1
                st.session_state.workflow_completed = []
                st.session_state.workflow_data = {}
                st.rerun()


elif mode == "🛰️ R-Cabi構成把握":
    st.title("🛰️ R-Cabi構成把握")
    st.markdown("R-Cabinetの構成を一覧化します。")

    # セッション状態の初期化
    if "folders_loaded" not in st.session_state:
        st.session_state.folders_loaded = False
        st.session_state.folders_data = None
        st.session_state.folders_error = None
    if "images_loaded" not in st.session_state:
        st.session_state.images_loaded = False
        st.session_state.images_data = None

    # フォルダ一覧を自動取得（初回のみ）- 再実行を挟まず同一runで続行し、残像を回避
    if not st.session_state.folders_loaded:
        with st.spinner("フォルダ一覧を取得中..."):
            folders, error = get_all_folders()
        st.session_state.folders_data = folders
        st.session_state.folders_error = error
        st.session_state.folders_loaded = True
    else:
        folders = st.session_state.folders_data
        error = st.session_state.folders_error

    if error:
        st.error(error)
        if st.button("🔄 再試行"):
            st.session_state.folders_loaded = False
            st.cache_data.clear()
            st.rerun()
        st.stop()

    if not folders:
        st.warning("フォルダがありません。")
        st.stop()

    # 総ファイル数を計算
    total_files = sum(f['FileCount'] for f in folders)

    # サイドバーにフォルダ情報（再取得ボタンは最新一覧を取得に統合したため不要）
    with st.sidebar:
        st.success(f"📁 {len(folders)} フォルダ")
        st.info(f"📷 {total_files} 画像（全体）")

    xlsx_col1, xlsx_col2, xlsx_col3 = st.columns([3, 3, 4])
    with xlsx_col1:
        xlsx_latest_btn = st.button(
            "🔄 最新をAPIから取得してダウンロード",
            help="R-Cabinet APIから最新のフォルダ構成・画像一覧を取得してExcel生成（時間がかかります）",
            key="xlsx_latest_btn",
        )
    with xlsx_col2:
        xlsx_db_btn = st.button(
            "📋 前回実行時の状態をダウンロード",
            help="DBに保存済みのデータ（前回同期時点のスナップショット）からExcel生成",
            key="xlsx_db_btn",
        )

    if xlsx_latest_btn:
        with st.spinner("フォルダ一覧を取得中..."):
            latest_folders, f_err = get_all_folders()
        if f_err or not latest_folders:
            st.error(f"フォルダ取得エラー: {f_err or 'データなし'}")
        else:
            target_prefixes = [prefix for prefix, _ in FOLDER_MANAGEMENT_SHEETS]
            # カテゴリ名ルックアップ用に親フォルダ（例: /comic）も対象に含める
            needed_ancestors = set()
            for prefix in target_prefixes:
                parts = [p for p in prefix.strip('/').split('/') if p]
                for i in range(1, len(parts)):
                    needed_ancestors.add('/' + '/'.join(parts[:i]))

            def _is_target_folder(fp: str) -> bool:
                if not fp:
                    return False
                if fp in needed_ancestors:
                    return True
                return any(fp == p or fp.startswith(p + '/') for p in target_prefixes)

            target_folders = [f for f in latest_folders if _is_target_folder(f.get('FolderPath', ''))]
            skipped_count = len(latest_folders) - len(target_folders)

            if not target_folders:
                st.error("対象フォルダが見つかりません（FOLDER_MANAGEMENT_SHEETSのプレフィックス配下が0件）")
            else:
                st.info(f"対象フォルダに絞り込み: {len(target_folders)}件（対象外{skipped_count}件はスキップ）")
                all_files_for_xlsx = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                for i, folder in enumerate(target_folders):
                    status_text.text(f"取得中: {folder['FolderName']} ({i + 1}/{len(target_folders)})")
                    progress_bar.progress((i + 1) / len(target_folders))
                    files, _ = get_folder_files(int(folder['FolderId']))
                    if files:
                        for fl in files:
                            fl['FolderName'] = folder['FolderName']
                            fl['FolderPath'] = folder.get('FolderPath', '')
                        all_files_for_xlsx.extend(files)
                    time.sleep(0.3)
                progress_bar.empty()
                status_text.empty()
                with st.spinner("Excel生成中..."):
                    xlsx_bytes = build_folder_management_xlsx(target_folders, all_files_for_xlsx)
                st.session_state.xlsx_latest_bytes = xlsx_bytes
                st.session_state.xlsx_latest_count = len(all_files_for_xlsx)
                st.success(f"✅ 最新版Excel生成完了（{len(target_folders)}フォルダ / {len(all_files_for_xlsx)}ファイル）")

    if xlsx_db_btn:
        with st.spinner("DBから読込中..."):
            db_images, _msg = load_images_from_db()
        if not db_images:
            st.warning("DBにデータがありません。日次同期が完了してから再実行してください。")
        else:
            with st.spinner("Excel生成中..."):
                xlsx_bytes = build_folder_management_xlsx(folders, db_images)
            st.session_state.xlsx_db_bytes = xlsx_bytes
            st.session_state.xlsx_db_count = len(db_images)
            st.success(f"✅ 前回データExcel生成完了（{len(folders)}フォルダ / {len(db_images)}ファイル）")

    if st.session_state.get("xlsx_latest_bytes"):
        st.download_button(
            label=f"⬇️ 最新版をダウンロード（{st.session_state.xlsx_latest_count}件）",
            data=st.session_state.xlsx_latest_bytes,
            file_name=f"楽天RMS画像フォルダ管理シート_最新_{datetime.now(JST).strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="xlsx_latest_dl",
        )
    if st.session_state.get("xlsx_db_bytes"):
        st.download_button(
            label=f"⬇️ 前回データをダウンロード（{st.session_state.xlsx_db_count}件）",
            data=st.session_state.xlsx_db_bytes,
            file_name=f"楽天RMS画像フォルダ管理シート_前回_{datetime.now(JST).strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="xlsx_db_dl",
        )


# ============================================================
# フォルダ一括作成
# ============================================================
elif mode == "🏗️ R-Cabiフォルダ制作":
    st.header("🏗️ R-Cabiフォルダ制作")
    st.markdown("R-Cabinetに複数のフォルダをまとめて作成します。パス形式で階層構造を指定でき、上位フォルダのIDは自動で解決されます。")

    st.divider()

    # CSV入力説明
    st.subheader("CSV入力")
    st.caption("カンマ区切りで「フォルダパス, ディレクトリ名」を入力")
    st.markdown("""
| 列 | 項目 | 必須 | 説明 |
|---|---|---|---|
| 1列目 | **フォルダパス** | ○ | `/`区切りで階層を表現。最下層がフォルダ名（最大50バイト） |
| 2列目 | ディレクトリ名 | | a-z, 0-9, -, _ のみ。最大20文字。省略時は自動採番 |

- 上位フォルダは**パスから自動判定**されます（IDの手動指定は不要）
- 上位フォルダが同じCSV内にあれば、作成後のIDを自動で引き継ぎます
- 上位フォルダがR-Cabinetに既存の場合も自動で検出します
""")

    # テンプレートCSVダウンロード
    template_csv = "フォルダパス,ディレクトリ名\nコミック,comic\nコミック/セット,set\nコミック/セット/セット1,set1\nコミック/セット/セット2,set2\nコミック/セット/セット3,set3\nコミック/単品,tanpin\nコミック/単品/単品1,tanpin1\nコミック/単品/単品2,tanpin2\nコミック/単品/単品3,tanpin3\nコミック/予約,yoyaku\nコミック/予約/予約1,yoyaku1\nコミック/予約/予約2,yoyaku2\nコミック/予約/予約3,yoyaku3"
    st.download_button(
        label="📥 テンプレートCSVをダウンロード",
        data=template_csv.encode('utf-8-sig'),
        file_name="folder_template.csv",
        mime="text/csv"
    )

    st.markdown("")

    csv_input = st.text_area(
        "CSV入力",
        height=250,
        placeholder="例:\nコミック,comic\nコミック/セット,set\nコミック/セット/セット1,set1\nコミック/セット/セット2,set2\nコミック/単品,tanpin\nコミック/単品/単品1,tanpin1",
        label_visibility="collapsed"
    )

    # CSVファイルアップロード
    uploaded_csv = st.file_uploader("またはCSVファイルをアップロード", type=['csv'])

    # 入力ソースの決定
    raw_lines = []
    if uploaded_csv:
        try:
            content = uploaded_csv.read().decode('utf-8')
        except UnicodeDecodeError:
            uploaded_csv.seek(0)
            content = uploaded_csv.read().decode('cp932')
        raw_lines = content.strip().splitlines()
    elif csv_input:
        raw_lines = csv_input.strip().splitlines()

    # ヘッダー行をスキップ
    if raw_lines and ("フォルダパス" in raw_lines[0] or "フォルダ名" in raw_lines[0]):
        raw_lines = raw_lines[1:]

    # パース
    import re as _re
    folder_entries = []
    for line in raw_lines:
        if not line.strip():
            continue
        parts = [p.strip() for p in line.split(",")]
        path = parts[0] if len(parts) > 0 else ""
        directory = parts[1] if len(parts) > 1 and parts[1] else None
        if path:
            # パスから階層を分解
            path_parts = [p.strip() for p in path.split("/") if p.strip()]
            folder_name = path_parts[-1]  # 最下層がフォルダ名
            parent_path = "/".join(path_parts[:-1]) if len(path_parts) > 1 else None
            folder_entries.append({
                "path": "/".join(path_parts),
                "name": folder_name,
                "directory": directory,
                "parent_path": parent_path
            })

    # プレビュー表示
    if folder_entries:
        st.divider()
        st.subheader(f"作成予定: {len(folder_entries)} フォルダ")

        preview_data = []
        for i, entry in enumerate(folder_entries, 1):
            # 階層の深さに応じてインデント
            depth = entry["path"].count("/")
            indent = "　" * depth
            preview_data.append({
                "No": i,
                "フォルダ構造": f"{indent}{entry['name']}",
                "パス": entry["path"],
                "ディレクトリ名": entry["directory"] or "（自動採番）",
                "上位フォルダ": entry["parent_path"] or "（ルート）"
            })

        st.dataframe(
            pd.DataFrame(preview_data),
            use_container_width=True,
            hide_index=True
        )

        # 既存フォルダ一覧を取得（バリデーション用）
        with st.spinner("既存フォルダを確認中..."):
            existing_folders, folder_list_error = get_all_folders()
        existing_name_set = set()
        if existing_folders:
            for f in existing_folders:
                existing_name_set.add(f['FolderName'])

        # バリデーション
        errors = []
        path_set = {e["path"] for e in folder_entries}
        for i, entry in enumerate(folder_entries, 1):
            if len(entry["name"].encode('utf-8')) > 50:
                errors.append(f"No.{i}「{entry['name']}」: フォルダ名が50バイトを超えています")
            if entry["directory"]:
                if len(entry["directory"]) > 20:
                    errors.append(f"No.{i}「{entry['directory']}」: ディレクトリ名が20文字を超えています")
                elif not _re.fullmatch(r'[a-z0-9_-]+', entry["directory"]):
                    errors.append(f"No.{i}「{entry['directory']}」: ディレクトリ名に使用不可の文字（a-z, 0-9, -, _ のみ）")
            if entry["parent_path"] and entry["parent_path"] not in path_set:
                # 既存フォルダからも検索（フォルダ名で一致）
                parent_name = entry["parent_path"].split("/")[-1]
                if parent_name not in existing_name_set:
                    errors.append(f"No.{i}「{entry['path']}」: 上位フォルダ「{entry['parent_path']}」が見つかりません")

        if errors:
            for err in errors:
                st.error(err)
        else:
            st.divider()
            if st.button("🚀 一括作成を実行", type="primary"):
                # 既存フォルダ一覧を取得（既存フォルダのID解決用）
                existing_folders, folder_list_error = get_all_folders()
                existing_name_to_id = {}
                if existing_folders:
                    for f in existing_folders:
                        existing_name_to_id[f['FolderName']] = f['FolderId']

                # パス→フォルダIDのマッピング（作成済みフォルダのID記録用）
                path_to_id = {}

                results = []
                progress = st.progress(0, text="フォルダ作成中...")
                total = len(folder_entries)

                for i, entry in enumerate(folder_entries):
                    progress.progress((i + 1) / total, text=f"作成中... ({i + 1}/{total}) {entry['name']}")

                    # 上位フォルダIDを解決
                    upper_folder_id = None
                    if entry["parent_path"]:
                        if entry["parent_path"] in path_to_id:
                            # 同じCSV内で先に作成済み
                            upper_folder_id = path_to_id[entry["parent_path"]]
                        else:
                            # 既存フォルダから検索（フォルダ名で一致）
                            parent_name = entry["parent_path"].split("/")[-1]
                            if parent_name in existing_name_to_id:
                                upper_folder_id = existing_name_to_id[parent_name]

                        if not upper_folder_id:
                            results.append({
                                "フォルダ構造": entry["path"],
                                "フォルダ名": entry["name"],
                                "ディレクトリ名": entry["directory"] or "（自動）",
                                "結果": "❌ スキップ",
                                "フォルダID": "",
                                "エラー": f"上位フォルダ「{entry['parent_path']}」のIDが見つかりません"
                            })
                            continue

                    result = create_folder(
                        folder_name=entry["name"],
                        directory_name=entry["directory"],
                        upper_folder_id=upper_folder_id
                    )

                    if result["success"]:
                        path_to_id[entry["path"]] = result["folder_id"]

                    results.append({
                        "フォルダ構造": entry["path"],
                        "フォルダ名": entry["name"],
                        "ディレクトリ名": entry["directory"] or "（自動）",
                        "結果": "✅ 成功" if result["success"] else "❌ 失敗",
                        "フォルダID": result.get("folder_id", ""),
                        "エラー": result.get("error", "")
                    })
                    if i < total - 1:
                        time.sleep(0.5)  # API負荷軽減

                progress.empty()

                # 結果表示
                success_count = sum(1 for r in results if r["結果"] == "✅ 成功")
                fail_count = total - success_count

                if fail_count == 0:
                    st.success(f"全 {total} フォルダの作成が完了しました！")
                elif success_count == 0:
                    st.error(f"全 {total} フォルダの作成に失敗しました")
                else:
                    st.warning(f"成功: {success_count} / 失敗: {fail_count}")

                st.dataframe(
                    pd.DataFrame(results),
                    use_container_width=True,
                    hide_index=True
                )

# ============================
# 🔁 コピー：R-Cabi⇒R-Cabi
# ============================
elif mode == "🔁 コピー：R-Cabi⇒R-Cabi":
    st.markdown("## 🔁 コピー：R-Cabi⇒R-Cabi")
    st.markdown("CSVで指定した既存画像を、指定フォルダにコピーします。")

    # Excelテンプレートダウンロード（フォルダ一覧シート付き）
    if st.button("📥 テンプレートをダウンロード（フォルダ一覧付き）"):
        with st.spinner("フォルダ一覧を取得中..."):
            tmpl_folders, tmpl_error = get_all_folders()

        if tmpl_error:
            st.error(f"フォルダ一覧の取得に失敗しました: {tmpl_error}")
        else:
            styles, _ = get_openpyxl_styles()
            Font = styles['Font']
            PatternFill = styles['PatternFill']
            Alignment = styles['Alignment']
            from openpyxl import Workbook

            wb = Workbook()

            # シート1: テンプレート
            ws1 = wb.active
            ws1.title = "コピー指示"
            headers1 = ["フォルダ", "カテゴリ1", "カテゴリ2", "カテゴリ3", "ファイル名", "URL", "フォルダID"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
            cell_font = Font(name="Meiryo UI", size=10)

            for col_idx, header in enumerate(headers1, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # サンプル行
            sample = ["商品画像文庫", "コミック", "セット", "セット1", "10000",
                       "https://image.rakuten.co.jp/haru-uraraka/cabinet/shohinbunko/10000.jpg", "13054250"]
            for col_idx, val in enumerate(sample, 1):
                cell = ws1.cell(row=2, column=col_idx, value=val)
                cell.font = cell_font

            # 列幅調整
            ws1.column_dimensions['A'].width = 18
            ws1.column_dimensions['B'].width = 14
            ws1.column_dimensions['C'].width = 14
            ws1.column_dimensions['D'].width = 14
            ws1.column_dimensions['E'].width = 16
            ws1.column_dimensions['F'].width = 60
            ws1.column_dimensions['G'].width = 14

            # シート2: フォルダ一覧
            ws2 = wb.create_sheet(title="フォルダ一覧")
            headers2 = ["フォルダID", "フォルダ名", "ディレクトリパス", "ファイル数"]
            for col_idx, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, f in enumerate(tmpl_folders, 2):
                ws2.cell(row=row_idx, column=1, value=f["FolderId"]).font = cell_font
                ws2.cell(row=row_idx, column=2, value=f["FolderName"]).font = cell_font
                ws2.cell(row=row_idx, column=3, value=f["FolderPath"]).font = cell_font
                ws2.cell(row=row_idx, column=4, value=f["FileCount"]).font = cell_font

            ws2.column_dimensions['A'].width = 14
            ws2.column_dimensions['B'].width = 30
            ws2.column_dimensions['C'].width = 40
            ws2.column_dimensions['D'].width = 12

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.download_button(
                label="📥 ダウンロード",
                data=excel_buffer.getvalue(),
                file_name="image_copy_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_template_xlsx",
            )

    st.divider()

    # ファイルアップロード（CSV/Excel対応）
    upload_file = st.file_uploader("CSV / Excelファイルを選択", type=["csv", "xlsx"], key="csv_image_copy")

    if upload_file:
        # ファイル読み込み
        if upload_file.name.endswith(".xlsx"):
            upload_file.seek(0)
            df = pd.read_excel(upload_file, sheet_name=0, dtype=str).fillna("")
        else:
            try:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="utf-8-sig", dtype=str).fillna("")
            except Exception:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="cp932", dtype=str).fillna("")

        required_cols = ["ファイル名", "URL"]
        missing_cols = [c for c in required_cols if c not in df.columns]
        has_folder_id = "フォルダID" in df.columns
        has_categories = all(c in df.columns for c in ["カテゴリ1", "カテゴリ2", "カテゴリ3"])

        if missing_cols:
            st.error(f"必須列が不足しています: {', '.join(missing_cols)}")
        elif not has_folder_id and not has_categories:
            st.error("「フォルダID」列 または 「カテゴリ1〜3」列が必要です")
        elif len(df) == 0:
            st.warning("データがありません。")
        else:
            st.info(f"📎 {len(df)} 件のデータを読み込みました")

            # フォルダ一覧を取得（名前マッチング用）
            with st.spinner("フォルダ一覧を取得中..."):
                folders, folder_error = get_all_folders()

            if folder_error:
                st.error(f"フォルダ一覧の取得に失敗しました: {folder_error}")
            else:
                folder_name_to_id = {}
                folder_id_set = set()
                for f in folders:
                    folder_name_to_id[f["FolderName"]] = f["FolderId"]
                    folder_id_set.add(str(f["FolderId"]))

                # フォルダIDを解決
                preview_data = []
                for idx, row in df.iterrows():
                    cat_path = ""
                    if has_categories:
                        cat_path = "/".join([
                            row.get("カテゴリ1", ""),
                            row.get("カテゴリ2", ""),
                            row.get("カテゴリ3", ""),
                        ]).strip("/")

                    # フォルダID解決: フォルダID列を優先、なければ名前マッチング
                    folder_id = ""
                    target_folder_name = ""
                    if has_folder_id and row.get("フォルダID", "").strip():
                        fid = row["フォルダID"].strip()
                        if fid in folder_id_set:
                            folder_id = fid
                            status = "✅ OK"
                            # フォルダ名を逆引き
                            for f in folders:
                                if str(f["FolderId"]) == fid:
                                    target_folder_name = f["FolderName"]
                                    break
                        else:
                            status = "❌ フォルダID不正"
                    elif has_categories:
                        target_folder_name = row.get("カテゴリ3", "").strip()
                        if not target_folder_name:
                            target_folder_name = row.get("カテゴリ2", "").strip()
                        if not target_folder_name:
                            target_folder_name = row.get("カテゴリ1", "").strip()
                        folder_id = folder_name_to_id.get(target_folder_name, "")
                        status = "✅ OK" if folder_id else "❌ フォルダ未検出"
                    else:
                        status = "❌ フォルダ指定なし"

                    # URL検証
                    url = row["URL"].strip()
                    if not url:
                        status = "❌ URL未入力"
                    elif not url.startswith("http"):
                        status = "❌ URL不正"

                    file_name = row["ファイル名"].strip()
                    if not file_name:
                        status = "❌ ファイル名未入力"

                    preview_data.append({
                        "No": idx + 1,
                        "コピー先": f"{target_folder_name}（ID: {folder_id}）" if folder_id else target_folder_name,
                        "カテゴリパス": cat_path,
                        "ファイル名": file_name,
                        "URL": url,
                        "チェック": status,
                        "_folder_id": folder_id,
                        "_url": url,
                        "_file_name": file_name,
                    })

                # プレビュー表示
                preview_df = pd.DataFrame(preview_data)
                display_df = preview_df[["No", "コピー先", "カテゴリパス", "ファイル名", "URL", "チェック"]]

                ok_count = sum(1 for d in preview_data if d["チェック"] == "✅ OK")
                ng_count = len(preview_data) - ok_count

                if ng_count > 0:
                    st.warning(f"実行可能: {ok_count} 件 / エラー: {ng_count} 件（エラー行はスキップされます）")
                else:
                    st.success(f"全 {ok_count} 件 実行可能です")

                st.dataframe(display_df, use_container_width=True, hide_index=True)

                # 実行範囲（バッチ処理）
                st.divider()
                st.markdown("### 実行設定")
                if ok_count > 500:
                    st.info("⏱ Streamlit Cloudのタイムアウト対策のため、500件ずつの実行を推奨します")
                col_start, col_end = st.columns(2)
                with col_start:
                    batch_start = st.number_input("開始（No）", min_value=1, max_value=ok_count, value=1, step=1, key="batch_start")
                with col_end:
                    batch_end_default = min(500, ok_count)
                    batch_end = st.number_input("終了（No）", min_value=1, max_value=ok_count, value=batch_end_default, step=1, key="batch_end")

                batch_size = batch_end - batch_start + 1
                st.markdown(f"**実行対象: {batch_start}〜{batch_end} 件目（{batch_size} 件）**")

                # 上書きオプション
                overwrite = st.checkbox("同名ファイルが存在する場合は上書きする", value=False, key="csv_copy_overwrite")

                # 実行ボタン
                if ok_count > 0 and st.button(f"📋 コピー実行（{batch_start}〜{batch_end}件目）", type="primary"):
                    progress = st.progress(0, text="コピー中...")
                    results = []
                    all_ok_rows = [d for d in preview_data if d["チェック"] == "✅ OK"]
                    target_rows = all_ok_rows[batch_start - 1:batch_end]
                    total = len(target_rows)

                    for i, row_data in enumerate(target_rows):
                        progress.progress((i + 1) / total, text=f"コピー中... ({i + 1}/{total}) {row_data['_file_name']}")

                        # 画像をダウンロード
                        try:
                            img_response = requests.get(row_data["_url"], timeout=30)
                            if img_response.status_code != 200:
                                results.append({
                                    "No": row_data["No"],
                                    "ファイル名": row_data["_file_name"],
                                    "コピー先": row_data["コピー先"],
                                    "結果": "❌ 失敗",
                                    "エラー": f"画像ダウンロード失敗: HTTP {img_response.status_code}",
                                })
                                continue
                            file_data = img_response.content
                        except requests.exceptions.RequestException as e:
                            results.append({
                                "No": row_data["No"],
                                "ファイル名": row_data["_file_name"],
                                "コピー先": row_data["コピー先"],
                                "結果": "❌ 失敗",
                                "エラー": f"画像ダウンロード失敗: {str(e)}",
                            })
                            continue

                        # ファイル名の拡張子を元URLから取得
                        url_path = row_data["_url"].rsplit("/", 1)[-1]
                        ext = url_path.rsplit(".", 1)[-1] if "." in url_path else "jpg"

                        # fileName（画像名）: 拡張子なし
                        api_file_name = row_data["_file_name"]

                        # filePath（URLのファイル名）: 拡張子付き、20バイト制限
                        file_path_name = f"{row_data['_file_name']}.{ext}"
                        if len(file_path_name.encode("utf-8")) > 20:
                            stem = row_data["_file_name"]
                            while len(f"{stem}.{ext}".encode("utf-8")) > 20 and stem:
                                stem = stem[:-1]
                            file_path_name = f"{stem}.{ext}"

                        # アップロード
                        result = upload_image(
                            file_data=file_data,
                            file_name=api_file_name,
                            folder_id=row_data["_folder_id"],
                            file_path_name=file_path_name,
                            overwrite=overwrite,
                        )

                        results.append({
                            "No": row_data["No"],
                            "ファイル名": row_data["_file_name"],
                            "コピー先": row_data["コピー先"],
                            "結果": "✅ 成功" if result["success"] else "❌ 失敗",
                            "エラー": result.get("error", ""),
                        })

                        # API負荷軽減（3 req/sec制限）
                        if i < total - 1:
                            time.sleep(0.35)

                    progress.empty()

                    # 結果表示
                    success_count = sum(1 for r in results if r["結果"] == "✅ 成功")
                    fail_count = total - success_count

                    if fail_count == 0:
                        st.success(f"全 {total} 件のコピーが完了しました！")
                    elif success_count == 0:
                        st.error(f"全 {total} 件のコピーに失敗しました")
                    else:
                        st.warning(f"成功: {success_count} / 失敗: {fail_count}")

                    st.dataframe(
                        pd.DataFrame(results),
                        use_container_width=True,
                        hide_index=True
                    )
# ============================
# ☁️ コピー：Local⇒R-Cabi
# ============================
elif mode == "☁️ コピー：Local⇒R-Cabi":
    st.markdown("## ☁️ コピー：Local⇒R-Cabi")
    st.markdown("CSVで指定したローカル画像ファイルを、指定フォルダにアップロードします。")

    # Excelテンプレートダウンロード（フォルダ一覧シート付き）
    if st.button("📥 テンプレートをダウンロード（フォルダ一覧付き）", key="local_tmpl_btn"):
        with st.spinner("フォルダ一覧を取得中..."):
            tmpl_folders, tmpl_error = get_all_folders()

        if tmpl_error:
            st.error(f"フォルダ一覧の取得に失敗しました: {tmpl_error}")
        else:
            styles, _ = get_openpyxl_styles()
            Font = styles['Font']
            PatternFill = styles['PatternFill']
            Alignment = styles['Alignment']
            from openpyxl import Workbook

            wb = Workbook()

            # シート1: テンプレート
            ws1 = wb.active
            ws1.title = "アップロード指示"
            headers1 = ["フォルダID", "カテゴリ1", "カテゴリ2", "カテゴリ3", "ファイル名", "ファイルパス"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
            cell_font = Font(name="Meiryo UI", size=10)

            for col_idx, header in enumerate(headers1, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # サンプル行
            sample = ["13054250", "コミック", "セット", "セット1", "10000",
                       r"C:\Users\ssasa\Pictures\10000.jpg"]
            for col_idx, val in enumerate(sample, 1):
                cell = ws1.cell(row=2, column=col_idx, value=val)
                cell.font = cell_font

            ws1.column_dimensions['A'].width = 14
            ws1.column_dimensions['B'].width = 14
            ws1.column_dimensions['C'].width = 14
            ws1.column_dimensions['D'].width = 14
            ws1.column_dimensions['E'].width = 16
            ws1.column_dimensions['F'].width = 60

            # シート2: フォルダ一覧
            ws2 = wb.create_sheet(title="フォルダ一覧")
            headers2 = ["フォルダID", "フォルダ名", "ディレクトリパス", "ファイル数"]
            for col_idx, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, f in enumerate(tmpl_folders, 2):
                ws2.cell(row=row_idx, column=1, value=f["FolderId"]).font = cell_font
                ws2.cell(row=row_idx, column=2, value=f["FolderName"]).font = cell_font
                ws2.cell(row=row_idx, column=3, value=f["FolderPath"]).font = cell_font
                ws2.cell(row=row_idx, column=4, value=f["FileCount"]).font = cell_font

            ws2.column_dimensions['A'].width = 14
            ws2.column_dimensions['B'].width = 30
            ws2.column_dimensions['C'].width = 40
            ws2.column_dimensions['D'].width = 12

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.download_button(
                label="📥 ダウンロード",
                data=excel_buffer.getvalue(),
                file_name="local_upload_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_local_template_xlsx",
            )

    st.divider()

    st.warning("⚠ この機能はローカル実行時のみ使用できます（Streamlit Cloudではファイルパスにアクセスできません）")

    # ファイルアップロード（CSV/Excel対応）
    upload_file = st.file_uploader("CSV / Excelファイルを選択", type=["csv", "xlsx"], key="local_image_upload")

    if upload_file:
        # ファイル読み込み
        if upload_file.name.endswith(".xlsx"):
            upload_file.seek(0)
            df = pd.read_excel(upload_file, sheet_name=0, dtype=str).fillna("")
        else:
            try:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="utf-8-sig", dtype=str).fillna("")
            except Exception:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="cp932", dtype=str).fillna("")

        required_cols = ["ファイル名", "ファイルパス"]
        missing_cols = [c for c in required_cols if c not in df.columns]
        has_folder_id = "フォルダID" in df.columns
        has_categories = all(c in df.columns for c in ["カテゴリ1", "カテゴリ2", "カテゴリ3"])

        if missing_cols:
            st.error(f"必須列が不足しています: {', '.join(missing_cols)}")
        elif not has_folder_id and not has_categories:
            st.error("「フォルダID」列 または 「カテゴリ1〜3」列が必要です")
        elif len(df) == 0:
            st.warning("データがありません。")
        else:
            st.info(f"📎 {len(df)} 件のデータを読み込みました")

            # フォルダ一覧を取得
            with st.spinner("フォルダ一覧を取得中..."):
                folders, folder_error = get_all_folders()

            if folder_error:
                st.error(f"フォルダ一覧の取得に失敗しました: {folder_error}")
            else:
                folder_name_to_id = {}
                folder_id_set = set()
                for f in folders:
                    folder_name_to_id[f["FolderName"]] = f["FolderId"]
                    folder_id_set.add(str(f["FolderId"]))

                # フォルダID解決 & ファイル存在チェック
                preview_data = []
                for idx, row in df.iterrows():
                    cat_path = ""
                    if has_categories:
                        cat_path = "/".join([
                            row.get("カテゴリ1", ""),
                            row.get("カテゴリ2", ""),
                            row.get("カテゴリ3", ""),
                        ]).strip("/")

                    folder_id = ""
                    target_folder_name = ""
                    if has_folder_id and row.get("フォルダID", "").strip():
                        fid = row["フォルダID"].strip()
                        if fid in folder_id_set:
                            folder_id = fid
                            status = "✅ OK"
                            for f in folders:
                                if str(f["FolderId"]) == fid:
                                    target_folder_name = f["FolderName"]
                                    break
                        else:
                            status = "❌ フォルダID不正"
                    elif has_categories:
                        target_folder_name = row.get("カテゴリ3", "").strip()
                        if not target_folder_name:
                            target_folder_name = row.get("カテゴリ2", "").strip()
                        if not target_folder_name:
                            target_folder_name = row.get("カテゴリ1", "").strip()
                        folder_id = folder_name_to_id.get(target_folder_name, "")
                        status = "✅ OK" if folder_id else "❌ フォルダ未検出"
                    else:
                        status = "❌ フォルダ指定なし"

                    # ファイルパス検証
                    file_path_local = row["ファイルパス"].strip()
                    if not file_path_local:
                        status = "❌ ファイルパス未入力"
                    elif not os.path.isfile(file_path_local):
                        status = "❌ ファイルが見つかりません"

                    file_name = row["ファイル名"].strip()
                    if not file_name:
                        status = "❌ ファイル名未入力"

                    preview_data.append({
                        "No": idx + 1,
                        "コピー先": f"{target_folder_name}（ID: {folder_id}）" if folder_id else target_folder_name,
                        "カテゴリパス": cat_path,
                        "ファイル名": file_name,
                        "ファイルパス": file_path_local,
                        "チェック": status,
                        "_folder_id": folder_id,
                        "_file_path": file_path_local,
                        "_file_name": file_name,
                    })

                # プレビュー表示
                preview_df = pd.DataFrame(preview_data)
                display_df = preview_df[["No", "コピー先", "カテゴリパス", "ファイル名", "ファイルパス", "チェック"]]

                ok_count = sum(1 for d in preview_data if d["チェック"] == "✅ OK")
                ng_count = len(preview_data) - ok_count

                if ng_count > 0:
                    st.warning(f"実行可能: {ok_count} 件 / エラー: {ng_count} 件（エラー行はスキップされます）")
                else:
                    st.success(f"全 {ok_count} 件 実行可能です")

                st.dataframe(display_df, use_container_width=True, hide_index=True)

                # 実行範囲（バッチ処理）
                st.divider()
                st.markdown("### 実行設定")
                if ok_count > 500:
                    st.info("⏱ Streamlit Cloudのタイムアウト対策のため、500件ずつの実行を推奨します")
                col_start, col_end = st.columns(2)
                with col_start:
                    batch_start = st.number_input("開始（No）", min_value=1, max_value=max(ok_count, 1), value=1, step=1, key="local_batch_start")
                with col_end:
                    batch_end_default = min(500, ok_count)
                    batch_end = st.number_input("終了（No）", min_value=1, max_value=max(ok_count, 1), value=max(batch_end_default, 1), step=1, key="local_batch_end")

                batch_size = batch_end - batch_start + 1
                st.markdown(f"**実行対象: {batch_start}〜{batch_end} 件目（{batch_size} 件）**")

                # 上書きオプション
                overwrite = st.checkbox("同名ファイルが存在する場合は上書きする", value=False, key="local_overwrite")

                # 実行ボタン
                if ok_count > 0 and st.button(f"📤 アップロード実行（{batch_start}〜{batch_end}件目）", type="primary", key="local_exec"):
                    progress = st.progress(0, text="アップロード中...")
                    stop_placeholder = st.empty()
                    results = []
                    stopped = False
                    all_ok_rows = [d for d in preview_data if d["チェック"] == "✅ OK"]
                    target_rows = all_ok_rows[batch_start - 1:batch_end]
                    total = len(target_rows)

                    for i, row_data in enumerate(target_rows):
                        if stop_placeholder.button("⏹ 停止", key=f"local_stop_{i}"):
                            stopped = True
                            break
                        progress.progress((i + 1) / total, text=f"アップロード中... ({i + 1}/{total}) {row_data['_file_name']}")

                        # ローカルファイルを読み込み
                        try:
                            with open(row_data["_file_path"], "rb") as local_f:
                                file_data = local_f.read()
                        except Exception as e:
                            results.append({
                                "No": row_data["No"],
                                "ファイル名": row_data["_file_name"],
                                "コピー先": row_data["コピー先"],
                                "結果": "❌ 失敗",
                                "エラー": f"ファイル読み込み失敗: {str(e)}",
                            })
                            continue

                        # 2MB制限チェック
                        if len(file_data) > 2 * 1024 * 1024:
                            results.append({
                                "No": row_data["No"],
                                "ファイル名": row_data["_file_name"],
                                "コピー先": row_data["コピー先"],
                                "結果": "❌ 失敗",
                                "エラー": f"ファイルサイズ超過: {len(file_data) / 1024 / 1024:.1f}MB（上限2MB）",
                            })
                            continue

                        # 拡張子を取得
                        ext = os.path.splitext(row_data["_file_path"])[1].lstrip(".").lower()
                        if not ext:
                            ext = "jpg"

                        # fileName（画像名）: 拡張子なし
                        api_file_name = row_data["_file_name"]

                        # filePath（URLのファイル名）: 拡張子付き、20バイト制限
                        file_path_name = f"{row_data['_file_name']}.{ext}"
                        if len(file_path_name.encode("utf-8")) > 20:
                            stem = row_data["_file_name"]
                            while len(f"{stem}.{ext}".encode("utf-8")) > 20 and stem:
                                stem = stem[:-1]
                            file_path_name = f"{stem}.{ext}"

                        # アップロード
                        result = upload_image(
                            file_data=file_data,
                            file_name=api_file_name,
                            folder_id=row_data["_folder_id"],
                            file_path_name=file_path_name,
                            overwrite=overwrite,
                        )

                        results.append({
                            "No": row_data["No"],
                            "ファイル名": row_data["_file_name"],
                            "コピー先": row_data["コピー先"],
                            "結果": "✅ 成功" if result["success"] else "❌ 失敗",
                            "エラー": result.get("error", ""),
                        })

                        # API負荷軽減（3 req/sec制限）
                        if i < total - 1:
                            time.sleep(0.35)

                    progress.empty()
                    stop_placeholder.empty()

                    # 結果表示
                    success_count = sum(1 for r in results if r["結果"] == "✅ 成功")
                    fail_count = sum(1 for r in results if r["結果"] == "❌ 失敗")
                    processed = len(results)

                    if stopped:
                        st.warning(f"⏹ {processed} 件目で停止しました（成功: {success_count} / 失敗: {fail_count}）")
                    elif fail_count == 0:
                        st.success(f"全 {total} 件のアップロードが完了しました！")
                    elif success_count == 0:
                        st.error(f"全 {total} 件のアップロードに失敗しました")
                    else:
                        st.warning(f"成功: {success_count} / 失敗: {fail_count}")

                    st.dataframe(
                        pd.DataFrame(results),
                        use_container_width=True,
                        hide_index=True,
                    )

# ============================
# 💾 コピー：R-Cabi⇒Local
# ============================
elif mode == "💾 コピー：R-Cabi⇒Local":
    st.markdown("## 💾 コピー：R-Cabi⇒Local")
    st.markdown("指定フォルダの画像をフォルダ構成ごとローカルにダウンロードします。")

    st.warning("⚠ この機能はローカル実行時のみ使用できます（Streamlit Cloudではローカルにファイル保存できません）")

    # Excelテンプレートダウンロード
    if st.button("📥 テンプレートをダウンロード（フォルダ一覧付き）", key="dl_folder_tmpl_btn"):
        with st.spinner("フォルダ一覧を取得中..."):
            tmpl_folders, tmpl_error = get_all_folders()

        if tmpl_error:
            st.error(f"フォルダ一覧の取得に失敗しました: {tmpl_error}")
        else:
            styles, _ = get_openpyxl_styles()
            Font = styles['Font']
            PatternFill = styles['PatternFill']
            Alignment = styles['Alignment']
            from openpyxl import Workbook

            wb = Workbook()

            # シート1: ダウンロード指示
            ws1 = wb.active
            ws1.title = "ダウンロード指示"
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
            cell_font = Font(name="Meiryo UI", size=10)

            cell = ws1.cell(row=1, column=1, value="フォルダID")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

            ws1.cell(row=2, column=1, value="13054250").font = cell_font
            ws1.column_dimensions['A'].width = 16

            # シート2: フォルダ一覧
            ws2 = wb.create_sheet(title="フォルダ一覧")
            headers2 = ["フォルダID", "フォルダ名", "ディレクトリパス", "ファイル数"]
            for col_idx, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, f in enumerate(tmpl_folders, 2):
                ws2.cell(row=row_idx, column=1, value=f["FolderId"]).font = cell_font
                ws2.cell(row=row_idx, column=2, value=f["FolderName"]).font = cell_font
                ws2.cell(row=row_idx, column=3, value=f["FolderPath"]).font = cell_font
                ws2.cell(row=row_idx, column=4, value=f["FileCount"]).font = cell_font

            ws2.column_dimensions['A'].width = 14
            ws2.column_dimensions['B'].width = 30
            ws2.column_dimensions['C'].width = 40
            ws2.column_dimensions['D'].width = 12

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.download_button(
                label="📥 ダウンロード",
                data=excel_buffer.getvalue(),
                file_name="folder_download_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_folder_dl_template",
            )

    st.divider()

    # 保存先フォルダ
    save_root = st.text_input("保存先フォルダ", value=r"C:\Users\ssasa\Downloads\rcabinet_backup", key="dl_save_root")

    # CSV/Excelアップロード
    upload_file = st.file_uploader("CSV / Excelファイルを選択（フォルダID列が必要）", type=["csv", "xlsx"], key="folder_dl_file")

    if upload_file:
        # ファイル読み込み
        if upload_file.name.endswith(".xlsx"):
            upload_file.seek(0)
            df = pd.read_excel(upload_file, sheet_name=0, dtype=str).fillna("")
        else:
            try:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="utf-8-sig", dtype=str).fillna("")
            except Exception:
                upload_file.seek(0)
                df = pd.read_csv(upload_file, encoding="cp932", dtype=str).fillna("")

        if "フォルダID" not in df.columns:
            st.error("「フォルダID」列が必要です")
        elif len(df) == 0:
            st.warning("データがありません。")
        else:
            folder_ids_input = [r.strip() for r in df["フォルダID"] if r.strip()]
            st.info(f"📎 {len(folder_ids_input)} 件のフォルダIDを読み込みました")

            # フォルダ一覧を取得して検証
            with st.spinner("フォルダ一覧を取得中..."):
                folders, folder_error = get_all_folders()

            if folder_error:
                st.error(f"フォルダ一覧の取得に失敗しました: {folder_error}")
            else:
                # フォルダID → 情報のマップ
                folder_map = {}
                for f in folders:
                    folder_map[str(f["FolderId"])] = f

                # 指定フォルダIDの検証 & サブフォルダ収集
                target_folders = []
                preview_rows = []

                for fid in folder_ids_input:
                    if fid not in folder_map:
                        preview_rows.append({
                            "フォルダID": fid,
                            "フォルダ名": "",
                            "パス": "",
                            "ファイル数": 0,
                            "チェック": "❌ フォルダID不正",
                        })
                        continue

                    f_info = folder_map[fid]
                    f_path = f_info["FolderPath"]

                    # このフォルダ自体を追加
                    if f_info not in target_folders:
                        target_folders.append(f_info)
                        preview_rows.append({
                            "フォルダID": fid,
                            "フォルダ名": f_info["FolderName"],
                            "パス": f_path,
                            "ファイル数": f_info["FileCount"],
                            "チェック": "✅ OK",
                        })

                    # サブフォルダを再帰的に収集（パスの前方一致）
                    for sf in folders:
                        sf_path = sf["FolderPath"]
                        sf_id = str(sf["FolderId"])
                        if sf_id != fid and sf_path.startswith(f_path + "/"):
                            if sf not in target_folders:
                                target_folders.append(sf)
                                preview_rows.append({
                                    "フォルダID": sf_id,
                                    "フォルダ名": sf["FolderName"],
                                    "パス": sf_path,
                                    "ファイル数": sf["FileCount"],
                                    "チェック": "✅ OK（サブフォルダ）",
                                })

                ok_count = sum(1 for r in preview_rows if "OK" in r["チェック"])
                ng_count = len(preview_rows) - ok_count
                total_files = sum(int(r.get("ファイル数", 0)) for r in preview_rows if "OK" in r["チェック"])

                if ng_count > 0:
                    st.warning(f"対象フォルダ: {ok_count} 件 / エラー: {ng_count} 件")
                else:
                    st.success(f"対象フォルダ: {ok_count} 件（推定ファイル数: {total_files} 件）")

                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

                if ok_count > 0:
                    st.divider()

                    # 実行ボタン
                    if st.button(f"📥 ダウンロード実行（{ok_count} フォルダ / 推定 {total_files} ファイル）", type="primary", key="folder_dl_exec"):
                        os.makedirs(save_root, exist_ok=True)

                        progress = st.progress(0, text="ダウンロード準備中...")
                        stop_placeholder = st.empty()
                        results = []
                        stopped = False
                        total_folder_count = len(target_folders)

                        # FolderPath → FolderName マップを構築（パス→表示名変換用）
                        path_to_name = {}
                        for f in folders:
                            path_to_name[f["FolderPath"]] = f["FolderName"]

                        def build_display_path(folder_path):
                            """FolderPathの各階層をFolderNameに変換"""
                            parts = folder_path.strip("/").split("/")
                            display_parts = []
                            for i in range(len(parts)):
                                partial = "/" + "/".join(parts[:i + 1])
                                display_parts.append(path_to_name.get(partial, parts[i]))
                            return os.sep.join(display_parts)

                        for fi, folder_info in enumerate(target_folders):
                            if stop_placeholder.button("⏹ 停止", key=f"dl_stop_{fi}"):
                                stopped = True
                                break

                            fid = str(folder_info["FolderId"])
                            f_path = folder_info["FolderPath"]
                            f_name = folder_info["FolderName"]

                            progress.progress(
                                (fi + 1) / total_folder_count,
                                text=f"フォルダ取得中... ({fi + 1}/{total_folder_count}) {f_name}"
                            )

                            # ローカルのフォルダパスを構築（フォルダ名で作成）
                            display_path = build_display_path(f_path)
                            local_folder = os.path.join(save_root, display_path)
                            os.makedirs(local_folder, exist_ok=True)

                            # フォルダ内ファイル一覧を取得
                            files, files_error = get_folder_files(int(fid))

                            # パス分解
                            path_parts = f_path.split("/")
                            cat1 = path_parts[0] if len(path_parts) > 0 else ""
                            cat2 = path_parts[1] if len(path_parts) > 1 else ""
                            cat3 = path_parts[2] if len(path_parts) > 2 else ""

                            if files_error:
                                results.append({
                                    "カテゴリ1": cat1,
                                    "カテゴリ2": cat2,
                                    "カテゴリ3": cat3,
                                    "ファイル名": "",
                                    "結果": "❌ 失敗",
                                    "エラー": f"ファイル一覧取得失敗: {files_error}",
                                })
                                continue

                            if not files:
                                continue

                            for file_info in files:
                                if stopped:
                                    break

                                file_url = file_info["FileUrl"]
                                # FilePath（拡張子付き）を優先、なければURLから拡張子を補完
                                file_name_raw = file_info["FileName"]
                                file_path_name = file_info.get("FilePath", "")
                                if file_path_name:
                                    file_name = file_path_name
                                elif file_url and "." in file_url.rsplit("/", 1)[-1]:
                                    ext = file_url.rsplit(".", 1)[-1]
                                    file_name = f"{file_name_raw}.{ext}"
                                else:
                                    file_name = file_name_raw

                                progress.progress(
                                    (fi + 1) / total_folder_count,
                                    text=f"ダウンロード中... フォルダ({fi + 1}/{total_folder_count}) {file_name}"
                                )

                                try:
                                    img_response = requests.get(file_url, timeout=30)
                                    if img_response.status_code != 200:
                                        results.append({
                                            "カテゴリ1": cat1,
                                            "カテゴリ2": cat2,
                                            "カテゴリ3": cat3,
                                            "ファイル名": file_name,
                                            "結果": "❌ 失敗",
                                            "エラー": f"HTTP {img_response.status_code}",
                                        })
                                        continue

                                    local_file_path = os.path.join(local_folder, file_name)
                                    with open(local_file_path, "wb") as lf:
                                        lf.write(img_response.content)

                                    results.append({
                                        "カテゴリ1": cat1,
                                        "カテゴリ2": cat2,
                                        "カテゴリ3": cat3,
                                        "ファイル名": file_name,
                                        "結果": "✅ 成功",
                                        "エラー": "",
                                    })

                                except requests.exceptions.RequestException as e:
                                    results.append({
                                        "カテゴリ1": cat1,
                                        "カテゴリ2": cat2,
                                        "カテゴリ3": cat3,
                                        "ファイル名": file_name,
                                        "結果": "❌ 失敗",
                                        "エラー": str(e),
                                    })
                                except OSError as e:
                                    results.append({
                                        "カテゴリ1": cat1,
                                        "カテゴリ2": cat2,
                                        "カテゴリ3": cat3,
                                        "ファイル名": file_name,
                                        "結果": "❌ 失敗",
                                        "エラー": f"保存失敗: {str(e)}",
                                    })

                                time.sleep(0.2)

                            if stopped:
                                break
                            time.sleep(0.3)

                        progress.empty()
                        stop_placeholder.empty()

                        # 結果表示
                        success_count = sum(1 for r in results if r["結果"] == "✅ 成功")
                        fail_count = sum(1 for r in results if r["結果"] == "❌ 失敗")

                        if stopped:
                            st.warning(f"⏹ 停止しました（成功: {success_count} / 失敗: {fail_count}）")
                        elif fail_count == 0:
                            st.success(f"全 {success_count} 件のダウンロードが完了しました！")
                        elif success_count == 0:
                            st.error(f"全 {fail_count} 件のダウンロードに失敗しました")
                        else:
                            st.warning(f"成功: {success_count} / 失敗: {fail_count}")

                        st.markdown(f"**保存先:** `{save_root}`")

                        result_df = pd.DataFrame(results)
                        st.dataframe(result_df, use_container_width=True, hide_index=True)

                        # 結果Excelダウンロード
                        if results:
                            styles, _ = get_openpyxl_styles()
                            Font = styles['Font']
                            PatternFill = styles['PatternFill']
                            Alignment = styles['Alignment']
                            from openpyxl import Workbook

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "ダウンロード結果"

                            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
                            cell_font = Font(name="Meiryo UI", size=10)

                            headers = ["No", "カテゴリ1", "カテゴリ2", "カテゴリ3", "ファイル名", "結果", "エラー"]
                            for col_idx, header in enumerate(headers, 1):
                                cell = ws.cell(row=1, column=col_idx, value=header)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center")

                            for row_idx, r in enumerate(results, 2):
                                ws.cell(row=row_idx, column=1, value=row_idx - 1).font = cell_font
                                ws.cell(row=row_idx, column=2, value=r["カテゴリ1"]).font = cell_font
                                ws.cell(row=row_idx, column=3, value=r["カテゴリ2"]).font = cell_font
                                ws.cell(row=row_idx, column=4, value=r["カテゴリ3"]).font = cell_font
                                ws.cell(row=row_idx, column=5, value=r["ファイル名"]).font = cell_font
                                ws.cell(row=row_idx, column=6, value=r["結果"]).font = cell_font
                                ws.cell(row=row_idx, column=7, value=r.get("エラー", "")).font = cell_font

                            ws.column_dimensions['A'].width = 8
                            ws.column_dimensions['B'].width = 18
                            ws.column_dimensions['C'].width = 18
                            ws.column_dimensions['D'].width = 18
                            ws.column_dimensions['E'].width = 24
                            ws.column_dimensions['F'].width = 10
                            ws.column_dimensions['G'].width = 40

                            result_buffer = BytesIO()
                            wb.save(result_buffer)
                            result_buffer.seek(0)

                            st.download_button(
                                label="📥 結果をExcelでダウンロード",
                                data=result_buffer.getvalue(),
                                file_name=f"download_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_result_xlsx",
                            )
