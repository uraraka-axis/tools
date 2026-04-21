import streamlit as st
import pandas as pd
import requests
import time
import io
import re
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from supabase import create_client

st.set_page_config(page_title="楽天 競合価格チェッカー", layout="wide")


def check_password():
    """パスワード認証"""
    if "password" not in st.secrets:
        return True

    correct_password = st.secrets["password"]

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.title("ログイン")
    st.write("このツールを使用するにはパスワードが必要です。")
    password = st.text_input("パスワード", type="password", key="password_input")

    if st.button("ログイン", type="primary"):
        if password == correct_password:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが違います")

    return False


if not check_password():
    st.stop()

st.title("楽天 競合価格チェッカー")
st.caption("商品リスト(xlsx)をアップロードし、楽天市場での競合出品状況を調査します")


def style_excel_sheet(ws, col_widths=None):
    """Excelシートに共通書式を適用する（Meiryo UI、ヘッダー色、列幅）"""
    header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    body_font = Font(name="Meiryo UI", size=10)

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

# --- API設定（secrets.tomlから取得） ---
app_id = st.secrets.get("rakuten_app_id", "")
access_key = st.secrets.get("rakuten_access_key", "")

# --- Supabase接続 ---
supabase_url = st.secrets.get("supabase_url", "")
supabase_key = st.secrets.get("supabase_key", "")
supabase = None
if supabase_url and supabase_key:
    supabase = create_client(supabase_url, supabase_key)


def db_load_shipping_params():
    """DBから送料パラメータキャッシュを一括読み込み → session_stateに格納"""
    if supabase is None:
        return
    if "shipping_params_loaded" in st.session_state:
        return  # 既に読み込み済み
    try:
        resp = supabase.table("shipping_params").select("*").execute()
        cache = {}
        for row in resp.data:
            if row["postage_included"]:
                cache[row["item_url"]] = "送料込"
            else:
                cache[row["item_url"]] = {
                    "shopId": row["shop_id"],
                    "itemId": row["item_id"],
                    "deliverySetId": row.get("delivery_set_id"),
                    "customTariffId": row.get("custom_tariff_id"),
                    "itemCode": row.get("item_code"),
                    "price": row.get("price", 0),
                    "postageIncluded": False,
                }
        if "shipping_params_cache" not in st.session_state:
            st.session_state["shipping_params_cache"] = {}
        st.session_state["shipping_params_cache"].update(cache)
        st.session_state["shipping_params_loaded"] = True
    except Exception:
        pass  # DB接続失敗時はキャッシュなしで続行


def db_save_shipping_params(item_url, params):
    """送料パラメータを1件DBに保存（upsert）"""
    if supabase is None:
        return
    try:
        if params == "送料込":
            row = {"item_url": item_url, "postage_included": True}
        else:
            row = {
                "item_url": item_url,
                "shop_id": params.get("shopId"),
                "item_id": params.get("itemId"),
                "delivery_set_id": params.get("deliverySetId"),
                "custom_tariff_id": params.get("customTariffId"),
                "item_code": params.get("itemCode"),
                "price": params.get("price"),
                "postage_included": params.get("postageIncluded", False),
            }
        supabase.table("shipping_params").upsert(row).execute()
    except Exception:
        pass  # DB書き込み失敗は無視（メモリキャッシュは残る）


def db_save_price_summary(summary_df):
    """品番別サマリーをDBに保存（同日同JANは上書き）"""
    if supabase is None or summary_df.empty:
        return
    try:
        today = time.strftime("%Y-%m-%d")
        rows = []
        for _, r in summary_df.iterrows():
            rows.append({
                "brand": r.get("ブランド"),
                "jan_code": r.get("JANコード"),
                "product_code": r.get("品番"),
                "list_price": int(r["定価"]) if pd.notna(r.get("定価")) else None,
                "listing_count": int(r["出品数"]) if pd.notna(r.get("出品数")) else None,
                "min_total": int(r["最安値_合計"]) if pd.notna(r.get("最安値_合計")) else None,
                "min_body": int(r["最安値_本体"]) if pd.notna(r.get("最安値_本体")) else None,
                "min_shipping": int(r["最安値_送料"]) if pd.notna(r.get("最安値_送料")) else None,
                "min_ratio": float(r["最安値_定価比率"]) if pd.notna(r.get("最安値_定価比率")) else None,
                "max_total": int(r["最高値_合計"]) if pd.notna(r.get("最高値_合計")) else None,
                "max_body": int(r["最高値_本体"]) if pd.notna(r.get("最高値_本体")) else None,
                "max_shipping": int(r["最高値_送料"]) if pd.notna(r.get("最高値_送料")) else None,
                "avg_total": int(r["平均合計"]) if pd.notna(r.get("平均合計")) else None,
                "avg_ratio": float(r["平均_定価比率"]) if pd.notna(r.get("平均_定価比率")) else None,
                "min_url_link": r.get("最安値_URL"),
                "max_url_link": r.get("最高値_URL"),
                "checked_date": today,
            })
        supabase.table("price_summary").upsert(
            rows, on_conflict="jan_code,checked_date"
        ).execute()
    except Exception:
        pass


def db_load_price_history(jan_codes=None):
    """DBから価格推移データを取得する"""
    if supabase is None:
        return pd.DataFrame()
    try:
        query = supabase.table("price_summary").select("*").order("checked_date")
        if jan_codes:
            query = query.in_("jan_code", jan_codes)
        resp = query.execute()
        if resp.data:
            return pd.DataFrame(resp.data)
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


# 起動時にDBからキャッシュ読み込み
db_load_shipping_params()

# --- サイドバー: 検索設定 ---
with st.sidebar:
    st.header("検索設定")
    max_items = st.number_input("1商品あたりの最大取得件数", min_value=1, max_value=30, value=5)
    delay = st.number_input("リクエスト間隔（秒）", min_value=0.5, max_value=5.0, value=1.0, step=0.5)

    # 送料パラメータキャッシュ状況
    cache_count = len(st.session_state.get("shipping_params_cache", {}))
    if cache_count > 0:
        st.caption(f"送料キャッシュ: {cache_count}件")
        if st.button("キャッシュクリア", key="clear_cache"):
            st.session_state.pop("shipping_params_cache", None)
            st.rerun()

# --- テンプレートダウンロード ---
st.subheader("1. テンプレート")


@st.cache_data
def create_template():
    """入力テンプレートExcelを生成する"""
    df = pd.DataFrame({
        "No.": [1, 2, 3],
        "ブランド": ["東谷", "東谷", "東谷"],
        "JANコード": ["4985155188260", "4985155220151", "4985155194988"],
        "品番": ["WE-332LBR", "HS-68BK", "SS-117"],
        "定価": [510000, 158000, 110000],
    })
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="商品リスト", index=False)
        ws = writer.sheets["商品リスト"]
        style_excel_sheet(ws, col_widths=[6, 14, 18, 16, 12])
    output.seek(0)
    return output.getvalue()


st.download_button(
    label="入力テンプレートをダウンロード",
    data=create_template(),
    file_name="rakuten_price_checker_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# --- メイン: ファイルアップロード ---
st.subheader("2. 商品リストをアップロード")
uploaded_file = st.file_uploader(
    "商品リスト(.xlsx)をアップロード",
    type=["xls", "xlsx"],
)

if not uploaded_file:
    st.info("テンプレートに沿ったExcelファイルをアップロードしてください")
    st.stop()


# --- Excel読み込み ---
@st.cache_data
def load_excel(file_bytes, file_name):
    """商品リストを読み込む（テンプレート形式: No., ブランド, JANコード, 品番, 定価）"""
    engine = "xlrd" if file_name.endswith(".xls") else "openpyxl"
    df = pd.read_excel(io.BytesIO(file_bytes), header=0, engine=engine)

    # カラム名を正規化
    df.columns = [str(c).strip() for c in df.columns]

    products = []
    for _, row in df.iterrows():
        jan = str(row.get("JANコード", "")).strip()
        code = str(row.get("品番", "")).strip()
        if not jan and not code:
            continue
        brand = str(row.get("ブランド", "")).strip() if pd.notna(row.get("ブランド")) else ""
        price_raw = row.get("定価")
        price = int(float(price_raw)) if pd.notna(price_raw) and str(price_raw).strip() else 0

        products.append({
            "ブランド": brand,
            "JANコード": jan,
            "品番": code,
            "定価": price,
        })
    return pd.DataFrame(products)


file_bytes = uploaded_file.read()
products_df = load_excel(file_bytes, uploaded_file.name)

st.write(f"読み込み結果: **{len(products_df)}商品**")
st.dataframe(products_df, use_container_width=True)

# --- フィルタリング ---
st.markdown("---")
st.subheader("3. 検索対象の絞り込み（任意）")
col1, col2 = st.columns(2)
with col1:
    keyword_filter = st.text_input("品番・ブランドで絞り込み", "")
with col2:
    price_min = st.number_input("定価（最小）", value=0, step=10000)
    price_max = st.number_input("定価（最大）", value=9999999, step=10000)

filtered_df = products_df.copy()
if keyword_filter:
    mask = (
        filtered_df["品番"].str.contains(keyword_filter, case=False, na=False)
        | filtered_df["ブランド"].str.contains(keyword_filter, case=False, na=False)
    )
    filtered_df = filtered_df[mask]
filtered_df = filtered_df[
    (filtered_df["定価"] >= price_min) & (filtered_df["定価"] <= price_max)
]
# JAN未設定の行を除外
filtered_df = filtered_df[filtered_df["JANコード"].str.len() >= 8]

st.info(f"検索対象: {len(filtered_df)}商品")


# --- API検索 ---
def search_rakuten(app_id, access_key, keyword, hits=30):
    """楽天市場商品検索APIで検索し、商品リストを返す"""
    url = "https://openapi.rakuten.co.jp/ichibams/api/IchibaItem/Search/20260401"
    params = {
        "applicationId": app_id,
        "accessKey": access_key,
        "format": "json",
        "formatVersion": 2,
        "keyword": keyword,
        "hits": hits,
        "sort": "+itemPrice",
    }
    headers = {
        "Origin": "https://rakuten.co.jp",
        "Referer": "https://rakuten.co.jp/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
    }
    resp = requests.get(url, params=params, headers=headers, timeout=15)
    if resp.status_code == 200:
        return resp.json()
    elif resp.status_code == 404:
        return None
    elif resp.status_code == 429:
        raise Exception("API制限に達しました。しばらく待ってから再実行してください。")
    else:
        raise Exception(f"APIエラー: {resp.status_code} - {resp.text[:200]}")


def extract_price_info(api_result, brand, jan, code, list_price):
    """APIレスポンスから価格情報を抽出する"""
    rows = []
    if not api_result or "Items" not in api_result:
        rows.append({
            "ブランド": brand,
            "JANコード": jan,
            "品番": code,
            "定価": list_price,
            "楽天ショップ名": "（出品なし）",
            "楽天商品名": None,
            "キャッチコピー": None,
            "楽天商品コード": None,
            "商品説明": None,
            "販売価格": None,
            "送料区分": None,
            "送料金額": None,
            "合計金額": None,
            "定価比率": None,
            "レビュー数": None,
            "レビュー平均": None,
            "商品URL": None,
        })
        return rows

    for item in api_result["Items"]:
        price = item.get("itemPrice", 0)
        postage_flag = "送料込" if item.get("postageFlag") == 0 else "送料別"
        ratio = round(price / list_price * 100, 1) if list_price > 0 else None

        # 送料込の場合、送料金額=0、合計=販売価格
        shipping_fee = 0 if postage_flag == "送料込" else None
        total = price if postage_flag == "送料込" else None

        rows.append({
            "ブランド": brand,
            "JANコード": jan,
            "品番": code,
            "定価": list_price,
            "楽天ショップ名": item.get("shopName", ""),
            "楽天商品名": item.get("itemName", ""),
            "キャッチコピー": item.get("catchcopy", ""),
            "楽天商品コード": item.get("itemCode", ""),
            "商品説明": item.get("itemCaption", ""),
            "販売価格": price,
            "送料区分": postage_flag,
            "送料金額": shipping_fee,
            "合計金額": total,
            "定価比率": ratio,
            "レビュー数": item.get("reviewCount", 0),
            "レビュー平均": item.get("reviewAverage", 0),
            "商品URL": item.get("itemUrl", ""),
        })
    return rows


_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
]


def _create_scrape_session():
    """ブラウザに近いセッションを作成する"""
    session = requests.Session()
    session.headers.update({
        "User-Agent": random.choice(_USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
    })
    return session


def _extract_shipping_params(html_text):
    """静的HTMLから送料API呼び出しに必要なパラメータを抽出する"""
    params = {}

    # shopId
    m = re.search(r'data-shop-id="(\d+)"', html_text)
    if m:
        params["shopId"] = m.group(1)

    # itemId
    m = re.search(r'data-item-id="(\d+)"', html_text)
    if m:
        params["itemId"] = m.group(1)

    # customTariffId（shopAreaSoryoPatternId）
    m = re.search(r'"shopAreaSoryoPatternId"\s*:\s*(\d+)', html_text)
    if m:
        params["customTariffId"] = m.group(1)

    # deliverySetId（shippingMethodGroup）
    m = re.search(r'"shippingMethodGroup"\s*:\s*"(\d+)"', html_text)
    if m:
        params["deliverySetId"] = m.group(1)

    # price
    m = re.search(r'data-price="(\d+)"', html_text)
    if m:
        params["price"] = int(m.group(1))

    # postageIncluded
    m = re.search(r'"postageIncluded"\s*:\s*(true|false)', html_text)
    if m:
        params["postageIncluded"] = m.group(1) == "true"

    return params


def _call_shipping_api(params):
    """楽天送料APIを呼び出して送料を取得する"""
    api_url = "https://gateway-api.global.rakuten.com/shippingx/v2/shippingCalculation"
    api_key = "s7589xj4EGA5gOsLmncgJDT6yCATYFwp"

    item_code = params.get("itemCode", "")
    item_key = f"{params['itemId']}:{item_code}"

    # customTariffIdはnull許容
    custom_tariff = params.get("customTariffId")
    if custom_tariff is not None:
        custom_tariff = str(custom_tariff)

    body = {
        "marketplaceId": "JP",
        "resultType": "DATE_FEE",
        "calculationMode": "CHEAPEST",
        "shippingUnits": {
            "unit01": {
                "shipTo": {"level1": "JP", "level2": "13"},  # 東京
                "shopShippingUnits": {
                    "shopUnit1": {
                        "shopId": params["shopId"],
                        "items": {
                            item_key: {
                                "quantity": 1,
                                "data": {
                                    "price": params.get("price", 0),
                                    "individualShipping": False,
                                    "customShipping": {
                                        "postageSegment1": 0,
                                        "postageSegment2": 0,
                                        "customTariffId": custom_tariff,
                                    },
                                    "deliverySetId": int(params["deliverySetId"]) if params.get("deliverySetId") else None,
                                    "includedPostage": params.get("postageIncluded", False),
                                    "inventory": 1,
                                    "handlingTime": {"inStockDelvTemplateId": 2},
                                },
                            }
                        },
                    }
                },
            }
        },
        "calculationSettings": {
            "showAvailableThresholdDiscounts": True,
            "showCalculationGroups": True,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": random.choice(_USER_AGENTS),
    }

    resp = requests.post(
        f"{api_url}?apikey={api_key}",
        json=body,
        headers=headers,
        timeout=15,
    )

    if resp.status_code != 200:
        return None

    data = resp.json()
    try:
        results = data["shippingUnits"]["unit01"]["shopShippingUnits"]["shopUnit1"]["results"]
        final_fee = results[0]["fees"]["finalFee"]
        return int(final_fee)
    except (KeyError, IndexError, TypeError):
        return None


def _get_params_cache():
    """送料パラメータキャッシュを取得（なければ初期化）"""
    if "shipping_params_cache" not in st.session_state:
        st.session_state["shipping_params_cache"] = {}
    return st.session_state["shipping_params_cache"]


def scrape_shipping_cost(item_url, session=None):
    """楽天商品ページから送料を取得する（Shipping API経由）

    Returns:
        int: 送料金額（円）。0の場合は送料無料。
        str: "送料込" / "送料無料" / "取得不可"
    """
    if session is None:
        session = _create_scrape_session()

    params_cache = _get_params_cache()

    # キャッシュにパラメータがあればHTML取得をスキップ
    if item_url in params_cache:
        cached = params_cache[item_url]
        if cached == "送料込":
            return "送料込"
        # キャッシュ済みパラメータで送料APIを直接呼び出し
        fee = _call_shipping_api(cached)
        if fee is not None:
            return fee
        return "取得不可"

    session.headers["Referer"] = "https://search.rakuten.co.jp/"

    # ランダムな遅延（0.5〜1.5秒）
    time.sleep(random.uniform(0.5, 1.5))

    # Step 1: 商品ページの静的HTMLを取得してパラメータ抽出
    resp = session.get(item_url, timeout=15)
    resp.encoding = resp.apparent_encoding
    html = resp.text

    # 送料込みチェック
    if re.search(r'"postageIncluded"\s*:\s*true', html):
        params_cache[item_url] = "送料込"
        db_save_shipping_params(item_url, "送料込")
        return "送料込"

    params = _extract_shipping_params(html)

    # itemCodeはURLパスから取得
    m = re.search(r'item\.rakuten\.co\.jp/[^/]+/([^/?]+)', item_url)
    if m:
        params["itemCode"] = m.group(1)

    # 必須パラメータが揃っているか確認（customTariffId, deliverySetIdはnull許容）
    required = ["shopId", "itemId"]
    if not all(k in params for k in required):
        return "取得不可"

    # パラメータをキャッシュに保存（メモリ＋DB）
    params_cache[item_url] = params
    db_save_shipping_params(item_url, params)

    # Step 2: 送料APIを呼び出し
    fee = _call_shipping_api(params)
    if fee is not None:
        return fee  # int型で返す（0=送料無料）

    return "取得不可"


def _run_search(filtered_df, app_id, access_key, max_items, delay, resume=False):
    """検索を実行する（途中再開対応）"""
    # 処理済み品番セットを取得（再開時）
    if resume and "search_progress" in st.session_state:
        all_results = list(st.session_state["search_progress"]["results"])
        processed_jans = set(st.session_state["search_progress"]["processed_jans"])
    else:
        all_results = []
        processed_jans = set()
        st.session_state.pop("search_progress", None)

    # 未処理の商品だけ対象にする
    pending_indices = [
        idx for idx in filtered_df.index
        if filtered_df.loc[idx, "JANコード"] not in processed_jans
    ]
    total = len(filtered_df)
    done_count = total - len(pending_indices)

    progress_bar = st.progress(done_count / total if total > 0 else 0)
    status_text = st.empty()

    if done_count > 0:
        status_text.text(f"再開: {done_count}件処理済み → 残り{len(pending_indices)}件")

    for idx in pending_indices:
        row = filtered_df.loc[idx]
        done_count += 1
        progress = done_count / total
        progress_bar.progress(progress)
        status_text.text(f"検索中... {done_count}/{total} - {row['品番']}")

        try:
            search_key = row["JANコード"] if row["JANコード"] else row["品番"]
            result = search_rakuten(app_id, access_key, search_key, hits=max_items)
            rows = extract_price_info(
                result, row["ブランド"], row["JANコード"], row["品番"], row["定価"]
            )

            # 送料別の商品について送料を並列取得（最大3並列）
            betsu_rows = [r for r in rows if r["送料区分"] == "送料別" and r["商品URL"]]
            if betsu_rows:
                status_text.text(
                    f"送料取得中... {done_count}/{total} - {row['品番']} ({len(betsu_rows)}件)"
                )

                def _fetch_shipping(r):
                    s = _create_scrape_session()
                    try:
                        return r, scrape_shipping_cost(r["商品URL"], session=s)
                    except Exception:
                        return r, "取得不可"

                with ThreadPoolExecutor(max_workers=3) as executor:
                    futures = [executor.submit(_fetch_shipping, r) for r in betsu_rows]
                    for future in as_completed(futures):
                        r, fee = future.result()
                        if isinstance(fee, int):
                            r["送料金額"] = fee
                            r["合計金額"] = r["販売価格"] + fee
                        elif fee == "送料込":
                            r["送料区分"] = "送料込"
                            r["送料金額"] = 0
                            r["合計金額"] = r["販売価格"]
                        elif fee == "送料無料":
                            r["送料区分"] = "送料無料"
                            r["送料金額"] = 0
                            r["合計金額"] = r["販売価格"]
                        else:
                            r["送料金額"] = None
                            r["合計金額"] = None

            all_results.extend(rows)
        except Exception as e:
            st.warning(f"{row['品番']}: {e}")
            all_results.append({
                "ブランド": row["ブランド"],
                "JANコード": row["JANコード"],
                "品番": row["品番"],
                "定価": row["定価"],
                "楽天ショップ名": f"エラー: {e}",
                "楽天商品名": None,
                "キャッチコピー": None,
                "楽天商品コード": None,
                "商品説明": None,
                "販売価格": None,
                "送料区分": None,
                "送料金額": None,
                "合計金額": None,
                "定価比率": None,
                "レビュー数": None,
                "レビュー平均": None,
                "商品URL": None,
            })

        # チェックポイント: 1商品ごとにsession_stateに保存
        processed_jans.add(row["JANコード"])
        st.session_state["search_progress"] = {
            "results": all_results,
            "processed_jans": processed_jans,
        }

        time.sleep(delay)

    progress_bar.progress(1.0)
    status_text.text("検索完了！")

    results_df = pd.DataFrame(all_results)
    st.session_state["results_df"] = results_df
    # 完了したら進捗データをクリア
    st.session_state.pop("search_progress", None)


# --- 実行ボタン ---
st.markdown("---")
st.subheader("4. 検索実行")

# 途中経過がある場合、再開ボタンを表示
has_progress = "search_progress" in st.session_state
if has_progress:
    progress_info = st.session_state["search_progress"]
    done = len(progress_info["processed_jans"])
    total = len(filtered_df)
    remaining = total - done
    st.warning(f"前回の検索が途中で中断されています（{done}/{total}件処理済み、残り{remaining}件）")

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("途中から再開", type="primary"):
            if not app_id or not access_key:
                st.error("secrets.tomlにrakuten_app_idとrakuten_access_keyを設定してください")
                st.stop()
            _run_search(filtered_df, app_id, access_key, max_items, delay, resume=True)
    with col_btn2:
        if st.button("最初からやり直す"):
            st.session_state.pop("search_progress", None)
            st.session_state.pop("results_df", None)
            st.rerun()
else:
    if st.button("競合価格を検索", type="primary"):
        if not app_id or not access_key:
            st.error("secrets.tomlにrakuten_app_idとrakuten_access_keyを設定してください")
            st.stop()
        _run_search(filtered_df, app_id, access_key, max_items, delay, resume=False)

# 途中経過があれば部分結果も表示
if has_progress and "results_df" not in st.session_state:
    partial = st.session_state["search_progress"]["results"]
    if partial:
        st.info(f"途中経過: {len(partial)}件の出品データを取得済み")
        partial_df = pd.DataFrame(partial)
        st.session_state["results_df"] = partial_df

# --- 結果表示 ---
if "results_df" in st.session_state:
    results_df = st.session_state["results_df"]

    st.markdown("---")
    st.subheader("検索結果")

    # サマリー
    col1, col2, col3, col4 = st.columns(4)
    total_products = results_df["品番"].nunique()
    found = results_df[results_df["楽天ショップ名"] != "（出品なし）"]["品番"].nunique()
    not_found = total_products - found
    total_listings = len(results_df[results_df["販売価格"].notna()])
    col1.metric("検索商品数", total_products)
    col2.metric("出品あり", found)
    col3.metric("出品なし", not_found)
    col4.metric("総出品件数", total_listings)

    # 結果テーブル（F〜I列は非表示）
    display_df = results_df.drop(columns=["楽天商品名", "キャッチコピー", "楽天商品コード", "商品説明"], errors="ignore")
    st.dataframe(
        display_df.style.format({
            "定価": "¥{:,.0f}",
            "販売価格": lambda x: f"¥{x:,.0f}" if pd.notna(x) else "",
            "送料金額": lambda x: f"¥{x:,.0f}" if pd.notna(x) else "不明",
            "合計金額": lambda x: f"¥{x:,.0f}" if pd.notna(x) else "",
            "定価比率": lambda x: f"{x}%" if pd.notna(x) else "",
        }),
        use_container_width=True,
        height=600,
    )

    # --- サマリーテーブル（品番ごとの最安・最高・平均 - 合計金額ベース） ---
    st.subheader("品番別サマリー（本体＋送料の合計金額ベース）")
    valid = results_df[results_df["合計金額"].notna()].copy()
    if not valid.empty:
        def summarize_group(g):
            min_idx = g["合計金額"].idxmin()
            max_idx = g["合計金額"].idxmax()
            min_total = int(g.loc[min_idx, "合計金額"])
            max_total = int(g.loc[max_idx, "合計金額"])
            avg_total = int(round(g["合計金額"].mean()))
            list_price = g["定価"].iloc[0]
            return pd.Series({
                "出品数": len(g),
                "最安値_合計": min_total,
                "最安値_本体": int(g.loc[min_idx, "販売価格"]),
                "最安値_送料": int(g.loc[min_idx, "送料金額"]) if pd.notna(g.loc[min_idx, "送料金額"]) else None,
                "最安値_定価比率": round(min_total / list_price * 100, 1) if list_price > 0 else None,
                "最高値_合計": max_total,
                "最高値_本体": int(g.loc[max_idx, "販売価格"]),
                "最高値_送料": int(g.loc[max_idx, "送料金額"]) if pd.notna(g.loc[max_idx, "送料金額"]) else None,
                "平均合計": avg_total,
                "平均_定価比率": round(avg_total / list_price * 100, 1) if list_price > 0 else None,
                "最安値_URL": g.loc[min_idx, "商品URL"],
                "最高値_URL": g.loc[max_idx, "商品URL"],
            })

        summary = valid.groupby(["ブランド", "JANコード", "品番", "定価"]).apply(summarize_group).reset_index()

        # サマリーをDBに保存
        db_save_price_summary(summary)

        # 送料取得不可の件数を注記
        unknown_count = len(results_df[
            (results_df["販売価格"].notna()) & (results_df["合計金額"].isna())
        ])
        if unknown_count > 0:
            st.warning(f"送料取得不可の商品が {unknown_count}件 あり、サマリーから除外されています。")

        st.dataframe(
            summary.style.format({
                "定価": "¥{:,.0f}",
                "最安値_合計": "¥{:,.0f}",
                "最安値_本体": "¥{:,.0f}",
                "最安値_送料": lambda x: f"¥{x:,.0f}" if pd.notna(x) else "不明",
                "最安値_定価比率": lambda x: f"{x}%" if pd.notna(x) else "",
                "最高値_合計": "¥{:,.0f}",
                "最高値_本体": "¥{:,.0f}",
                "最高値_送料": lambda x: f"¥{x:,.0f}" if pd.notna(x) else "不明",
                "平均合計": "¥{:,.0f}",
                "平均_定価比率": lambda x: f"{x}%" if pd.notna(x) else "",
            }),
            use_container_width=True,
        )

    # --- Excelダウンロード ---
    st.subheader("ダウンロード")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="全出品一覧", index=False)
        ws1 = writer.sheets["全出品一覧"]
        style_excel_sheet(ws1, col_widths=[14, 18, 16, 12, 24, 30, 30, 20, 40, 12, 10, 12, 12, 12, 12, 12, 40])
        if not valid.empty:
            summary.to_excel(writer, sheet_name="品番別サマリー", index=False)
            ws2 = writer.sheets["品番別サマリー"]
            style_excel_sheet(ws2)
    output.seek(0)
    st.download_button(
        label="Excelでダウンロード",
        data=output,
        file_name="rakuten_price_check.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# --- 価格推移 ---
if supabase is not None:
    st.markdown("---")
    st.subheader("価格推移")

    history_df = db_load_price_history()
    if history_df.empty:
        st.info("まだ履歴データがありません。検索を実行するとデータが蓄積されます。")
    else:
        # 品番選択
        products = sorted(history_df["product_code"].dropna().unique().tolist())
        selected = st.multiselect("品番を選択", products, default=products[:5] if len(products) > 5 else products)

        if selected:
            filtered_history = history_df[history_df["product_code"].isin(selected)].copy()
            filtered_history["checked_date"] = pd.to_datetime(filtered_history["checked_date"])

            # データが2日以上あればチャート表示
            date_count = filtered_history["checked_date"].nunique()
            if date_count >= 2:
                # 最安値の推移チャート
                chart_data = filtered_history.pivot_table(
                    index="checked_date", columns="product_code", values="min_total"
                )
                st.caption("最安値（合計）の推移")
                st.line_chart(chart_data)

            # テーブル表示
            display_cols = [
                "checked_date", "product_code", "list_price", "listing_count",
                "min_total", "min_ratio", "max_total", "avg_total", "avg_ratio",
            ]
            existing_cols = [c for c in display_cols if c in filtered_history.columns]
            st.dataframe(
                filtered_history[existing_cols].sort_values(
                    ["product_code", "checked_date"], ascending=[True, False]
                ),
                use_container_width=True,
                column_config={
                    "checked_date": st.column_config.DateColumn("調査日"),
                    "product_code": "品番",
                    "list_price": st.column_config.NumberColumn("定価", format="¥%d"),
                    "listing_count": "出品数",
                    "min_total": st.column_config.NumberColumn("最安値", format="¥%d"),
                    "min_ratio": st.column_config.NumberColumn("最安_定価比率", format="%.1f%%"),
                    "max_total": st.column_config.NumberColumn("最高値", format="¥%d"),
                    "avg_total": st.column_config.NumberColumn("平均", format="¥%d"),
                    "avg_ratio": st.column_config.NumberColumn("平均_定価比率", format="%.1f%%"),
                },
            )
