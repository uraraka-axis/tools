#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
楽天市場 カテゴリ抽出ツール - Streamlit版
Webブラウザから使用可能なバージョン
"""

import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict
import random
import time
import re
import io


@dataclass
class Category:
    """カテゴリ情報を保持するクラス"""
    name: str
    category_id: str
    url: str
    count: int
    level: int
    parent_path: List[str] = field(default_factory=list)


class RakutenCategoryScraper:
    """楽天市場カテゴリスクレイパー (requests版)"""

    BASE_URL = "https://www.rakuten.co.jp"

    def __init__(self):
        self.session = requests.Session()
        # ブラウザヘッダーを設定（bot検出回避）
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Cache-Control': 'max-age=0',
            'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
        })
        self.stop_flag = False
        self.categories: List[Category] = []
        self.visited_ids = set()
        self.root_category_name = ""
        self.root_category_id = ""
        self.total_requests = 0
        self.log_messages = []

        # 待機時間設定
        self.min_delay = 1.5
        self.max_delay = 4.0

    def log(self, message: str):
        """ログ出力"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_messages.append(log_message)

    def stop(self):
        """処理を停止"""
        self.stop_flag = True

    def random_delay(self):
        """ランダムな待機時間"""
        delay = random.uniform(self.min_delay, self.max_delay)
        if random.random() < 0.1:
            delay += random.uniform(1.0, 3.0)
        time.sleep(delay)

    def extract_category_id_from_url(self, url: str) -> str:
        """URLからカテゴリIDを抽出"""
        match = re.search(r'/category/(\d+)/?', url)
        if match:
            return match.group(1)
        return ""

    def fetch_page(self, url: str):
        """ページを取得"""
        try:
            self.total_requests += 1

            if self.total_requests > 1:
                self.random_delay()

            response = self.session.get(url, timeout=30)
            response.raise_for_status()

            return BeautifulSoup(response.text, 'html.parser')

        except Exception as e:
            self.log(f"  ⚠️ ページ取得エラー: {e}")
            return None

    def get_root_category_name(self, soup: BeautifulSoup) -> str:
        """ルートカテゴリ名を取得"""
        # パンくずリストから取得
        breadcrumb = soup.find('div', class_='dui-container breadcrumb')
        if breadcrumb:
            items = breadcrumb.find_all('a', class_='item')
            if items:
                return items[-1].get_text(strip=True)

        # -activeクラスを持つspanから取得
        active = soup.find('span', class_=re.compile(r'-active'))
        if active:
            return active.get_text(strip=True)

        # h1から取得
        h1 = soup.find('h1')
        if h1:
            return h1.get_text(strip=True)

        return "カテゴリ"

    def get_subcategories_from_page(self, soup: BeautifulSoup, current_category_id: str) -> List[Dict]:
        """ページからサブカテゴリを抽出"""
        subcategories = []

        # ジャンルフィルターを探す
        genre_filter = None

        # サイドバーから探す
        sidebar = soup.find('div', class_=re.compile(r'sidebar|side-menu'))
        if sidebar:
            genre_section = sidebar.find('div', class_=re.compile(r'genre|category'))
            if genre_section:
                genre_filter = genre_section

        if not genre_filter:
            genre_filter = soup.find('div', class_=re.compile(r'genrefilter|genre_filter|genre-list'))

        if not genre_filter:
            genre_filter = soup.find('div', class_='dui-filter-menu')

        if not genre_filter:
            self.log("    ⚠️ ジャンルセクションが見つかりませんでした")
            # HTMLフォールバック
            return self._extract_categories_from_html(soup, current_category_id)

        # 現在のカテゴリ（-activeクラスを持つspan）を探す
        active_element = genre_filter.find(['span', 'div'], class_=re.compile(r'-active'))

        if not active_element:
            self.log("    ⚠️ 現在のカテゴリ要素が見つかりませんでした")
            return self._extract_categories_from_html(soup, current_category_id)

        # 現在のカテゴリの次の兄弟要素（div.item）を取得
        child_container = active_element.find_next_sibling('div', class_='item')

        if not child_container:
            self.log("    ℹ️ このカテゴリには子カテゴリがありません")
            return []

        # 子コンテナ内のdui-listを探す
        child_list = child_container.find('div', class_='dui-list')

        if not child_list:
            self.log("    ℹ️ このカテゴリには子カテゴリがありません")
            return []

        # 子リスト内のリンクを取得
        category_links = child_list.find_all('a', href=re.compile(r'/category/\d+/?'))

        self.log(f"    検出リンク数: {len(category_links)}, 現在ID: {current_category_id}")

        for link in category_links:
            href = link.get('href', '')

            category_id = self.extract_category_id_from_url(href)
            if not category_id:
                continue

            if category_id == current_category_id:
                continue

            if category_id in self.visited_ids:
                continue

            # -activeクラスを持つリンクは除外
            link_classes = link.get('class', [])
            if link_classes and '-active' in ' '.join(link_classes):
                continue

            # 名前を取得
            name = link.get('title', '')
            if not name:
                name_elem = link.find('div', class_='_ellipsis')
                if name_elem:
                    name = name_elem.get_text(strip=True)
                else:
                    name = link.get_text(strip=True)

            # 件数を除去
            name = re.sub(r'\s*[\(（]\s*[\d,]+\s*[件点]\s*[\)）]\s*$', '', name)
            name = name.strip()

            if not name:
                continue

            # URLを正規化
            if href.startswith('//'):
                full_url = 'https:' + href
            elif href.startswith('/'):
                full_url = self.BASE_URL + href
            else:
                full_url = href

            subcategories.append({
                'name': name,
                'url': full_url,
                'category_id': category_id,
                'count': 0
            })

        # 重複除去
        seen = set()
        unique = []
        for cat in subcategories:
            if cat['category_id'] not in seen and cat['name']:
                seen.add(cat['category_id'])
                unique.append(cat)

        return unique

    def _extract_categories_from_html(self, soup: BeautifulSoup, current_category_id: str) -> List[Dict]:
        """HTMLから直接カテゴリリンクを抽出（フォールバック）"""
        subcategories = []
        seen_ids = set()

        # カテゴリリンクのパターン
        category_pattern = re.compile(r'/category/(\d+)/?')

        # すべてのカテゴリリンクを探す
        for link in soup.find_all('a', href=category_pattern):
            href = link.get('href', '')
            match = category_pattern.search(href)
            if not match:
                continue

            category_id = match.group(1)

            # 自分自身は除外
            if category_id == current_category_id:
                continue

            # 重複除去
            if category_id in seen_ids:
                continue
            seen_ids.add(category_id)

            # 名前を取得
            name = link.get('title', '') or link.get_text(strip=True)

            # 件数を除去
            name = re.sub(r'\s*[\(（]\s*[\d,]+\s*[件点]\s*[\)）]\s*$', '', name)
            name = re.sub(r'[\d,]+件$', '', name).strip()

            if not name:
                continue

            # 無効な名前をスキップ
            if name in ['もっと見る', 'すべて見る', '詳細を見る', '閉じる']:
                continue

            # URLを正規化
            if href.startswith('//'):
                full_url = 'https:' + href
            elif href.startswith('/'):
                full_url = self.BASE_URL + href
            else:
                full_url = href

            subcategories.append({
                'name': name,
                'url': full_url,
                'category_id': category_id,
                'count': 0
            })

        self.log(f"    [DEBUG] HTMLから {len(subcategories)} 件のカテゴリリンクを検出")
        return subcategories

    def scrape_categories_recursive(
        self,
        url: str,
        level: int = 0,
        parent_path: List[str] = None,
        max_depth: int = 5,
        progress_callback=None
    ):
        """カテゴリを再帰的に取得"""
        if parent_path is None:
            parent_path = []

        if level > max_depth or self.stop_flag:
            return

        indent = "  " * level
        self.log(f"{indent}📂 取得中: {url}")

        soup = self.fetch_page(url)
        if not soup:
            return

        current_id = self.extract_category_id_from_url(url)
        self.visited_ids.add(current_id)

        is_root = (level == 0)
        if is_root:
            self.root_category_name = self.get_root_category_name(soup)
            self.root_category_id = current_id
            self.log(f"📌 ルートカテゴリ: {self.root_category_name} (ID: {current_id})")

        subcategories = self.get_subcategories_from_page(soup, current_id)

        self.log(f"{indent}  → {len(subcategories)}件のサブカテゴリを発見")

        for subcat in subcategories:
            if self.stop_flag:
                break

            if subcat['category_id'] in self.visited_ids:
                continue

            self.visited_ids.add(subcat['category_id'])

            cat = Category(
                name=subcat['name'],
                category_id=subcat['category_id'],
                url=subcat['url'],
                count=subcat['count'],
                level=level + 1,
                parent_path=parent_path.copy()
            )
            self.categories.append(cat)

            self.log(f"{indent}  ✓ {subcat['name']} [ID: {subcat['category_id']}]")

            # 進捗コールバック
            if progress_callback:
                progress_callback(len(self.categories), parent_path + [subcat['name']])

            # 再帰
            if level + 1 < max_depth:
                new_parent_path = parent_path + [subcat['name']]
                self.scrape_categories_recursive(
                    subcat['url'],
                    level + 1,
                    new_parent_path,
                    max_depth,
                    progress_callback
                )

    def scrape(self, start_url: str, max_depth: int = 5, progress_callback=None) -> List[Category]:
        """スクレイピング開始"""
        self.stop_flag = False
        self.categories = []
        self.visited_ids = set()
        self.total_requests = 0
        self.log_messages = []

        self.log("=" * 50)
        self.log("🛒 楽天市場 カテゴリ抽出開始")
        self.log("=" * 50)
        self.log(f"🔗 URL: {start_url}")
        self.log(f"📊 最大取得階層: {max_depth}")
        self.log("")

        self.scrape_categories_recursive(start_url, max_depth=max_depth, progress_callback=progress_callback)

        if not self.stop_flag:
            self.log("")
            self.log(f"✅ 合計 {len(self.categories)} カテゴリを取得しました")
            self.log(f"📡 総リクエスト数: {self.total_requests}")

        return self.categories

    def export_to_excel(self) -> bytes:
        """Excelファイルをバイトストリームで出力"""
        if not self.categories:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "カテゴリ一覧"

        base_font = Font(name="Meiryo UI", size=10)
        header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="BF0000")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(name="Meiryo UI", bold=True, size=14, color="BF0000")

        thin_border = Border(
            left=Side(style='thin', color='595959'),
            right=Side(style='thin', color='595959'),
            top=Side(style='thin', color='595959'),
            bottom=Side(style='thin', color='595959')
        )

        max_level = max((cat.level for cat in self.categories), default=1)

        level_counts = defaultdict(int)
        for cat in self.categories:
            level_counts[cat.level] += 1

        summary_level_col = 2 + max_level + 4
        summary_count_col = 2 + max_level + 5

        title_col_end = get_column_letter(summary_count_col)
        ws.merge_cells(f'B1:{title_col_end}1')
        ws['B1'] = f"【楽天市場】{self.root_category_name}のジャンル一覧"
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 8

        headers = ["#", "ジャンル1"]
        for i in range(max_level):
            headers.append(f"ジャンル{i + 2}")
        headers.append("カテゴリID")
        headers.append("ページURL")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[3].height = 24

        summary_level_header = ws.cell(row=3, column=summary_level_col, value="レベル")
        summary_level_header.font = header_font
        summary_level_header.fill = header_fill
        summary_level_header.alignment = header_alignment
        summary_level_header.border = thin_border

        summary_count_header = ws.cell(row=3, column=summary_count_col, value="カテゴリ数")
        summary_count_header.font = header_font
        summary_count_header.fill = header_fill
        summary_count_header.alignment = header_alignment
        summary_count_header.border = thin_border

        prev_values = [""] * (max_level + 1)

        for idx, cat in enumerate(self.categories, 1):
            row = idx + 3

            current_values = [self.root_category_name] + [""] * max_level

            for i, parent_name in enumerate(cat.parent_path):
                if i < max_level:
                    current_values[i + 1] = parent_name

            if cat.level <= max_level:
                current_values[cat.level] = cat.name

            cell = ws.cell(row=row, column=1, value=idx)
            cell.border = thin_border
            cell.font = base_font

            for col, value in enumerate(current_values, 2):
                cell = ws.cell(row=row, column=col)

                show_value = value
                if idx > 1 and col - 2 < len(prev_values):
                    if value == prev_values[col - 2]:
                        has_change = any(
                            current_values[j] != prev_values[j]
                            for j in range(col - 1, len(current_values))
                            if j < len(prev_values)
                        )
                        if not has_change:
                            show_value = ""

                cell.value = show_value
                cell.border = thin_border
                cell.font = base_font

            id_col = 2 + max_level + 1
            id_cell = ws.cell(row=row, column=id_col, value=cat.category_id)
            id_cell.border = thin_border
            id_cell.font = base_font

            url_col = 2 + max_level + 2
            url_cell = ws.cell(row=row, column=url_col, value=cat.url)
            url_cell.hyperlink = cat.url
            url_cell.style = "Hyperlink"
            url_cell.border = thin_border
            url_cell.font = Font(name="Meiryo UI", size=10, color="0563C1", underline="single")

            prev_values = current_values.copy()

        summary_row = 4

        level_cell = ws.cell(row=summary_row, column=summary_level_col, value="ジャンル1")
        level_cell.border = thin_border
        level_cell.font = base_font

        count_cell = ws.cell(row=summary_row, column=summary_count_col, value=1)
        count_cell.border = thin_border
        count_cell.font = base_font
        count_cell.alignment = Alignment(horizontal="right")

        summary_row += 1

        for level in sorted(level_counts.keys()):
            level_cell = ws.cell(row=summary_row, column=summary_level_col, value=f"ジャンル{level + 1}")
            level_cell.border = thin_border
            level_cell.font = base_font

            count_cell = ws.cell(row=summary_row, column=summary_count_col, value=level_counts[level])
            count_cell.border = thin_border
            count_cell.font = base_font
            count_cell.alignment = Alignment(horizontal="right")

            summary_row += 1

        total_cell = ws.cell(row=summary_row, column=summary_level_col, value="合計")
        total_cell.border = thin_border
        total_cell.font = Font(name="Meiryo UI", size=10, bold=True)

        total_count_cell = ws.cell(row=summary_row, column=summary_count_col, value=len(self.categories) + 1)
        total_count_cell.border = thin_border
        total_count_cell.font = Font(name="Meiryo UI", size=10, bold=True)
        total_count_cell.alignment = Alignment(horizontal="right")

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        for i in range(max_level):
            col_letter = get_column_letter(3 + i)
            ws.column_dimensions[col_letter].width = 22
        ws.column_dimensions[get_column_letter(3 + max_level)].width = 12
        # URL列の幅を最長URLに合わせて調整
        max_url_length = max((len(cat.url) for cat in self.categories), default=50)
        url_col_width = min(max(max_url_length * 1.1, 50), 120)
        ws.column_dimensions[get_column_letter(4 + max_level)].width = url_col_width
        ws.column_dimensions[get_column_letter(5 + max_level)].width = 3
        ws.column_dimensions[get_column_letter(summary_level_col)].width = 12
        ws.column_dimensions[get_column_letter(summary_count_col)].width = 12

        ws.freeze_panes = 'A4'

        # バイトストリームとして出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def check_password():
    """パスワード認証"""

    # secretsにパスワードが設定されているか確認
    if "password" not in st.secrets:
        # パスワード未設定の場合はそのままアクセス可能
        return True

    correct_password = st.secrets["password"]

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown("""
        <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 40px;
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        </style>
    """, unsafe_allow_html=True)

    st.title("🔐 ログイン")
    st.write("このツールを使用するにはパスワードが必要です。")

    password = st.text_input("パスワード", type="password", key="password_input")

    if st.button("ログイン", type="primary"):
        if password == correct_password:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが違います")

    return False


def main():
    """メイン関数"""

    st.set_page_config(
        page_title="楽天市場 カテゴリ抽出ツール",
        page_icon="🛒",
        layout="wide"
    )

    # パスワード認証
    if not check_password():
        return

    # カスタムCSS
    st.markdown("""
        <style>
        .main-header {
            background: linear-gradient(135deg, #BF0000 0%, #990000 100%);
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        .main-header h1 {
            color: white !important;
            margin: 0;
        }
        .main-header p {
            color: #FFD4D4;
            margin: 5px 0 0 0;
        }
        .stats-card {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }
        .stats-value {
            font-size: 2em;
            font-weight: bold;
            color: #BF0000;
        }
        </style>
    """, unsafe_allow_html=True)

    # ヘッダー
    st.markdown("""
        <div class="main-header">
            <h1>🛒 楽天市場 カテゴリ抽出ツール</h1>
            <p>Rakuten Category Extractor</p>
        </div>
    """, unsafe_allow_html=True)

    # セッションステート初期化
    if "scraper" not in st.session_state:
        st.session_state.scraper = None
    if "is_running" not in st.session_state:
        st.session_state.is_running = False
    if "excel_data" not in st.session_state:
        st.session_state.excel_data = None
    if "log_messages" not in st.session_state:
        st.session_state.log_messages = []
    if "total_categories" not in st.session_state:
        st.session_state.total_categories = 0

    # 入力フォーム
    with st.container():
        st.subheader("📝 設定")

        col1, col2 = st.columns([3, 1])

        with col1:
            url = st.text_input(
                "カテゴリURL",
                value="https://www.rakuten.co.jp/category/101354/",
                help="例: https://www.rakuten.co.jp/category/101354/ （DVD）"
            )

        with col2:
            depth = st.number_input(
                "取得階層数",
                min_value=1,
                max_value=10,
                value=3,
                help="1〜10階層まで指定可能"
            )

    # ボタンエリア
    start_disabled = st.session_state.is_running
    start_clicked = st.button("🚀 抽出開始", disabled=start_disabled, type="primary")

    # 抽出処理
    if start_clicked and url:
        st.session_state.is_running = True
        st.session_state.excel_data = None
        st.session_state.log_messages = []
        st.session_state.total_categories = 0

        # スクレイパー実行
        scraper = RakutenCategoryScraper()
        st.session_state.scraper = scraper

        progress_bar = st.progress(0)
        status_text = st.empty()
        log_container = st.empty()

        def update_progress(count, path):
            st.session_state.total_categories = count
            status_text.text(f"取得中: {count}件 | {' > '.join(path)}")

        try:
            with st.spinner("カテゴリを取得中..."):
                categories = scraper.scrape(url, max_depth=depth, progress_callback=update_progress)

            st.session_state.log_messages = scraper.log_messages

            if categories:
                st.session_state.excel_data = scraper.export_to_excel()
                st.session_state.total_categories = len(categories)
                st.success(f"✅ {len(categories)}件のカテゴリを取得しました！")
            else:
                st.warning("カテゴリが取得できませんでした")

        except Exception as e:
            st.error(f"エラーが発生しました: {e}")
        finally:
            st.session_state.is_running = False
            st.rerun()

    elif start_clicked and not url:
        st.error("URLを入力してください")

    # 統計情報
    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f"""
            <div class="stats-card">
                <p>取得カテゴリ数</p>
                <div class="stats-value">{st.session_state.total_categories}件</div>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        # ダウンロードボタン
        if st.session_state.excel_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📥 Excelファイルをダウンロード",
                data=st.session_state.excel_data,
                file_name=f"rakuten_categories_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

    # ログ表示
    if st.session_state.log_messages:
        with st.expander("📋 ログ", expanded=False):
            log_text = "\n".join(st.session_state.log_messages)
            st.code(log_text, language=None)


if __name__ == "__main__":
    main()
