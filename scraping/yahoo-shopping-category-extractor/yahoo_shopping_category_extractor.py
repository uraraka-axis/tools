#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Yahoo!ショッピング カテゴリ抽出ツール
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin
from typing import Optional, List, Dict, Callable
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict
import threading
import random
import time
import re
import os
import json


@dataclass
class Category:
    """カテゴリ情報を保持するクラス"""
    name: str
    category_id: str
    url: str
    count: int
    level: int
    parent_path: List[str] = field(default_factory=list)


@dataclass
class ProcessingStats:
    """処理統計情報"""
    start_time: float = 0.0
    categories_by_level: Dict[int, int] = field(default_factory=lambda: defaultdict(int))
    total_categories: int = 0
    current_path: List[str] = field(default_factory=list)
    requests_count: int = 0
    
    def get_elapsed_time(self) -> str:
        """経過時間を取得"""
        if self.start_time == 0:
            return "00:00:00"
        elapsed = int(time.time() - self.start_time)
        hours = elapsed // 3600
        minutes = (elapsed % 3600) // 60
        seconds = elapsed % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


class YahooCategoryScraper:
    """Yahoo!ショッピングカテゴリスクレイパー"""
    
    BASE_URL = "https://shopping.yahoo.co.jp"
    
    def __init__(self, log_callback: Callable = None, progress_callback: Callable = None):
        self.session = requests.Session()
        self.driver = None
        self.log_callback = log_callback or print
        self.progress_callback = progress_callback
        self.stop_flag = False
        self.categories: List[Category] = []
        self.root_category_name = ""
        self.root_category_id = ""
        self.total_requests = 0
        self.stats = ProcessingStats()
        
        # ランダム待機時間の設定（秒）
        self.min_delay = 1.5
        self.max_delay = 4.0
    
    def log(self, message: str):
        """ログ出力"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_callback(log_message)
    
    def update_progress(self, stats: ProcessingStats):
        """プログレス更新"""
        if self.progress_callback:
            self.progress_callback(stats)
    
    def stop(self):
        """処理を停止"""
        self.stop_flag = True
        self.log("⏸️ 停止リクエストを受信しました")
    
    def random_delay(self):
        """ランダムな待機時間（bot判定回避）"""
        delay = random.uniform(self.min_delay, self.max_delay)
        # 時々長めの待機を入れる（より人間らしく）
        if random.random() < 0.1:
            delay += random.uniform(1.0, 3.0)
        time.sleep(delay)
    
    def setup_driver(self):
        """Seleniumドライバーを初期化"""
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                use_manager = True
            except ImportError:
                use_manager = False
            
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-extensions')
            options.add_argument('--disable-logging')
            options.add_argument('--log-level=3')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option('excludeSwitches', ['enable-automation'])
            options.add_experimental_option('useAutomationExtension', False)
            
            # ランダムなUser-Agent
            user_agents = [
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
            ]
            options.add_argument(f'--user-agent={random.choice(user_agents)}')
            
            if use_manager:
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
            else:
                self.driver = webdriver.Chrome(options=options)
            
            self.driver.set_page_load_timeout(30)
            
            # webdriver検出回避
            self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                'source': '''
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                '''
            })
            
            self.log("✅ ブラウザ起動完了")
            return True
            
        except Exception as e:
            self.log(f"❌ ブラウザ起動失敗: {e}")
            self.log("💡 pip install selenium webdriver-manager を実行してください")
            return False
    
    def close_driver(self):
        """Seleniumドライバーを終了"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
    
    def extract_category_id_from_url(self, url: str) -> str:
        """URLからカテゴリIDパスを抽出"""
        # /category/2517/881/883/list → 2517/881/883
        # /category/881/list → 881
        # /category/2517/list → 2517
        match = re.search(r'/category/([\d/]+)/list', url)
        if match:
            return match.group(1).rstrip('/')
        return ""
    
    def get_last_category_id(self, category_path: str) -> str:
        """カテゴリパスから最後のIDを取得"""
        # 2517/881/883 → 883
        # 881 → 881
        if '/' in category_path:
            return category_path.split('/')[-1]
        return category_path
    
    def fetch_page(self, url: str) -> Optional[BeautifulSoup]:
        """ページを取得"""
        try:
            self.total_requests += 1
            self.stats.requests_count += 1
            
            # ランダム待機
            if self.total_requests > 1:
                self.random_delay()
            
            self.driver.get(url)
            
            # ページ読み込み待機（ランダム）
            time.sleep(random.uniform(1.0, 2.0))
            
            # 「もっと見る」ボタンをクリックして全カテゴリを表示
            try:
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                
                # 複数の「もっと見る」ボタンがある場合があるので全部クリック
                more_buttons = self.driver.find_elements(By.XPATH, "//button[contains(@class, 'toggleButton')]")
                for btn in more_buttons:
                    try:
                        self.driver.execute_script("arguments[0].click();", btn)
                        time.sleep(random.uniform(0.3, 0.6))
                    except:
                        pass
            except:
                pass
            
            # たまにスクロールする（人間らしい動作）
            if random.random() < 0.3:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 3);")
                time.sleep(random.uniform(0.3, 0.8))
            
            html = self.driver.page_source
            return BeautifulSoup(html, 'html.parser')
            
        except Exception as e:
            self.log(f"  ⚠️ ページ取得エラー: {e}")
            return None
    
    def parse_category_count(self, text: str) -> int:
        """カテゴリ名から件数を抽出"""
        # "708,280件" や "708,280" から数値を抽出
        match = re.search(r'([\d,]+)', text.replace('件', ''))
        if match:
            return int(match.group(1).replace(',', ''))
        return 0
    
    def extract_category_name(self, text: str) -> str:
        """カテゴリ名から件数を除去"""
        return re.sub(r'[\d,]+件?$', '', text).strip()
    
    def get_root_category_name(self, soup: BeautifulSoup) -> str:
        """ルートカテゴリ名を取得"""
        # h1タグから取得
        h1 = soup.find('h1')
        if h1:
            name = h1.get_text(strip=True)
            # 不要な接尾辞を除去
            name = re.sub(r'映像ソフト$|おすすめ.*$', '', name).strip()
            return name
        
        # カテゴリセクションのタイトルから取得を試みる
        title_span = soup.find('span', class_=re.compile(r'listItemTitle'))
        if title_span:
            parent_div = title_span.find_parent('div', class_=re.compile(r'listItem(?!--)'))
            if parent_div and not parent_div.find_parent('a'):
                return title_span.get_text(strip=True)
        
        return "カテゴリ"
    
    def get_subcategories_from_page(self, soup: BeautifulSoup, current_category_id: str, is_root: bool = False) -> List[Dict]:
        """ページからサブカテゴリを抽出（__NEXT_DATA__のJSONから取得）"""
        subcategories = []

        # __NEXT_DATA__ scriptタグからJSONデータを取得
        next_data_script = soup.find('script', id='__NEXT_DATA__')
        if not next_data_script:
            self.log("    [DEBUG] __NEXT_DATA__ が見つかりません。旧方式で試行します。")
            return self._get_subcategories_legacy(soup, current_category_id, is_root)

        try:
            json_data = json.loads(next_data_script.string)

            # カテゴリデータへのパスを探索
            categories_data = self._extract_categories_from_json(json_data)

            if not categories_data:
                self.log("    [DEBUG] JSONからカテゴリデータを取得できませんでした")
                return []

            self.log(f"    [DEBUG] JSONから {len(categories_data)} 件のカテゴリを検出")

            for cat_data in categories_data:
                name = cat_data.get('text', '')
                url = cat_data.get('url', '')
                count = cat_data.get('count', 0)

                if not name or not url:
                    continue

                # URLからカテゴリIDを抽出
                category_path = self.extract_category_id_from_url(url)
                if not category_path:
                    continue

                last_id = self.get_last_category_id(category_path)

                # URLを正規化
                if not url.startswith('http'):
                    url = self.BASE_URL + url

                # クエリパラメータを除去
                url = re.sub(r'\?.*$', '', url)
                if not url.endswith('/list'):
                    url = url.rstrip('/') + '/list'

                subcategories.append({
                    'name': name,
                    'url': url,
                    'category_id': category_path,
                    'last_id': last_id,
                    'count': count
                })

        except json.JSONDecodeError as e:
            self.log(f"    [DEBUG] JSON解析エラー: {e}")
            return []
        except Exception as e:
            self.log(f"    [DEBUG] カテゴリ抽出エラー: {e}")
            return []

        # 重複除去（last_idでユニーク化）
        seen = set()
        unique = []
        for cat in subcategories:
            if cat['last_id'] not in seen and cat['name']:
                seen.add(cat['last_id'])
                unique.append(cat)

        return unique

    def _extract_categories_from_json(self, json_data: dict) -> List[Dict]:
        """JSONデータからカテゴリ情報を抽出"""
        categories = []

        try:
            # props > pageProps > initialState > bff > advancedFilter > sections > category > categories
            initial_state = json_data.get('props', {}).get('pageProps', {}).get('initialState', {})
            bff = initial_state.get('bff', {})
            advanced_filter = bff.get('advancedFilter', {})
            sections = advanced_filter.get('sections', {})
            category_section = sections.get('category', {})
            categories_data = category_section.get('categories', {})

            # suggestedCategories から取得
            suggested = categories_data.get('suggestedCategories', [])
            if suggested:
                categories.extend(suggested)

            # toggleAreaCategoryItems から取得（「もっと見る」で表示されるカテゴリ）
            toggle_items = categories_data.get('toggleAreaCategoryItems', [])
            if toggle_items:
                categories.extend(toggle_items)

            # childCategories がある場合も取得
            for cat in suggested + toggle_items:
                child_cats = cat.get('childCategories', [])
                if child_cats:
                    # 子カテゴリは別途再帰取得されるので、ここでは追加しない
                    pass

        except Exception as e:
            self.log(f"    [DEBUG] JSON構造解析エラー: {e}")

        return categories

    def _get_subcategories_legacy(self, soup: BeautifulSoup, current_category_id: str, is_root: bool = False) -> List[Dict]:
        """旧方式でサブカテゴリを抽出（フォールバック用）"""
        subcategories = []

        # 現在のカテゴリの最後のID
        current_last_id = self.get_last_category_id(current_category_id)
        current_path_parts = current_category_id.split('/')
        current_depth = len(current_path_parts)

        # カテゴリリンクを探す
        category_links = soup.find_all('a', href=re.compile(r'/category/[\d/]+/list'))

        self.log(f"    [DEBUG] 旧方式: 検出リンク数: {len(category_links)}")

        for link in category_links:
            href = link.get('href', '')

            # カテゴリIDパスを抽出
            category_path = self.extract_category_id_from_url(href)
            if not category_path:
                continue

            # 最後のカテゴリIDを取得
            last_id = self.get_last_category_id(category_path)
            path_parts = category_path.split('/')

            # 自分自身は除外
            if category_path == current_category_id:
                continue

            is_child = False

            if is_root:
                if category_path.startswith(current_category_id + '/'):
                    if len(path_parts) == current_depth + 1:
                        is_child = True
            else:
                if current_last_id in path_parts:
                    current_idx = path_parts.index(current_last_id)
                    if current_idx == len(path_parts) - 2:
                        is_child = True

            if not is_child:
                continue

            # カテゴリ名を取得（テキストノードから）
            name = link.get_text(strip=True)
            if not name:
                name = link.get('title', '')
            if not name:
                continue

            # 件数を除去
            name = self.extract_category_name(name)

            # URLを正規化
            if href.startswith('//'):
                full_url = 'https:' + href
            elif href.startswith('/'):
                full_url = self.BASE_URL + href
            else:
                full_url = href

            full_url = re.sub(r'\?.*$', '', full_url)
            if not full_url.endswith('/list'):
                full_url = full_url.rstrip('/') + '/list'

            subcategories.append({
                'name': name,
                'url': full_url,
                'category_id': category_path,
                'last_id': last_id,
                'count': 0
            })

        # 重複除去
        seen = set()
        unique = []
        for cat in subcategories:
            if cat['last_id'] not in seen and cat['name']:
                seen.add(cat['last_id'])
                unique.append(cat)

        return unique
    
    def scrape_categories_recursive(
        self,
        url: str,
        level: int = 0,
        parent_path: List[str] = None,
        max_depth: int = 5,
        parent_id: str = ""
    ):
        """カテゴリを再帰的に取得"""
        if parent_path is None:
            parent_path = []
        
        if level > max_depth or self.stop_flag:
            return
        
        indent = "  " * level
        self.log(f"{indent}📂 取得中: {url}")
        
        self.stats.current_path = parent_path.copy()
        self.update_progress(self.stats)
        
        soup = self.fetch_page(url)
        if not soup:
            return
        
        current_id = self.extract_category_id_from_url(url)
        current_last_id = self.get_last_category_id(current_id)
        
        # ルートカテゴリ情報
        is_root = (level == 0)
        if is_root:
            self.root_category_name = self.get_root_category_name(soup)
            self.root_category_id = current_id
            self.log(f"📌 ルートカテゴリ: {self.root_category_name} (ID: {current_id})")
        
        # サブカテゴリ取得（現在のカテゴリIDとルートフラグを渡す）
        subcategories = self.get_subcategories_from_page(soup, current_id, is_root=is_root)
        
        self.log(f"{indent}  → {len(subcategories)}件のサブカテゴリを発見")
        
        for subcat in subcategories:
            if self.stop_flag:
                break
            
            cat = Category(
                name=subcat['name'],
                category_id=subcat['category_id'],
                url=subcat['url'],
                count=subcat['count'],
                level=level + 1,
                parent_path=parent_path.copy()
            )
            self.categories.append(cat)
            
            # 統計更新
            self.stats.total_categories += 1
            self.stats.categories_by_level[level + 1] += 1
            self.stats.current_path = parent_path + [subcat['name']]
            self.update_progress(self.stats)
            
            self.log(f"{indent}  ✓ {subcat['name']} ({subcat['count']:,}件) [ID: {subcat['last_id']}]")
            
            # 再帰（次の階層を取得）
            if level + 1 < max_depth:
                new_parent_path = parent_path + [subcat['name']]
                self.scrape_categories_recursive(
                    subcat['url'],
                    level + 1,
                    new_parent_path,
                    max_depth,
                    subcat['category_id']
                )
    
    def scrape(self, start_url: str, max_depth: int = 5) -> List[Category]:
        """スクレイピング開始"""
        self.stop_flag = False
        self.categories = []
        self.total_requests = 0
        self.stats = ProcessingStats()
        self.stats.start_time = time.time()
        
        self.log("\n" + "=" * 50)
        self.log("🛒 Yahoo!ショッピング カテゴリ抽出開始")
        self.log("=" * 50)
        self.log(f"🔗 URL: {start_url}")
        self.log(f"📊 最大取得階層: {max_depth}")
        self.log(f"⏱️ 待機時間: {self.min_delay}〜{self.max_delay}秒（ランダム）")
        self.log("")
        
        if not self.setup_driver():
            raise Exception("ブラウザの起動に失敗しました")
        
        try:
            self.scrape_categories_recursive(start_url, max_depth=max_depth)
        finally:
            self.close_driver()
        
        if not self.stop_flag:
            self.log("")
            self.log(f"✅ 合計 {len(self.categories)} カテゴリを取得しました")
            self.log(f"📡 総リクエスト数: {self.total_requests}")
            self.log(f"⏱️ 処理時間: {self.stats.get_elapsed_time()}")
        
        return self.categories
    
    def export_to_excel(self, output_path: str):
        """Excelファイルに出力（楽天版と同じ形式）"""
        if not self.categories:
            self.log("⚠️ カテゴリが取得されていません")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ジャンル一覧"
        
        # Meiryo UIフォント
        base_font = Font(name="Meiryo UI", size=10)
        header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="ff0033")  # Yahoo!の赤色
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(name="Meiryo UI", bold=True, size=14, color="ff0033")
        
        # 罫線
        thin_border = Border(
            left=Side(style='thin', color='595959'),
            right=Side(style='thin', color='595959'),
            top=Side(style='thin', color='595959'),
            bottom=Side(style='thin', color='595959')
        )
        
        max_level = max((cat.level for cat in self.categories), default=1)
        
        # レベル別カテゴリ数を集計
        level_counts = defaultdict(int)
        for cat in self.categories:
            level_counts[cat.level] += 1
        
        # 集計表の列位置を計算（G列を空列にし、H列とI列に配置）
        summary_level_col = 2 + max_level + 4  # カテゴリID、ページURL、空列の後
        summary_count_col = 2 + max_level + 5
        
        # タイトルは集計表まで含める
        title_col_end = get_column_letter(summary_count_col)
        ws.merge_cells(f'B1:{title_col_end}1')
        ws['B1'] = f"【Yahoo!ショッピング】{self.root_category_name}のジャンル一覧"
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 8
        
        # ヘッダー：ジャンル1, ジャンル2, ジャンル3...
        headers = ["#", "ジャンル1"]
        for i in range(max_level):
            headers.append(f"ジャンル{i + 2}")
        headers.append("カテゴリID")
        headers.append("ページURL")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[3].height = 24
        
        # 集計表ヘッダー
        summary_level_header = ws.cell(row=3, column=summary_level_col, value="レベル")
        summary_level_header.font = header_font
        summary_level_header.fill = header_fill
        summary_level_header.alignment = header_alignment
        summary_level_header.border = thin_border
        
        summary_count_header = ws.cell(row=3, column=summary_count_col, value="カテゴリ数")
        summary_count_header.font = header_font
        summary_count_header.fill = header_fill
        summary_count_header.alignment = header_alignment
        summary_count_header.border = thin_border
        
        prev_values = [""] * (max_level + 1)
        
        for idx, cat in enumerate(self.categories, 1):
            row = idx + 3
            
            current_values = [self.root_category_name] + [""] * max_level
            
            for i, parent_name in enumerate(cat.parent_path):
                if i < max_level:
                    current_values[i + 1] = parent_name
            
            if cat.level <= max_level:
                current_values[cat.level] = cat.name
            
            # 番号
            cell = ws.cell(row=row, column=1, value=idx)
            cell.border = thin_border
            cell.font = base_font
            
            # カテゴリ列
            for col, value in enumerate(current_values, 2):
                cell = ws.cell(row=row, column=col)
                
                show_value = value
                if idx > 1 and col - 2 < len(prev_values):
                    if value == prev_values[col - 2]:
                        has_change = any(
                            current_values[j] != prev_values[j]
                            for j in range(col - 1, len(current_values))
                            if j < len(prev_values)
                        )
                        if not has_change:
                            show_value = ""
                
                cell.value = show_value
                cell.border = thin_border
                cell.font = base_font
            
            # カテゴリID（最下層のIDのみ）
            id_col = 2 + max_level + 1
            last_id = self.get_last_category_id(cat.category_id)
            id_cell = ws.cell(row=row, column=id_col, value=last_id)
            id_cell.border = thin_border
            id_cell.font = base_font
            
            # ページURL（ハイパーリンク設定）
            url_col = 2 + max_level + 2
            url_cell = ws.cell(row=row, column=url_col, value=cat.url)
            url_cell.hyperlink = cat.url
            url_cell.style = "Hyperlink"
            url_cell.border = thin_border
            url_cell.font = Font(name="Meiryo UI", size=10, color="0563C1", underline="single")
            
            prev_values = current_values.copy()
        
        # 集計表を作成
        summary_row = 4  # データの開始行
        
        # ジャンル1（ルートカテゴリ）を追加
        level_cell = ws.cell(row=summary_row, column=summary_level_col, value="ジャンル1")
        level_cell.border = thin_border
        level_cell.font = base_font
        
        count_cell = ws.cell(row=summary_row, column=summary_count_col, value=1)
        count_cell.border = thin_border
        count_cell.font = base_font
        count_cell.alignment = Alignment(horizontal="right")
        
        summary_row += 1
        
        # ジャンル2以降
        for level in sorted(level_counts.keys()):
            level_cell = ws.cell(row=summary_row, column=summary_level_col, value=f"ジャンル{level + 1}")
            level_cell.border = thin_border
            level_cell.font = base_font
            
            count_cell = ws.cell(row=summary_row, column=summary_count_col, value=level_counts[level])
            count_cell.border = thin_border
            count_cell.font = base_font
            count_cell.alignment = Alignment(horizontal="right")
            
            summary_row += 1
        
        # 合計行
        total_cell = ws.cell(row=summary_row, column=summary_level_col, value="合計")
        total_cell.border = thin_border
        total_cell.font = Font(name="Meiryo UI", size=10, bold=True)
        
        total_count_cell = ws.cell(row=summary_row, column=summary_count_col, value=len(self.categories) + 1)  # ルートカテゴリを含める
        total_count_cell.border = thin_border
        total_count_cell.font = Font(name="Meiryo UI", size=10, bold=True)
        total_count_cell.alignment = Alignment(horizontal="right")
        
        # 列幅調整
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        for i in range(max_level):
            col_letter = get_column_letter(3 + i)
            ws.column_dimensions[col_letter].width = 22
        ws.column_dimensions[get_column_letter(3 + max_level)].width = 18  # カテゴリID列
        ws.column_dimensions[get_column_letter(4 + max_level)].width = 50  # URL列
        ws.column_dimensions[get_column_letter(5 + max_level)].width = 3   # 空白列（G列）
        ws.column_dimensions[get_column_letter(summary_level_col)].width = 12  # レベル列（H列）
        ws.column_dimensions[get_column_letter(summary_count_col)].width = 12  # カテゴリ数列（I列）
        
        ws.freeze_panes = 'A4'
        
        wb.save(output_path)
        self.log(f"📄 Excelファイルを保存しました: {output_path}")


class YahooCategoryExtractorGUI:
    """Yahoo!ショッピングカテゴリ抽出GUI"""
    
    # Yahoo!カラー #ff0033
    YAHOO_RED = "#ff0033"
    PRIMARY = "#ff0033"
    PRIMARY_HOVER = "#cc0029"
    BACKGROUND = "#F8F9FA"
    CARD_BG = "#FFFFFF"
    TEXT_PRIMARY = "#212529"
    TEXT_SECONDARY = "#6C757D"
    BORDER = "#DEE2E6"
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Yahoo!ショッピング カテゴリ抽出ツール")
        self.root.geometry("820x850")
        self.root.configure(bg=self.BACKGROUND)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.scraper = None
        self.is_running = False
        self.timer_id = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIセットアップ"""
        container = tk.Frame(self.root, bg=self.BACKGROUND)
        container.pack(fill=tk.BOTH, expand=True, padx=24, pady=20)
        
        self.create_header(container)
        self.create_input_form(container)
        self.create_button_area(container)
        self.create_stats_card(container)
        self.create_log_card(container)
    
    def create_header(self, parent):
        """ヘッダー作成"""
        header = tk.Frame(parent, bg=self.PRIMARY, height=80)
        header.pack(fill=tk.X, pady=(0, 20))
        header.pack_propagate(False)
        
        content = tk.Frame(header, bg=self.PRIMARY)
        content.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        title = tk.Label(
            content,
            text="🛒 Yahoo!ショッピング カテゴリ抽出ツール",
            font=("メイリオ", 16, "bold"),
            bg=self.PRIMARY,
            fg="white"
        )
        title.pack()
        
        subtitle = tk.Label(
            content,
            text="Yahoo! Shopping Category Extractor",
            font=("メイリオ", 9),
            bg=self.PRIMARY,
            fg="#FFE0B2"
        )
        subtitle.pack(pady=(2, 0))
    
    def create_input_form(self, parent):
        """入力フォーム作成"""
        card = self.create_card(parent)
        card.pack(fill=tk.X, pady=(0, 16))
        
        content = tk.Frame(card, bg=self.CARD_BG, padx=20, pady=20)
        content.pack(fill=tk.X)
        
        # URL入力
        url_label = tk.Label(
            content,
            text="カテゴリURL",
            font=("メイリオ", 10, "bold"),
            bg=self.CARD_BG,
            fg=self.TEXT_PRIMARY
        )
        url_label.pack(anchor=tk.W, pady=(0, 6))
        
        url_frame = tk.Frame(content, bg=self.CARD_BG)
        url_frame.pack(fill=tk.X, pady=(0, 4))
        
        self.url_entry = tk.Entry(
            url_frame,
            font=("メイリオ", 10),
            relief=tk.SOLID,
            borderwidth=1,
            highlightthickness=0
        )
        self.url_entry.pack(fill=tk.X, ipady=6)
        self.url_entry.insert(0, "https://shopping.yahoo.co.jp/category/2517/list")
        
        hint = tk.Label(
            content,
            text="例: https://shopping.yahoo.co.jp/category/2517/list （DVD、映像ソフト）",
            font=("メイリオ", 8),
            bg=self.CARD_BG,
            fg=self.TEXT_SECONDARY
        )
        hint.pack(anchor=tk.W, pady=(0, 16))
        
        # 設定行
        settings = tk.Frame(content, bg=self.CARD_BG)
        settings.pack(fill=tk.X)
        
        # 階層数
        depth_frame = tk.Frame(settings, bg=self.CARD_BG)
        depth_frame.pack(side=tk.LEFT, padx=(0, 24))
        
        depth_label = tk.Label(
            depth_frame,
            text="取得階層数",
            font=("メイリオ", 10, "bold"),
            bg=self.CARD_BG,
            fg=self.TEXT_PRIMARY
        )
        depth_label.pack(anchor=tk.W, pady=(0, 6))
        
        self.depth_var = tk.StringVar(value="3")
        depth_spin = tk.Spinbox(
            depth_frame,
            from_=1, to=10,
            textvariable=self.depth_var,
            width=6,
            font=("メイリオ", 10),
            relief=tk.SOLID,
            borderwidth=1
        )
        depth_spin.pack(anchor=tk.W)
        
        # 出力フォルダ
        output_frame = tk.Frame(settings, bg=self.CARD_BG)
        output_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        output_label = tk.Label(
            output_frame,
            text="出力フォルダ",
            font=("メイリオ", 10, "bold"),
            bg=self.CARD_BG,
            fg=self.TEXT_PRIMARY
        )
        output_label.pack(anchor=tk.W, pady=(0, 6))
        
        output_row = tk.Frame(output_frame, bg=self.CARD_BG)
        output_row.pack(fill=tk.X)
        
        self.output_entry = tk.Entry(
            output_row,
            font=("メイリオ", 10),
            relief=tk.SOLID,
            borderwidth=1,
            state='readonly'
        )
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)
        
        browse_btn = tk.Button(
            output_row,
            text="参照",
            command=self.browse_output,
            bg=self.TEXT_SECONDARY,
            fg="white",
            font=("メイリオ", 9),
            relief=tk.FLAT,
            padx=16,
            pady=6,
            cursor="hand2"
        )
        browse_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.add_hover_effect(browse_btn, self.TEXT_SECONDARY, "#5A6268")
    
    def create_button_area(self, parent):
        """ボタンエリア作成"""
        button_container = tk.Frame(parent, bg=self.BACKGROUND)
        button_container.pack(pady=20)
        
        self.start_button = tk.Button(
            button_container,
            text="🚀 抽出開始",
            command=self.start_extraction,
            bg=self.PRIMARY,
            fg="white",
            font=("メイリオ", 11, "bold"),
            relief=tk.FLAT,
            padx=50,
            pady=10,
            cursor="hand2"
        )
        self.start_button.pack(side=tk.LEFT, padx=8)
        self.add_hover_effect(self.start_button, self.PRIMARY, self.PRIMARY_HOVER)
        
        self.stop_button = tk.Button(
            button_container,
            text="⏹ 停止",
            command=self.stop_extraction,
            bg="#6C757D",
            fg="white",
            font=("メイリオ", 11, "bold"),
            relief=tk.FLAT,
            padx=50,
            pady=10,
            state=tk.DISABLED,
            cursor="hand2"
        )
        self.stop_button.pack(side=tk.LEFT, padx=8)
    
    def create_stats_card(self, parent):
        """統計情報カード作成"""
        card = self.create_card(parent)
        card.pack(fill=tk.X, pady=(0, 16))
        
        content = tk.Frame(card, bg=self.CARD_BG, padx=24, pady=20)
        content.pack(fill=tk.X)
        
        # ヘッダー
        header = tk.Label(
            content,
            text="処理状況",
            font=("メイリオ", 11, "bold"),
            bg=self.CARD_BG,
            fg=self.TEXT_PRIMARY
        )
        header.pack(anchor=tk.W, pady=(0, 16))
        
        # メイングリッド
        main_grid = tk.Frame(content, bg=self.CARD_BG)
        main_grid.pack(fill=tk.X)
        
        # 左側：取得数
        left_box = tk.Frame(main_grid, bg="#F8F9FA", relief=tk.FLAT, bd=0)
        left_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        
        left_content = tk.Frame(left_box, bg="#F8F9FA", padx=20, pady=16)
        left_content.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(
            left_content,
            text="取得数",
            font=("メイリオ", 9),
            bg="#F8F9FA",
            fg=self.TEXT_SECONDARY
        ).pack(anchor=tk.W)
        
        self.total_label = tk.Label(
            left_content,
            text="0件",
            font=("メイリオ", 20, "bold"),
            bg="#F8F9FA",
            fg=self.PRIMARY
        )
        self.total_label.pack(anchor=tk.W, pady=(4, 0))
        
        # 右側：処理時間
        right_box = tk.Frame(main_grid, bg="#F8F9FA", relief=tk.FLAT, bd=0)
        right_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        right_content = tk.Frame(right_box, bg="#F8F9FA", padx=20, pady=16)
        right_content.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(
            right_content,
            text="処理時間",
            font=("メイリオ", 9),
            bg="#F8F9FA",
            fg=self.TEXT_SECONDARY
        ).pack(anchor=tk.W)
        
        self.time_label = tk.Label(
            right_content,
            text="00:00:00",
            font=("メイリオ", 20, "bold"),
            bg="#F8F9FA",
            fg=self.TEXT_PRIMARY
        )
        self.time_label.pack(anchor=tk.W, pady=(4, 0))
        
        # 階層別情報
        self.level_label = tk.Label(
            content,
            text="",
            font=("メイリオ", 9),
            bg=self.CARD_BG,
            fg=self.TEXT_SECONDARY,
            anchor=tk.W
        )
        self.level_label.pack(fill=tk.X, pady=(16, 0))
        
        # 区切り線
        separator = tk.Frame(content, bg=self.BORDER, height=1)
        separator.pack(fill=tk.X, pady=(12, 12))
        
        # 処理中パス
        tk.Label(
            content,
            text="処理中",
            font=("メイリオ", 9, "bold"),
            bg=self.CARD_BG,
            fg=self.TEXT_SECONDARY
        ).pack(anchor=tk.W, pady=(0, 6))
        
        self.path_label = tk.Label(
            content,
            text="待機中...",
            font=("メイリオ", 9),
            bg=self.CARD_BG,
            fg=self.TEXT_PRIMARY,
            anchor=tk.W,
            wraplength=720,
            justify=tk.LEFT
        )
        self.path_label.pack(fill=tk.X)
    
    def create_log_card(self, parent):
        """ログカード作成"""
        card = self.create_card(parent)
        card.pack(fill=tk.BOTH, expand=True)
        
        # ヘッダー
        header_frame = tk.Frame(card, bg="#34495E", height=40)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        header_label = tk.Label(
            header_frame,
            text="📋 ログ",
            font=("メイリオ", 10, "bold"),
            bg="#34495E",
            fg="white"
        )
        header_label.pack(side=tk.LEFT, padx=20, pady=10)
        
        # ログエリア
        log_container = tk.Frame(card, bg=self.CARD_BG, padx=12, pady=12)
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(
            log_container,
            font=("Consolas", 9),
            bg="#1E1E1E",
            fg="#D4D4D4",
            relief=tk.FLAT,
            wrap=tk.WORD,
            padx=10,
            pady=10
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
    
    def create_card(self, parent):
        """カードウィジェット作成"""
        card = tk.Frame(
            parent,
            bg=self.CARD_BG,
            relief=tk.FLAT,
            highlightbackground=self.BORDER,
            highlightthickness=1
        )
        return card
    
    def add_hover_effect(self, button, normal_color, hover_color):
        """ボタンホバーエフェクト"""
        button.bind('<Enter>', lambda e: button.config(bg=hover_color))
        button.bind('<Leave>', lambda e: button.config(bg=normal_color) if button['state'] != tk.DISABLED else None)
    
    def browse_output(self):
        """出力フォルダ選択"""
        folder = filedialog.askdirectory(title="出力フォルダを選択")
        if folder:
            self.output_entry.config(state='normal')
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder)
            self.output_entry.config(state='readonly')
    
    def log(self, message: str):
        """ログ出力"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def update_progress(self, stats: ProcessingStats):
        """進捗更新"""
        self.total_label.config(text=f"{stats.total_categories}件")
        
        if stats.categories_by_level:
            level_text = "階層別: " + ", ".join(
                f"Lv{level}={count}件" 
                for level, count in sorted(stats.categories_by_level.items())
            )
            self.level_label.config(text=level_text)
        
        if stats.current_path:
            path_str = " > ".join(stats.current_path)
            self.path_label.config(text=path_str)
        
        self.root.update_idletasks()
    
    def update_timer(self):
        """タイマー更新（1秒ごと）"""
        if self.is_running:
            if self.scraper and hasattr(self.scraper, 'stats'):
                if self.scraper.stats.start_time > 0:
                    elapsed = self.scraper.stats.get_elapsed_time()
                    self.time_label.config(text=elapsed)
                else:
                    self.time_label.config(text="00:00:00")
            
            self.timer_id = self.root.after(1000, self.update_timer)
    
    def start_extraction(self):
        """抽出開始"""
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showerror("エラー", "URLを入力してください")
            return
        
        output_folder = self.output_entry.get().strip()
        if not output_folder:
            messagebox.showerror("エラー", "出力フォルダを選択してください")
            return
        
        try:
            depth = int(self.depth_var.get())
            if depth < 1 or depth > 10:
                raise ValueError()
        except ValueError:
            messagebox.showerror("エラー", "階層数は1〜10の整数を入力してください")
            return
        
        # UI更新
        self.is_running = True
        self.start_button.config(state=tk.DISABLED, bg="#CCCCCC")
        self.stop_button.config(state=tk.NORMAL, bg="#DC3545")
        self.log_text.delete(1.0, tk.END)
        
        # 統計情報をリセット
        self.total_label.config(text="0件")
        self.time_label.config(text="00:00:00")
        self.level_label.config(text="")
        self.path_label.config(text="待機中...")
        
        # バックグラウンドで実行
        thread = threading.Thread(
            target=self.run_extraction,
            args=(url, depth, output_folder)
        )
        thread.daemon = True
        thread.start()
        
        # タイマーをすぐに開始
        self.update_timer()
    
    def run_extraction(self, url: str, depth: int, output_folder: str):
        """抽出実行（バックグラウンド）"""
        try:
            self.scraper = YahooCategoryScraper(
                log_callback=self.log,
                progress_callback=self.update_progress
            )
            
            categories = self.scraper.scrape(url, max_depth=depth)
            
            if categories and not self.scraper.stop_flag:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"{output_folder}/yahoo_categories_{timestamp}.xlsx"
                
                self.scraper.export_to_excel(output_path)
                
                self.root.after(0, lambda: messagebox.showinfo(
                    "完了",
                    f"カテゴリ抽出が完了しました！\n\n"
                    f"取得数: {len(categories)}件\n"
                    f"出力先: {output_path}"
                ))
            elif self.scraper.stop_flag:
                self.log("\n⏸️ 処理を中断しました")
            else:
                self.log("\n⚠️ カテゴリが取得できませんでした")
                
        except Exception as e:
            self.log(f"\n❌ エラーが発生しました: {e}")
            self.root.after(0, lambda: messagebox.showerror("エラー", str(e)))
        finally:
            self.root.after(0, self.extraction_finished)
    
    def extraction_finished(self):
        """抽出完了後の処理"""
        self.is_running = False
        self.start_button.config(state=tk.NORMAL, bg=self.PRIMARY)
        self.stop_button.config(state=tk.DISABLED, bg="#6C757D")
        
        # タイマー停止
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
    
    def stop_extraction(self):
        """抽出停止"""
        if self.scraper:
            self.scraper.stop()
    
    def on_closing(self):
        """ウィンドウクローズ時"""
        if self.scraper:
            self.scraper.stop()
            self.scraper.close_driver()
        
        self.root.quit()
        self.root.destroy()
    
    def run(self):
        """GUIを実行"""
        self.root.mainloop()


def main():
    """メイン関数"""
    app = YahooCategoryExtractorGUI()
    app.run()


if __name__ == "__main__":
    main()