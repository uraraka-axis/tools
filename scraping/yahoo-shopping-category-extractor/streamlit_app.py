#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Yahoo!ショッピング カテゴリ抽出ツール - Streamlit版
Webブラウザから使用可能なバージョン
"""

import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict
import random
import time
import re
import json
import io


@dataclass
class Category:
    """カテゴリ情報を保持するクラス"""
    name: str
    category_id: str
    url: str
    count: int
    level: int
    parent_path: List[str] = field(default_factory=list)


class YahooCategoryScraper:
    """Yahoo!ショッピングカテゴリスクレイパー (requests版)"""

    BASE_URL = "https://shopping.yahoo.co.jp"

    def __init__(self):
        self.session = requests.Session()
        # より完全なブラウザヘッダーを設定（bot検出回避）
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Cache-Control': 'max-age=0',
            'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
        })
        self.stop_flag = False
        self.categories: List[Category] = []
        self.root_category_name = ""
        self.root_category_id = ""
        self.total_requests = 0
        self.log_messages = []

        # 待機時間設定
        self.min_delay = 1.5
        self.max_delay = 4.0

    def log(self, message: str):
        """ログ出力"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_messages.append(log_message)

    def stop(self):
        """処理を停止"""
        self.stop_flag = True

    def random_delay(self):
        """ランダムな待機時間"""
        delay = random.uniform(self.min_delay, self.max_delay)
        if random.random() < 0.1:
            delay += random.uniform(1.0, 3.0)
        time.sleep(delay)

    def extract_category_id_from_url(self, url: str) -> str:
        """URLからカテゴリIDパスを抽出"""
        match = re.search(r'/category/([\d/]+)/list', url)
        if match:
            return match.group(1).rstrip('/')
        return ""

    def get_last_category_id(self, category_path: str) -> str:
        """カテゴリパスから最後のIDを取得"""
        if '/' in category_path:
            return category_path.split('/')[-1]
        return category_path

    def fetch_page(self, url: str):
        """ページを取得"""
        try:
            self.total_requests += 1

            if self.total_requests > 1:
                self.random_delay()

            response = self.session.get(url, timeout=30)
            response.raise_for_status()

            return BeautifulSoup(response.text, 'html.parser')

        except Exception as e:
            self.log(f"  ⚠️ ページ取得エラー: {e}")
            return None

    def get_root_category_name(self, soup: BeautifulSoup) -> str:
        """ルートカテゴリ名を取得"""
        h1 = soup.find('h1')
        if h1:
            name = h1.get_text(strip=True)
            name = re.sub(r'映像ソフト$|おすすめ.*$', '', name).strip()
            return name
        return "カテゴリ"

    def get_subcategories_from_page(self, soup: BeautifulSoup, current_category_id: str, is_root: bool = False) -> List[Dict]:
        """ページからサブカテゴリを抽出（__NEXT_DATA__のJSONから取得）"""
        subcategories = []

        next_data_script = soup.find('script', id='__NEXT_DATA__')
        if not next_data_script:
            self.log("    [DEBUG] __NEXT_DATA__ が見つかりません")
            # HTMLフォールバックを試す
            return self._extract_categories_from_html(soup, current_category_id)

        try:
            json_data = json.loads(next_data_script.string)

            # デバッグ: JSONの構造を確認
            props = json_data.get('props', {})
            page_props = props.get('pageProps', {})
            initial_state = page_props.get('initialState', {})

            self.log(f"    [DEBUG] props keys: {list(props.keys())[:5]}")
            self.log(f"    [DEBUG] pageProps keys: {list(page_props.keys())[:5]}")
            self.log(f"    [DEBUG] initialState keys: {list(initial_state.keys())[:5]}")

            categories_data = self._extract_categories_from_json(json_data)

            if not categories_data:
                self.log("    [DEBUG] JSONからカテゴリデータを取得できませんでした")
                # フォールバック: 別のパスを試す
                categories_data = self._extract_categories_fallback(json_data)
                if categories_data:
                    self.log(f"    [DEBUG] フォールバックで {len(categories_data)} 件取得")

            # JSONから1件以下の場合はHTMLフォールバックを試す
            if len(categories_data) <= 1:
                self.log("    [DEBUG] JSON結果が不十分、HTMLフォールバックを試行")
                html_categories = self._extract_categories_from_html(soup, current_category_id)
                if len(html_categories) > len(categories_data):
                    self.log(f"    [DEBUG] HTMLから {len(html_categories)} 件取得")
                    return html_categories

            if not categories_data:
                return []

            self.log(f"    [DEBUG] JSONから {len(categories_data)} 件のカテゴリを検出")

            for cat_data in categories_data:
                name = cat_data.get('text', '')
                url = cat_data.get('url', '')
                count = cat_data.get('count', 0)

                if not name or not url:
                    continue

                category_path = self.extract_category_id_from_url(url)
                if not category_path:
                    continue

                last_id = self.get_last_category_id(category_path)

                if not url.startswith('http'):
                    url = self.BASE_URL + url

                url = re.sub(r'\?.*$', '', url)
                if not url.endswith('/list'):
                    url = url.rstrip('/') + '/list'

                subcategories.append({
                    'name': name,
                    'url': url,
                    'category_id': category_path,
                    'last_id': last_id,
                    'count': count
                })

        except json.JSONDecodeError as e:
            self.log(f"    [DEBUG] JSON解析エラー: {e}")
            return []
        except Exception as e:
            self.log(f"    [DEBUG] カテゴリ抽出エラー: {e}")
            return []

        # 重複除去
        seen = set()
        unique = []
        for cat in subcategories:
            if cat['last_id'] not in seen and cat['name']:
                seen.add(cat['last_id'])
                unique.append(cat)

        return unique

    def _extract_categories_from_json(self, json_data: dict) -> List[Dict]:
        """JSONデータからカテゴリ情報を抽出"""
        categories = []

        try:
            page_props = json_data.get('props', {}).get('pageProps', {})

            # パス1: initialState.bff.advancedFilter経由（従来方式）
            initial_state = page_props.get('initialState', {})
            if initial_state:
                bff = initial_state.get('bff', {})
                advanced_filter = bff.get('advancedFilter', {})
                sections = advanced_filter.get('sections', {})
                category_section = sections.get('category', {})
                categories_data = category_section.get('categories', {})

                suggested = categories_data.get('suggestedCategories', [])
                if suggested:
                    categories.extend(suggested)

                toggle_items = categories_data.get('toggleAreaCategoryItems', [])
                if toggle_items:
                    categories.extend(toggle_items)

            # パス2: ptahV2InitialData経由（新方式）
            if not categories:
                ptah_data = page_props.get('ptahV2InitialData', {})
                if ptah_data:
                    # ptahV2InitialDataが文字列の場合はJSONとしてパース
                    if isinstance(ptah_data, str):
                        try:
                            ptah_data = json.loads(ptah_data)
                            self.log(f"    [DEBUG] ptahV2InitialData parsed from string")
                        except json.JSONDecodeError:
                            self.log(f"    [DEBUG] ptahV2InitialData is not valid JSON")
                            ptah_data = {}

                    if isinstance(ptah_data, dict):
                        self.log(f"    [DEBUG] ptahV2InitialData keys: {list(ptah_data.keys())[:10]}")
                        # ptahV2InitialData内を探索
                        categories = self._search_categories_in_ptah(ptah_data)

        except Exception as e:
            self.log(f"    [DEBUG] JSON構造解析エラー: {e}")

        return categories

    def _search_categories_in_ptah(self, ptah_data: dict) -> List[Dict]:
        """ptahV2InitialData内からカテゴリを探索"""
        categories = []

        def search(obj, depth=0):
            if depth > 15:
                return []
            found = []

            if isinstance(obj, dict):
                # advancedFilterセクションを探す
                if 'advancedFilter' in obj:
                    af = obj['advancedFilter']
                    if isinstance(af, dict) and 'sections' in af:
                        sections = af['sections']
                        if isinstance(sections, dict) and 'category' in sections:
                            cat_section = sections['category']
                            if isinstance(cat_section, dict) and 'categories' in cat_section:
                                cat_data = cat_section['categories']
                                if isinstance(cat_data, dict):
                                    suggested = cat_data.get('suggestedCategories', [])
                                    toggle = cat_data.get('toggleAreaCategoryItems', [])
                                    found.extend(suggested)
                                    found.extend(toggle)
                                    if found:
                                        return found

                # suggestedCategoriesを直接探す
                if 'suggestedCategories' in obj:
                    items = obj['suggestedCategories']
                    if isinstance(items, list) and items:
                        if isinstance(items[0], dict) and 'text' in items[0]:
                            found.extend(items)
                            toggle = obj.get('toggleAreaCategoryItems', [])
                            found.extend(toggle)
                            return found

                # 再帰探索
                for v in obj.values():
                    result = search(v, depth + 1)
                    if result:
                        return result

            elif isinstance(obj, list):
                for item in obj:
                    result = search(item, depth + 1)
                    if result:
                        return result

            return found

        try:
            categories = search(ptah_data)
        except Exception as e:
            self.log(f"    [DEBUG] ptah探索エラー: {e}")

        return categories

    def _extract_categories_fallback(self, json_data: dict) -> List[Dict]:
        """フォールバック: 別のパスでカテゴリを探す（suggestedCategories + toggleAreaCategoryItems）"""
        categories = []

        def find_categories(obj, depth=0):
            """再帰的にカテゴリ配列を探す（両方のリストを返す）"""
            if depth > 10:
                return None
            if isinstance(obj, dict):
                # suggestedCategories を探す（toggleAreaCategoryItemsも一緒に取得）
                if 'suggestedCategories' in obj:
                    result = []
                    suggested = obj['suggestedCategories']
                    if isinstance(suggested, list):
                        result.extend(suggested)
                    # toggleAreaCategoryItems も取得（「もっと見る」のカテゴリ）
                    toggle_items = obj.get('toggleAreaCategoryItems', [])
                    if isinstance(toggle_items, list):
                        result.extend(toggle_items)
                    if result:
                        return result
                # categories キーを探す
                if 'categories' in obj and isinstance(obj['categories'], dict):
                    cat_data = obj['categories']
                    result = []
                    suggested = cat_data.get('suggestedCategories', [])
                    if isinstance(suggested, list):
                        result.extend(suggested)
                    toggle_items = cat_data.get('toggleAreaCategoryItems', [])
                    if isinstance(toggle_items, list):
                        result.extend(toggle_items)
                    if result:
                        return result
                # categories が配列の場合
                if 'categories' in obj and isinstance(obj['categories'], list):
                    items = obj['categories']
                    if items and isinstance(items[0], dict) and 'text' in items[0]:
                        return items
                # 再帰的に探索
                for v in obj.values():
                    result = find_categories(v, depth + 1)
                    if result:
                        return result
            elif isinstance(obj, list) and obj:
                for item in obj:
                    result = find_categories(item, depth + 1)
                    if result:
                        return result
            return None

        try:
            found = find_categories(json_data)
            if found:
                categories = found
                self.log(f"    [DEBUG] フォールバックで suggestedCategories + toggleAreaCategoryItems を取得")
        except Exception as e:
            self.log(f"    [DEBUG] フォールバック探索エラー: {e}")

        return categories

    def _extract_categories_from_html(self, soup: BeautifulSoup, current_category_id: str) -> List[Dict]:
        """HTMLから直接カテゴリリンクを抽出（JSONが不十分な場合のフォールバック）"""
        subcategories = []
        seen_ids = set()

        # 現在のカテゴリIDの最後の部分を取得（フィルタリング用）
        current_last_id = self.get_last_category_id(current_category_id) if current_category_id else ""

        # 現在のカテゴリパスの階層数（親カテゴリ除外用）
        current_depth = len(current_category_id.split('/')) if current_category_id else 0

        # カテゴリリンクのパターン: /category/数字/list または /category/数字/数字/list など
        category_pattern = re.compile(r'/category/([\d/]+)/list')

        # すべてのリンクを探索
        for link in soup.find_all('a', href=True):
            href = link.get('href', '')
            match = category_pattern.search(href)
            if not match:
                continue

            category_path = match.group(1).rstrip('/')
            last_id = self.get_last_category_id(category_path)

            # カテゴリパスの階層数
            link_depth = len(category_path.split('/'))

            # 親カテゴリ（階層が浅い）は除外
            if link_depth <= current_depth:
                continue

            # 自分自身は除外
            if last_id == current_last_id:
                continue

            # 重複除去
            if last_id in seen_ids:
                continue
            seen_ids.add(last_id)

            # リンクテキストを取得
            name = link.get_text(strip=True)
            if not name:
                continue

            # 数字だけの場合はスキップ（件数表示など）
            if name.isdigit() or re.match(r'^[\d,]+件?$', name):
                continue

            # 「もっと見る」「すべて見る」などはスキップ
            if name in ['もっと見る', 'すべて見る', '詳細を見る', '閉じる']:
                continue

            url = href
            if not url.startswith('http'):
                url = self.BASE_URL + url

            url = re.sub(r'\?.*$', '', url)
            if not url.endswith('/list'):
                url = url.rstrip('/') + '/list'

            # 件数を取得（リンクの近くにある数字）
            count = 0
            try:
                count_text = link.find_next(string=re.compile(r'[\d,]+件?'))
                if count_text:
                    count_match = re.search(r'([\d,]+)', str(count_text))
                    if count_match:
                        count_str = count_match.group(1).replace(',', '')
                        if count_str:
                            count = int(count_str)
            except (ValueError, AttributeError):
                count = 0

            subcategories.append({
                'name': name,
                'url': url,
                'category_id': category_path,
                'last_id': last_id,
                'count': count
            })

        self.log(f"    [DEBUG] HTMLから {len(subcategories)} 件のカテゴリリンクを検出")
        return subcategories

    def scrape_categories_recursive(
        self,
        url: str,
        level: int = 0,
        parent_path: List[str] = None,
        max_depth: int = 5,
        progress_callback=None
    ):
        """カテゴリを再帰的に取得"""
        if parent_path is None:
            parent_path = []

        if level > max_depth or self.stop_flag:
            return

        indent = "  " * level
        self.log(f"{indent}📂 取得中: {url}")

        soup = self.fetch_page(url)
        if not soup:
            return

        current_id = self.extract_category_id_from_url(url)

        is_root = (level == 0)
        if is_root:
            self.root_category_name = self.get_root_category_name(soup)
            self.root_category_id = current_id
            self.log(f"📌 ルートカテゴリ: {self.root_category_name} (ID: {current_id})")

        subcategories = self.get_subcategories_from_page(soup, current_id, is_root=is_root)

        self.log(f"{indent}  → {len(subcategories)}件のサブカテゴリを発見")

        for subcat in subcategories:
            if self.stop_flag:
                break

            cat = Category(
                name=subcat['name'],
                category_id=subcat['category_id'],
                url=subcat['url'],
                count=subcat['count'],
                level=level + 1,
                parent_path=parent_path.copy()
            )
            self.categories.append(cat)

            self.log(f"{indent}  ✓ {subcat['name']} ({subcat['count']:,}件) [ID: {subcat['last_id']}]")

            # 進捗コールバック
            if progress_callback:
                progress_callback(len(self.categories), parent_path + [subcat['name']])

            # 再帰
            if level + 1 < max_depth:
                new_parent_path = parent_path + [subcat['name']]
                self.scrape_categories_recursive(
                    subcat['url'],
                    level + 1,
                    new_parent_path,
                    max_depth,
                    progress_callback
                )

    def scrape(self, start_url: str, max_depth: int = 5, progress_callback=None) -> List[Category]:
        """スクレイピング開始"""
        self.stop_flag = False
        self.categories = []
        self.total_requests = 0
        self.log_messages = []

        self.log("=" * 50)
        self.log("🛒 Yahoo!ショッピング カテゴリ抽出開始")
        self.log("=" * 50)
        self.log(f"🔗 URL: {start_url}")
        self.log(f"📊 最大取得階層: {max_depth}")
        self.log("")

        self.scrape_categories_recursive(start_url, max_depth=max_depth, progress_callback=progress_callback)

        if not self.stop_flag:
            self.log("")
            self.log(f"✅ 合計 {len(self.categories)} カテゴリを取得しました")
            self.log(f"📡 総リクエスト数: {self.total_requests}")

        return self.categories

    def export_to_excel(self) -> bytes:
        """Excelファイルをバイトストリームで出力"""
        if not self.categories:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "ジャンル一覧"

        base_font = Font(name="Meiryo UI", size=10)
        header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="ff0033")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(name="Meiryo UI", bold=True, size=14, color="ff0033")

        thin_border = Border(
            left=Side(style='thin', color='595959'),
            right=Side(style='thin', color='595959'),
            top=Side(style='thin', color='595959'),
            bottom=Side(style='thin', color='595959')
        )

        max_level = max((cat.level for cat in self.categories), default=1)

        level_counts = defaultdict(int)
        for cat in self.categories:
            level_counts[cat.level] += 1

        summary_level_col = 2 + max_level + 4
        summary_count_col = 2 + max_level + 5

        title_col_end = get_column_letter(summary_count_col)
        ws.merge_cells(f'B1:{title_col_end}1')
        ws['B1'] = f"【Yahoo!ショッピング】{self.root_category_name}のジャンル一覧"
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 8

        headers = ["#", "ジャンル1"]
        for i in range(max_level):
            headers.append(f"ジャンル{i + 2}")
        headers.append("カテゴリID")
        headers.append("ページURL")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[3].height = 24

        summary_level_header = ws.cell(row=3, column=summary_level_col, value="レベル")
        summary_level_header.font = header_font
        summary_level_header.fill = header_fill
        summary_level_header.alignment = header_alignment
        summary_level_header.border = thin_border

        summary_count_header = ws.cell(row=3, column=summary_count_col, value="カテゴリ数")
        summary_count_header.font = header_font
        summary_count_header.fill = header_fill
        summary_count_header.alignment = header_alignment
        summary_count_header.border = thin_border

        prev_values = [""] * (max_level + 1)

        for idx, cat in enumerate(self.categories, 1):
            row = idx + 3

            current_values = [self.root_category_name] + [""] * max_level

            for i, parent_name in enumerate(cat.parent_path):
                if i < max_level:
                    current_values[i + 1] = parent_name

            if cat.level <= max_level:
                current_values[cat.level] = cat.name

            cell = ws.cell(row=row, column=1, value=idx)
            cell.border = thin_border
            cell.font = base_font

            for col, value in enumerate(current_values, 2):
                cell = ws.cell(row=row, column=col)

                show_value = value
                if idx > 1 and col - 2 < len(prev_values):
                    if value == prev_values[col - 2]:
                        has_change = any(
                            current_values[j] != prev_values[j]
                            for j in range(col - 1, len(current_values))
                            if j < len(prev_values)
                        )
                        if not has_change:
                            show_value = ""

                cell.value = show_value
                cell.border = thin_border
                cell.font = base_font

            id_col = 2 + max_level + 1
            last_id = self.get_last_category_id(cat.category_id)
            id_cell = ws.cell(row=row, column=id_col, value=last_id)
            id_cell.border = thin_border
            id_cell.font = base_font

            url_col = 2 + max_level + 2
            url_cell = ws.cell(row=row, column=url_col, value=cat.url)
            url_cell.hyperlink = cat.url
            url_cell.style = "Hyperlink"
            url_cell.border = thin_border
            url_cell.font = Font(name="Meiryo UI", size=10, color="0563C1", underline="single")

            prev_values = current_values.copy()

        summary_row = 4

        level_cell = ws.cell(row=summary_row, column=summary_level_col, value="ジャンル1")
        level_cell.border = thin_border
        level_cell.font = base_font

        count_cell = ws.cell(row=summary_row, column=summary_count_col, value=1)
        count_cell.border = thin_border
        count_cell.font = base_font
        count_cell.alignment = Alignment(horizontal="right")

        summary_row += 1

        for level in sorted(level_counts.keys()):
            level_cell = ws.cell(row=summary_row, column=summary_level_col, value=f"ジャンル{level + 1}")
            level_cell.border = thin_border
            level_cell.font = base_font

            count_cell = ws.cell(row=summary_row, column=summary_count_col, value=level_counts[level])
            count_cell.border = thin_border
            count_cell.font = base_font
            count_cell.alignment = Alignment(horizontal="right")

            summary_row += 1

        total_cell = ws.cell(row=summary_row, column=summary_level_col, value="合計")
        total_cell.border = thin_border
        total_cell.font = Font(name="Meiryo UI", size=10, bold=True)

        total_count_cell = ws.cell(row=summary_row, column=summary_count_col, value=len(self.categories) + 1)
        total_count_cell.border = thin_border
        total_count_cell.font = Font(name="Meiryo UI", size=10, bold=True)
        total_count_cell.alignment = Alignment(horizontal="right")

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        for i in range(max_level):
            col_letter = get_column_letter(3 + i)
            ws.column_dimensions[col_letter].width = 22
        ws.column_dimensions[get_column_letter(3 + max_level)].width = 18
        ws.column_dimensions[get_column_letter(4 + max_level)].width = 50
        ws.column_dimensions[get_column_letter(5 + max_level)].width = 3
        ws.column_dimensions[get_column_letter(summary_level_col)].width = 12
        ws.column_dimensions[get_column_letter(summary_count_col)].width = 12

        ws.freeze_panes = 'A4'

        # バイトストリームとして出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def check_password():
    """パスワード認証"""

    # secretsにパスワードが設定されているか確認
    if "password" not in st.secrets:
        # パスワード未設定の場合はそのままアクセス可能
        return True

    correct_password = st.secrets["password"]

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown("""
        <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 40px;
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        </style>
    """, unsafe_allow_html=True)

    st.title("🔐 ログイン")
    st.write("このツールを使用するにはパスワードが必要です。")

    password = st.text_input("パスワード", type="password", key="password_input")

    if st.button("ログイン", type="primary"):
        if password == correct_password:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが違います")

    return False


def main():
    """メイン関数"""

    st.set_page_config(
        page_title="Yahoo!ショッピング カテゴリ抽出ツール",
        page_icon="🛒",
        layout="wide"
    )

    # パスワード認証
    if not check_password():
        return

    # カスタムCSS
    st.markdown("""
        <style>
        .main-header {
            background: linear-gradient(135deg, #ff0033 0%, #cc0029 100%);
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        .main-header h1 {
            color: white !important;
            margin: 0;
        }
        .main-header p {
            color: #FFE0B2;
            margin: 5px 0 0 0;
        }
        .stats-card {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }
        .stats-value {
            font-size: 2em;
            font-weight: bold;
            color: #ff0033;
        }
        </style>
    """, unsafe_allow_html=True)

    # ヘッダー
    st.markdown("""
        <div class="main-header">
            <h1>🛒 Yahoo!ショッピング カテゴリ抽出ツール</h1>
            <p>Yahoo! Shopping Category Extractor</p>
        </div>
    """, unsafe_allow_html=True)

    # セッションステート初期化
    if "scraper" not in st.session_state:
        st.session_state.scraper = None
    if "is_running" not in st.session_state:
        st.session_state.is_running = False
    if "excel_data" not in st.session_state:
        st.session_state.excel_data = None
    if "log_messages" not in st.session_state:
        st.session_state.log_messages = []
    if "total_categories" not in st.session_state:
        st.session_state.total_categories = 0

    # 入力フォーム
    with st.container():
        st.subheader("📝 設定")

        col1, col2 = st.columns([3, 1])

        with col1:
            url = st.text_input(
                "カテゴリURL",
                value="https://shopping.yahoo.co.jp/category/2517/list",
                help="例: https://shopping.yahoo.co.jp/category/2517/list （DVD、映像ソフト）"
            )

        with col2:
            depth = st.number_input(
                "取得階層数",
                min_value=1,
                max_value=10,
                value=3,
                help="1〜10階層まで指定可能"
            )

    # ボタンエリア
    start_disabled = st.session_state.is_running
    start_clicked = st.button("🚀 抽出開始", disabled=start_disabled, type="primary")

    # 抽出処理
    if start_clicked and url:
        st.session_state.is_running = True
        st.session_state.excel_data = None
        st.session_state.log_messages = []
        st.session_state.total_categories = 0

        # スクレイパー実行
        scraper = YahooCategoryScraper()
        st.session_state.scraper = scraper

        progress_bar = st.progress(0)
        status_text = st.empty()
        log_container = st.empty()

        def update_progress(count, path):
            st.session_state.total_categories = count
            status_text.text(f"取得中: {count}件 | {' > '.join(path)}")

        try:
            with st.spinner("カテゴリを取得中..."):
                categories = scraper.scrape(url, max_depth=depth, progress_callback=update_progress)

            st.session_state.log_messages = scraper.log_messages

            if categories:
                st.session_state.excel_data = scraper.export_to_excel()
                st.session_state.total_categories = len(categories)
                st.success(f"✅ {len(categories)}件のカテゴリを取得しました！")
            else:
                st.warning("カテゴリが取得できませんでした")

        except Exception as e:
            st.error(f"エラーが発生しました: {e}")
        finally:
            st.session_state.is_running = False
            st.rerun()

    elif start_clicked and not url:
        st.error("URLを入力してください")

    # 統計情報
    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f"""
            <div class="stats-card">
                <p>取得カテゴリ数</p>
                <div class="stats-value">{st.session_state.total_categories}件</div>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        # ダウンロードボタン
        if st.session_state.excel_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📥 Excelファイルをダウンロード",
                data=st.session_state.excel_data,
                file_name=f"yahoo_categories_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

    # ログ表示
    if st.session_state.log_messages:
        with st.expander("📋 ログ", expanded=False):
            log_text = "\n".join(st.session_state.log_messages)
            st.code(log_text, language=None)


if __name__ == "__main__":
    main()
