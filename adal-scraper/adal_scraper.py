# -*- coding: utf-8 -*-
"""
ADAL ONLINE SHOP (https://adal-online.shop/) 商品スクレイパー

全カテゴリの全商品をスクレイピングし、Excel に出力します。
品番が複数ある場合は 1品番=1行 に展開します。

GUI版:
    python adal_scraper.py            # 引数なし → GUI起動
    （ビルドした adal_scraper.exe をダブルクリックでもGUI起動）

CLI版:
    python adal_scraper.py --categories others        # 「その他」のみ
    python adal_scraper.py --limit 5                  # 各カテゴリ先頭5商品
    python adal_scraper.py --out 結果.xlsx --delay 1.0

依存: requests, beautifulsoup4, openpyxl
"""
import argparse
import re
import sys
import time
import unicodedata
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = "https://adal-online.shop"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
}

# カテゴリ: 表示名 -> URLスラッグ（巡回順）
CATEGORIES = [
    ("チェア", "chair"),
    ("ソファ", "sofa"),
    ("ベンチ", "bench"),
    ("キッズ", "kids"),
    ("アウトドア", "outdoors"),
    ("テーブル", "table"),
    ("その他", "others"),
    ("クリアランスセール", "outlet-option"),
]

# 出力ヘッダ（A〜P列）
HEADERS_ROW = [
    "No.",          # A
    "カタログNo.",   # B
    "商品名",        # C
    "品番",          # D
    "カラー",        # E
    "材質",          # F
    "重量",          # G
    "サイズ",        # H
    "保証期間",      # I
    "配送",          # J
    "お届け目安",    # K
    "カテゴリ",      # L
    "URL",          # M
    "カタログ価格",  # N
    "法人会員価格",  # O
    "商品説明",      # P
]


# ============================================================
#  共通ヘルパ
# ============================================================
def build_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def clean_text(s):
    """連続空白を1つに、前後トリム。"""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", s).strip()


# 型番らしさ: 英数字を含み、かつ数字を含み、英字/数字/記号(-_/.()) のみ
CODE_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-_/.()]*$")


def parse_color_code(line):
    """「ウォームグレー P3002-10JEC」-> (color, code)。
    末尾の型番トークンを code、残りを color とする。
    型番が見つからなければ (line, '')。"""
    tokens = line.split(" ")
    if len(tokens) >= 2:
        last = tokens[-1]
        if CODE_RE.match(last) and any(ch.isdigit() for ch in last):
            color = " ".join(tokens[:-1]).strip()
            return color, last
    if len(tokens) == 1 and CODE_RE.match(tokens[0]) and any(ch.isdigit() for ch in tokens[0]):
        return "", tokens[0]
    return line, ""


def expand_hinban(lines):
    """品番の行リストを (カラー, 品番) のペアに展開する。
    2レイアウトに対応:
      A) 「カラー名 型番」が同一行 （例: ウォームグレー P3002-10JEC）
      B) 「ラベル行」→「型番行」が交互 （例: 左テーブル / X4017-99LX）
    """
    pairs = []
    pending = None
    for line in lines:
        color, code = parse_color_code(line)
        if code and color:
            if pending is not None:
                pairs.append(("", pending))
                pending = None
            pairs.append((color, code))
        elif code and not color:
            if pending is not None:
                pairs.append((pending, code))
                pending = None
            else:
                pairs.append(("", code))
        else:
            if pending is not None:
                pairs.append(("", pending))
            pending = line
    if pending is not None:
        pairs.append(("", pending))
    return pairs


def get_spec_dd(soup, label):
    """スペック表(dl)から指定ラベルの dd テキストを返す。<a>リンクは除去。"""
    for dt in soup.select("dt.item-detail-info__spec--title"):
        if dt.get_text(strip=True) == label:
            dd = dt.find_next_sibling("dd")
            if dd is None:
                return "", None
            dd_copy = BeautifulSoup(str(dd), "html.parser")
            for a in dd_copy.find_all("a"):
                a.decompose()
            return clean_text(dd_copy.get_text(" ")), dd
    return "", None


def split_hinban_lines(dd):
    if dd is None:
        return []
    lines = [clean_text(x) for x in dd.get_text("\n").split("\n")]
    return [ln for ln in lines if ln]


def disp_width(s):
    """全角を2、半角を1として文字幅を概算。"""
    w = 0
    for ch in str(s):
        w += 2 if unicodedata.east_asian_width(ch) in ("F", "W", "A") else 1
    return w


def write_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ADAL商品"

    header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    body_font = Font(name="Meiryo UI")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.append(HEADERS_ROW)
    for col, _ in enumerate(HEADERS_ROW, start=1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r["catalog_no"], r["name"], r["hinban"], r["color"], r["material"],
            r["weight"], r["size"], r["warranty"], r["shipping"], r["delivery"],
            r["category"], r["url"], r["catalog_price"], r["member_price"],
            r["description"],
        ])

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=len(HEADERS_ROW)):
        for c in row:
            c.font = body_font
            c.border = border
            c.alignment = center if c.column == 1 else left

    for col in range(1, len(HEADERS_ROW) + 1):
        letter = get_column_letter(col)
        max_w = disp_width(HEADERS_ROW[col - 1])
        for row in range(2, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_w = max(max_w, disp_width(v))
        ws.column_dimensions[letter].width = min(max_w + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ============================================================
#  スクレイパー本体（コールバック・停止対応）
# ============================================================
class AdalScraper:
    def __init__(self, delay=0.7, limit=None, log_callback=None, progress_callback=None):
        self.delay = delay
        self.limit = limit  # 各カテゴリの最大商品数（None=全件）
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self._stop = False
        self.session = build_session()

    # --- 通知系 ---
    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            try:
                print(message)
            except UnicodeEncodeError:
                enc = getattr(sys.stdout, "encoding", "utf-8") or "utf-8"
                print(message.encode(enc, "replace").decode(enc))

    def progress(self, current, total):
        if self.progress_callback:
            self.progress_callback(current, total)

    def stop(self):
        self._stop = True

    # --- 取得 ---
    def get_soup(self, url, retries=3, timeout=30):
        for attempt in range(1, retries + 1):
            if self._stop:
                return None
            try:
                r = self.session.get(url, timeout=timeout)
                if r.status_code == 200:
                    r.encoding = "UTF-8"
                    return BeautifulSoup(r.text, "html.parser")
                if r.status_code == 404:
                    return None
                self.log(f"  [warn] {url} status={r.status_code} (try {attempt}/{retries})")
            except requests.RequestException as e:
                self.log(f"  [warn] {url} error={e} (try {attempt}/{retries})")
            time.sleep(1.5 * attempt)
        return None

    def collect_detail_urls(self, slug):
        """カテゴリの全ページを巡回し、商品詳細URL（重複なし・出現順）を返す。"""
        detail_urls = []
        seen = set()
        pageno = 1
        max_pages = 80
        while pageno <= max_pages and not self._stop:
            url = f"{BASE}/{slug}?pageno={pageno}"
            soup = self.get_soup(url)
            if soup is None:
                break
            anchors = soup.select('ul.itembox a[href*="/products/detail/"]')
            new_on_page = 0
            for a in anchors:
                m = re.search(r"/products/detail/(\d+)", a["href"])
                if not m:
                    continue
                pid = m.group(1)
                if pid in seen:
                    continue
                seen.add(pid)
                detail_urls.append(f"{BASE}/products/detail/{pid}")
                new_on_page += 1
                if self.limit and len(detail_urls) >= self.limit:
                    return detail_urls
            if new_on_page == 0:
                break
            pageno += 1
            time.sleep(self.delay)
        return detail_urls

    def parse_detail(self, soup, url, category_label):
        """詳細ページから情報を抽出し、品番ごとに展開した行(dict)リストを返す。"""
        h1 = soup.select_one("h1")
        name = clean_text(h1.get_text()) if h1 else ""

        cat_el = soup.select_one("p.item-detail__heading--catalog")
        catalog_no = ""
        if cat_el:
            catalog_no = re.sub(r"^カタログ[:：]\s*", "", clean_text(cat_el.get_text()))

        cat_price_el = soup.select_one("p.item-detail-price__catalog")
        catalog_price = ""
        if cat_price_el:
            catalog_price = clean_text(cat_price_el.get_text()).replace("カタログ価格", "").strip()
        member_price_el = soup.select_one("p.item-detail-price__member")
        member_price = ""
        if member_price_el:
            member_price = clean_text(member_price_el.get_text()).replace("法人会員価格", "").strip()

        delivery_el = soup.select_one("div.item-detail-cart__form--input--delivery")
        delivery_est = clean_text(delivery_el.get_text()) if delivery_el else ""

        desc_el = soup.select_one("p.item-detail-info__description--text")
        description = clean_text(desc_el.get_text()) if desc_el else ""

        material, _ = get_spec_dd(soup, "材質")
        weight, _ = get_spec_dd(soup, "重量")
        size, _ = get_spec_dd(soup, "サイズ")
        warranty, _ = get_spec_dd(soup, "保証期間")
        shipping, _ = get_spec_dd(soup, "配送")
        _, hinban_dd = get_spec_dd(soup, "品番")
        hinban_lines = split_hinban_lines(hinban_dd)

        common = {
            "catalog_no": catalog_no, "name": name, "material": material,
            "weight": weight, "size": size, "warranty": warranty,
            "shipping": shipping, "delivery": delivery_est,
            "category": category_label, "url": url,
            "catalog_price": catalog_price, "member_price": member_price,
            "description": description,
        }

        rows = []
        pairs = expand_hinban(hinban_lines)
        if pairs:
            for color, code in pairs:
                row = dict(common)
                row["hinban"] = code
                row["color"] = color
                rows.append(row)
        else:
            row = dict(common)
            row["hinban"] = ""
            row["color"] = ""
            rows.append(row)
        return rows

    def scrape(self, targets):
        """targets: [(label, slug), ...] を巡回し全行を返す。
        戻り値: dict(rows=[...], products=N, stopped=bool)"""
        # フェーズ1: 全カテゴリのURLを収集
        self.log("▼ 商品URLを収集しています...")
        plan = []  # (label, url)
        for label, slug in targets:
            if self._stop:
                break
            urls = self.collect_detail_urls(slug)
            self.log(f"  [{label}] {len(urls)} 商品")
            for u in urls:
                plan.append((label, u))

        total = len(plan)
        self.log(f"▼ 合計 {total} 商品の詳細を解析します")
        self.progress(0, total)

        # フェーズ2: 詳細ページ解析
        all_rows = []
        done = 0
        for label, url in plan:
            if self._stop:
                break
            soup = self.get_soup(url)
            done += 1
            if soup is None:
                self.log(f"  [skip] 取得失敗: {url}")
                self.progress(done, total)
                continue
            try:
                rows = self.parse_detail(soup, url, label)
                all_rows.extend(rows)
                name = rows[0]["name"] if rows else ""
                self.log(f"  [{done}/{total}] {name}（{len(rows)}品番）")
            except Exception as e:
                self.log(f"  [error] 解析失敗 {url}: {e}")
            self.progress(done, total)
            time.sleep(self.delay)

        return {"rows": all_rows, "products": total, "stopped": self._stop}


# ============================================================
#  CLI
# ============================================================
def run_cli():
    ap = argparse.ArgumentParser(description="ADAL ONLINE SHOP 商品スクレイパー")
    ap.add_argument("--categories", default="",
                    help="巡回するカテゴリのスラッグをカンマ区切り（省略時は全カテゴリ）。例: chair,sofa")
    ap.add_argument("--limit", type=int, default=None, help="各カテゴリの最大商品数（テスト用）")
    ap.add_argument("--delay", type=float, default=0.7, help="リクエスト間の待機秒数（既定0.7）")
    ap.add_argument("--out", default=None, help="出力Excelパス")
    ap.add_argument("--gui", action="store_true", help="GUIを起動")
    args = ap.parse_args()

    if args.gui:
        launch_gui()
        return

    if args.categories.strip():
        wanted = [c.strip() for c in args.categories.split(",") if c.strip()]
        targets = [(lbl, slug) for lbl, slug in CATEGORIES if slug in wanted]
        if not targets:
            print(f"指定カテゴリが見つかりません: {wanted}")
            print("有効なスラッグ:", ", ".join(s for _, s in CATEGORIES))
            sys.exit(1)
    else:
        targets = CATEGORIES

    out_path = args.out or f"adal_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    scraper = AdalScraper(delay=args.delay, limit=args.limit)
    result = scraper.scrape(targets)
    rows = result["rows"]
    print(f"\n出力行数: {len(rows)} 行")
    if rows:
        write_excel(rows, out_path)
        print(f"Excel を書き出しました: {out_path}")
    else:
        print("データが取得できませんでした。")


# ============================================================
#  GUI（参考: 東谷 ESSENCE MALL ダウンローダーと同テイスト）
# ============================================================
def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk
    import threading

    COLORS = {
        "primary": "#475569", "primary_hover": "#334155", "success": "#059669",
        "warning": "#d97706", "danger": "#dc2626", "bg_dark": "#1e293b",
        "bg_light": "#f8fafc", "text_dark": "#0f172a", "text_light": "#64748b",
        "border": "#e2e8f0", "accent": "#3b82f6",
    }

    class ScraperGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("ADAL ONLINE SHOP 商品一括スクレイパー v1.0")
            self.root.geometry("960x820")
            self.root.configure(bg=COLORS["bg_light"])
            self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

            self.out_path = tk.StringVar()
            self.delay = tk.StringVar(value="0.7")
            self.limit = tk.StringVar(value="0")  # 0=全件
            self.status_text = tk.StringVar(value="準備完了")
            self.cat_vars = {slug: tk.BooleanVar(value=True) for _, slug in CATEGORIES}

            self.scraper = None
            self._setup_styles()
            self._setup_ui()

        # --- styles ---
        def _setup_styles(self):
            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Title.TLabel", background=COLORS["bg_light"],
                            foreground=COLORS["text_dark"], font=("Segoe UI", 18, "bold"))
            style.configure("Subtitle.TLabel", background=COLORS["bg_light"],
                            foreground=COLORS["text_light"], font=("Segoe UI", 10))
            style.configure("Field.TLabel", background="white",
                            foreground=COLORS["text_dark"], font=("Segoe UI", 10, "bold"))
            style.configure("Modern.Horizontal.TProgressbar", troughcolor=COLORS["border"],
                            background=COLORS["primary"], borderwidth=0, thickness=20)

        # --- UI ---
        def _setup_ui(self):
            main = tk.Frame(self.root, bg=COLORS["bg_light"])
            main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            self._create_header(main)
            self._create_category_area(main)
            self._create_output_area(main)
            self._create_status_area(main)
            self._create_action_buttons(main)
            self._create_log_area(main)

        def _create_header(self, parent):
            f = tk.Frame(parent, bg=COLORS["bg_light"])
            f.pack(fill=tk.X, pady=(0, 18))
            ttk.Label(f, text="ADAL ONLINE SHOP 商品一括スクレイパー", style="Title.TLabel").pack()
            ttk.Label(f, text="全カテゴリの全商品を取得 → 品番ごとに行展開してExcel出力（A〜P列）",
                      style="Subtitle.TLabel").pack(pady=(5, 0))

        def _create_category_area(self, parent):
            card = tk.Frame(parent, bg="white", relief=tk.FLAT, bd=1)
            card.pack(fill=tk.X, pady=(0, 12))
            inner = tk.Frame(card, bg="white")
            inner.pack(fill=tk.BOTH, padx=20, pady=15)

            head = tk.Frame(inner, bg="white")
            head.pack(fill=tk.X, pady=(0, 8))
            tk.Label(head, text="取得するカテゴリ", bg="white", fg=COLORS["text_dark"],
                     font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
            tk.Button(head, text="全解除", command=lambda: self._toggle_all(False),
                      bg=COLORS["bg_light"], font=("Segoe UI", 9), relief=tk.FLAT,
                      cursor="hand2", padx=10, pady=2).pack(side=tk.RIGHT)
            tk.Button(head, text="全選択", command=lambda: self._toggle_all(True),
                      bg=COLORS["bg_light"], font=("Segoe UI", 9), relief=tk.FLAT,
                      cursor="hand2", padx=10, pady=2).pack(side=tk.RIGHT, padx=(0, 6))

            grid = tk.Frame(inner, bg="white")
            grid.pack(fill=tk.X)
            for i, (label, slug) in enumerate(CATEGORIES):
                r, c = divmod(i, 4)
                tk.Checkbutton(grid, text=label, variable=self.cat_vars[slug],
                               bg="white", activebackground="white",
                               font=("Segoe UI", 10), anchor=tk.W, width=16
                               ).grid(row=r, column=c, sticky=tk.W, padx=(0, 10), pady=3)

        def _create_output_area(self, parent):
            card = tk.Frame(parent, bg="white", relief=tk.FLAT, bd=1)
            card.pack(fill=tk.X, pady=(0, 12))
            inner = tk.Frame(card, bg="white")
            inner.pack(fill=tk.BOTH, padx=20, pady=18)

            # 出力ファイル
            of = tk.Frame(inner, bg="white")
            of.pack(fill=tk.X, pady=(0, 12))
            ttk.Label(of, text="出力Excelファイル（未指定なら自動命名）", style="Field.TLabel").pack(
                anchor=tk.W, pady=(0, 5))
            of_in = tk.Frame(of, bg="white")
            of_in.pack(fill=tk.X)
            tk.Entry(of_in, textvariable=self.out_path, font=("Segoe UI", 10),
                     relief=tk.SOLID, bd=1, bg="#f8fafc").pack(
                side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=8)
            tk.Button(of_in, text="参照", command=self.browse_output, bg=COLORS["primary"],
                      fg="white", font=("Segoe UI", 9), relief=tk.FLAT, cursor="hand2",
                      padx=15, pady=6).pack(side=tk.LEFT, padx=(10, 0))

            # オプション行
            opt = tk.Frame(inner, bg="white")
            opt.pack(fill=tk.X)
            tk.Label(opt, text="各カテゴリ最大件数", bg="white", fg=COLORS["text_dark"],
                     font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 8))
            tk.Spinbox(opt, from_=0, to=99999, textvariable=self.limit, width=6,
                       font=("Segoe UI", 10), relief=tk.SOLID, bd=1, bg="#f8fafc",
                       justify=tk.RIGHT).pack(side=tk.LEFT, ipady=3)
            tk.Label(opt, text="件（0=全件）", bg="white", fg=COLORS["text_light"],
                     font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(6, 0))

            tk.Label(opt, text="リクエスト間隔", bg="white", fg=COLORS["text_dark"],
                     font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(30, 8))
            tk.Spinbox(opt, from_=0.0, to=10.0, increment=0.1, textvariable=self.delay,
                       width=6, font=("Segoe UI", 10), relief=tk.SOLID, bd=1, bg="#f8fafc",
                       justify=tk.RIGHT).pack(side=tk.LEFT, ipady=3)
            tk.Label(opt, text="秒", bg="white", fg=COLORS["text_light"],
                     font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(6, 0))

        def _create_status_area(self, parent):
            f = tk.Frame(parent, bg=COLORS["bg_light"])
            f.pack(fill=tk.X, pady=(0, 12))
            tk.Label(f, textvariable=self.status_text, bg=COLORS["bg_light"],
                     fg=COLORS["text_light"], font=("Segoe UI", 9)).pack(anchor=tk.W, pady=(0, 5))
            self.progress = ttk.Progressbar(f, style="Modern.Horizontal.TProgressbar",
                                            mode="determinate", length=400)
            self.progress.pack(fill=tk.X)

        def _create_action_buttons(self, parent):
            f = tk.Frame(parent, bg=COLORS["bg_light"])
            f.pack(fill=tk.X, pady=(0, 12))
            cont = tk.Frame(f, bg=COLORS["bg_light"])
            cont.pack()
            self.run_button = tk.Button(cont, text="取得開始", command=self.run_scrape,
                                        bg=COLORS["success"], fg="white", font=("Segoe UI", 10),
                                        relief=tk.FLAT, cursor="hand2", padx=25, pady=10, borderwidth=0)
            self.run_button.pack(side=tk.LEFT, padx=(0, 8))
            self.stop_button = tk.Button(cont, text="停止", command=self.stop_scrape,
                                         bg=COLORS["danger"], fg="white", font=("Segoe UI", 10),
                                         relief=tk.FLAT, cursor="hand2", padx=25, pady=10,
                                         state=tk.DISABLED, borderwidth=0)
            self.stop_button.pack(side=tk.LEFT)

        def _create_log_area(self, parent):
            log_frame = tk.Frame(parent, bg="white", relief=tk.FLAT, bd=1)
            log_frame.pack(fill=tk.BOTH, expand=True)
            header = tk.Frame(log_frame, bg=COLORS["bg_light"])
            header.pack(fill=tk.X)
            tk.Label(header, text="処理ログ", bg=COLORS["bg_light"], fg=COLORS["text_dark"],
                     font=("Segoe UI", 10, "bold"), pady=10, padx=15).pack(anchor=tk.W)
            tk.Frame(log_frame, bg=COLORS["border"], height=1).pack(fill=tk.X)
            content = tk.Frame(log_frame, bg="white")
            content.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
            self.log_text = scrolledtext.ScrolledText(content, wrap=tk.WORD, font=("Consolas", 9),
                                                      bg="#0f172a", fg="#e2e8f0",
                                                      insertbackground="white", relief=tk.FLAT,
                                                      padx=10, pady=10)
            self.log_text.pack(fill=tk.BOTH, expand=True)

        # --- handlers ---
        def _toggle_all(self, val):
            for v in self.cat_vars.values():
                v.set(val)

        def browse_output(self):
            path = filedialog.asksaveasfilename(
                title="出力Excelファイル", defaultextension=".xlsx",
                filetypes=[("Excelファイル", "*.xlsx")],
                initialfile=f"adal_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if path:
                self.out_path.set(path)

        def log_message(self, message):
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.root.update_idletasks()

        def update_progress(self, current, total):
            value = (current / total) * 100 if total > 0 else 0
            self.progress["value"] = value
            self.status_text.set(f"処理中... {current}/{total} ({value:.1f}%)")
            self.root.update_idletasks()

        def stop_scrape(self):
            if self.scraper:
                self.scraper.stop()
                self.stop_button.config(state=tk.DISABLED, bg=COLORS["text_light"])
                self.log_message("\n⏹ 停止します。処理を中断しています...")

        def run_scrape(self):
            targets = [(lbl, slug) for lbl, slug in CATEGORIES if self.cat_vars[slug].get()]
            if not targets:
                messagebox.showwarning("警告", "カテゴリを1つ以上選択してください")
                return
            self.run_button.config(state=tk.DISABLED, bg=COLORS["text_light"])
            self.stop_button.config(state=tk.NORMAL, bg=COLORS["danger"])
            self.log_text.delete(1.0, tk.END)
            self.progress["value"] = 0
            self.status_text.set("処理を開始します...")
            threading.Thread(target=self._run_thread, args=(targets,), daemon=True).start()

        def _run_thread(self, targets):
            try:
                try:
                    delay = float(self.delay.get().strip() or "0.7")
                except ValueError:
                    delay = 0.7
                try:
                    lim = int(self.limit.get().strip() or "0")
                except ValueError:
                    lim = 0
                limit = lim if lim > 0 else None

                out_path = self.out_path.get().strip() or \
                    f"adal_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                self.scraper = AdalScraper(delay=delay, limit=limit,
                                           log_callback=self.log_message,
                                           progress_callback=self.update_progress)
                result = self.scraper.scrape(targets)
                rows = result["rows"]
                if rows:
                    write_excel(rows, out_path)
                self.root.after(0, lambda: self._on_complete(result, rows, out_path))
            except Exception as e:
                self.root.after(0, lambda: self._on_error(e))

        def _on_complete(self, result, rows, out_path):
            self.progress["value"] = 100
            self.run_button.config(state=tk.NORMAL, bg=COLORS["success"])
            self.stop_button.config(state=tk.DISABLED, bg=COLORS["text_light"])
            stopped = result.get("stopped")
            self.status_text.set("処理を停止しました" if stopped else "処理完了")
            if rows:
                msg = (f"処理を{'停止' if stopped else '完了'}しました。\n\n"
                       f"取得商品数: {result.get('products', 0)}\n"
                       f"出力行数: {len(rows)} 行\n"
                       f"出力先: {out_path}")
                messagebox.showinfo("停止" if stopped else "完了", msg)
            else:
                messagebox.showwarning("結果", "データが取得できませんでした。")

        def _on_error(self, error):
            self.status_text.set("エラーが発生しました")
            self.run_button.config(state=tk.NORMAL, bg=COLORS["success"])
            self.stop_button.config(state=tk.DISABLED, bg=COLORS["text_light"])
            self.log_message(f"\n❌ エラー: {error}")
            messagebox.showerror("エラー", f"エラーが発生しました:\n{error}")

        def on_closing(self):
            if self.scraper:
                self.scraper.stop()
            try:
                self.root.quit()
                self.root.destroy()
            except Exception:
                pass

    root = tk.Tk()
    ScraperGUI(root)
    root.mainloop()


def main():
    # 引数なし → GUI、引数あり → CLI
    if len(sys.argv) > 1:
        run_cli()
    else:
        launch_gui()


if __name__ == "__main__":
    main()
