#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
from datetime import datetime
import random
import re
import os
import tempfile
import zipfile
from pathlib import Path
from io import BytesIO

# ===== ãƒšãƒ¼ã‚¸è¨­å®š =====
st.set_page_config(page_title="å•†å“ç”»åƒãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ€ãƒ¼", page_icon="ğŸ“¦", layout="wide")


# ===== ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰èªè¨¼ =====
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    password_input = st.text_input("ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„", type="password")
    if password_input:
        if password_input == st.secrets.get("password", ""):
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒæ­£ã—ãã‚ã‚Šã¾ã›ã‚“")
    return False


if not check_password():
    st.stop()


# ===== å®šæ•° =====
SURUGAYA_SEARCH_URL = "https://www.suruga-ya.jp/kaitori/search_buy"
BOOKOFF_BASE_URL = "https://shopping.bookoff.co.jp/search/keyword/"
NO_IMAGE_PATTERNS = ['item_ll.gif', 'no_image', 'noimage', 'no-image', 'now_printing']
MIN_FILE_SIZE = 2 * 1024  # 2KBï¼ˆã“ã‚Œä»¥ä¸‹ã¯ã‚¹ã‚­ãƒƒãƒ—ï¼‰
MAX_FOLDER_SIZE = 23 * 1024 * 1024  # 23MBï¼ˆãƒ•ã‚©ãƒ«ãƒ€ä¸Šé™ï¼‰
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}


# ===== Selenium ã‚»ãƒƒãƒˆã‚¢ãƒƒãƒ— =====
def setup_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--log-level=3')
    options.add_argument('--disable-blink-features=AutomationControlled')

    # Streamlit Cloud (Debian) ã®ã‚·ã‚¹ãƒ†ãƒ  Chromium ã‚’æ¤œç´¢
    for chrome_path in ['/usr/bin/chromium-browser', '/usr/bin/chromium', '/usr/bin/google-chrome']:
        if os.path.exists(chrome_path):
            options.binary_location = chrome_path
            break

    driver = None
    for driver_path in ['/usr/bin/chromedriver', '/usr/lib/chromium-browser/chromedriver',
                        '/usr/lib/chromium/chromedriver']:
        if os.path.exists(driver_path):
            service = Service(driver_path)
            driver = webdriver.Chrome(service=service, options=options)
            break

    if driver is None:
        # ãƒ­ãƒ¼ã‚«ãƒ«ç’°å¢ƒ: webdriver-manager ã§ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

    # é§¿æ²³å±‹ã‚»ãƒ¼ãƒ•ã‚µãƒ¼ãƒè¨­å®š
    driver.get("https://www.suruga-ya.jp/")
    driver.add_cookie({'name': 'safe_search_option', 'value': '3', 'domain': '.suruga-ya.jp'})
    return driver


# ===== ãƒ˜ãƒ«ãƒ‘ãƒ¼é–¢æ•° =====
def is_no_image(url: str) -> bool:
    url_lower = url.lower()
    return any(p in url_lower for p in NO_IMAGE_PATTERNS)


def sanitize(name) -> str | None:
    if not name or str(name).strip() == "":
        return None
    return re.sub(r'[<>:"/\\|?*]', '_', str(name)).strip('. ')


# ===== ç”»åƒå–å¾—é–¢æ•° =====
def get_amazon_images(driver, asin: str, main_only: bool) -> list:
    url = f"https://www.amazon.co.jp/dp/{asin}"
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 3))
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        images = []

        main_img = soup.find('img', {'id': 'landingImage'})
        if main_img:
            src = main_img.get('data-old-hires') or main_img.get('src')
            if src:
                src = re.sub(r'_AC_[A-Z]{2}\d+_', '_AC_SL1500_', src)
                images.append(src)

        if not main_only:
            alt_div = soup.find('div', {'id': 'altImages'})
            if alt_div:
                for thumb in alt_div.find_all('img'):
                    t_src = thumb.get('src')
                    if t_src and 'video' not in t_src.lower():
                        h_res = re.sub(r'_AC_[A-Z]{2}\d+,?\d*_', '_AC_SL1500_', t_src)
                        h_res = re.sub(r'\._[A-Z]{2}\d+,?\d*_\.', '._SL1500_.', h_res)
                        if h_res not in images and not is_no_image(h_res):
                            images.append(h_res)
        return images
    except Exception:
        return []


def get_surugaya_images(driver, jan: str) -> list:
    url = f"{SURUGAYA_SEARCH_URL}?search_word={jan}&key_flag=1"
    try:
        driver.get(url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        title_a = soup.select_one('div.title a')
        if not title_a:
            return []

        detail_url = title_a['href']
        if detail_url.startswith('/'):
            detail_url = "https://www.suruga-ya.jp" + detail_url

        driver.get(detail_url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        img_up = soup.find('div', {'id': 'imgUp'})
        if img_up and img_up.find('a'):
            img_url = img_up.find('a')['href']
            if img_url.startswith('/'):
                img_url = "https://www.suruga-ya.jp" + img_url
            return [img_url]
    except Exception:
        pass
    return []


def get_bookoff_images(driver, jan: str) -> list:
    url = f"{BOOKOFF_BASE_URL}{jan}"
    try:
        driver.get(url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        img_tag = soup.select_one('img.js-gridImg, .productItem__image img')
        if img_tag and img_tag.get('src'):
            img_url = img_tag['src'].replace('/SS/', '/LL/').replace('SS.jpg', 'LL.jpg')
            return [img_url]
    except Exception:
        pass
    return []


# ===== ãƒ•ã‚©ãƒ«ãƒ€æŒ¯ã‚Šåˆ†ã‘ç®¡ç† =====
class FolderManager:
    """23MBä»¥ä¸‹ã§ãƒ•ã‚©ãƒ«ãƒ€ã‚’è‡ªå‹•æŒ¯ã‚Šåˆ†ã‘"""

    def __init__(self, output_base: Path):
        self.output_base = output_base
        self.folder_index = 1
        self.folder_size = 0
        self.folder_names = []  # å„ãƒ•ã‚©ãƒ«ãƒ€ã®ãƒ‘ã‚¹ã‚’è¨˜éŒ²

    def get_folder(self, files_size: int, base_fname: str) -> Path:
        """ç¾åœ¨ã®ãƒ•ã‚©ãƒ«ãƒ€ã‚’è¿”ã™ã€‚å®¹é‡è¶…éãªã‚‰æ¬¡ã®ãƒ•ã‚©ãƒ«ãƒ€ã¸ã€‚
        ç”»åƒã‚»ãƒƒãƒˆã¯åŒä¸€ãƒ•ã‚©ãƒ«ãƒ€ã«æ ¼ç´ã™ã‚‹ã€‚
        ãƒ•ã‚©ãƒ«ãƒ€åã¯æœ€åˆã«æ ¼ç´ã•ã‚Œã‚‹å•†å“ã®base_fnameã‚’ä½¿ç”¨ã€‚"""
        if self.folder_size > 0 and self.folder_size + files_size > MAX_FOLDER_SIZE:
            self.folder_index += 1
            self.folder_size = 0

        if self.folder_index > len(self.folder_names):
            # æ–°ã—ã„ãƒ•ã‚©ãƒ«ãƒ€: æœ€åˆã®å•†å“åã§å‘½å
            folder_name = base_fname.lower()
            folder_path = self.output_base / folder_name
            folder_path.mkdir(parents=True, exist_ok=True)
            self.folder_names.append(folder_path)
        else:
            folder_path = self.folder_names[self.folder_index - 1]

        return folder_path

    def add_size(self, size: int):
        self.folder_size += size


# ===== ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ï¼†ãƒ•ã‚£ãƒ«ã‚¿ =====
def download_and_filter_images(session, images, base_fname, folder_mgr, main_only):
    """ç”»åƒã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã€2KBä»¥ä¸‹ã‚’ã‚¹ã‚­ãƒƒãƒ—ã€ãƒ•ã‚©ãƒ«ãƒ€ã«æŒ¯ã‚Šåˆ†ã‘ä¿å­˜"""
    valid_images = []  # (fname, content) ã®ãƒªã‚¹ãƒˆ
    for idx, url in enumerate(images):
        if idx > 0 and main_only:
            break
        try:
            resp = session.get(url, timeout=30, headers=HEADERS)
            if resp.status_code == 200:
                content = resp.content
                if len(content) > MIN_FILE_SIZE:
                    suffix = "" if len(valid_images) == 0 else f"_{len(valid_images)}"
                    fname = f"{base_fname}{suffix}.jpg".lower()
                    valid_images.append((fname, content))
        except Exception:
            pass

    if not valid_images:
        return 0

    # ã‚»ãƒƒãƒˆå…¨ä½“ã®ã‚µã‚¤ã‚ºã‚’è¨ˆç®—
    total_size = sum(len(c) for _, c in valid_images)

    # ãƒ•ã‚©ãƒ«ãƒ€ã‚’å–å¾—ï¼ˆ23MBåˆ¶é™ã§è‡ªå‹•æŒ¯ã‚Šåˆ†ã‘ï¼‰
    folder = folder_mgr.get_folder(total_size, base_fname)

    saved_count = 0
    saved_size = 0
    for fname, content in valid_images:
        try:
            with open(folder / fname, 'wb') as f:
                f.write(content)
            saved_count += 1
            saved_size += len(content)
        except Exception:
            pass

    folder_mgr.add_size(saved_size)
    return saved_count


# ===== ZIP ä½œæˆ =====
def create_zip_files(output_base: Path, folder_mgr) -> Path:
    """å„ãƒ•ã‚©ãƒ«ãƒ€ã®ä¸­èº«ã‚’ãƒ•ãƒ©ãƒƒãƒˆãªZIPã«åœ§ç¸®ã€‚
    ZIPå = ãƒ•ã‚©ãƒ«ãƒ€å†…ã®æœ€åˆã®ãƒ•ã‚¡ã‚¤ãƒ«åï¼ˆæ‹¡å¼µå­ãªã—ï¼‰.zip"""
    zip_dir = output_base / "åœ§ç¸®ãƒ•ã‚¡ã‚¤ãƒ«"
    zip_dir.mkdir(parents=True, exist_ok=True)

    for folder_path in folder_mgr.folder_names:
        if not folder_path.exists():
            continue

        files = sorted([f for f in folder_path.iterdir() if f.is_file()], key=lambda x: x.name)
        if not files:
            continue

        zip_filename = f"{files[0].stem}.zip"
        zip_path = zip_dir / zip_filename

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for file_path in files:
                zf.write(file_path, file_path.name)  # ãƒ•ãƒ©ãƒƒãƒˆæ ¼ç´

    return zip_dir


def create_final_zip(zip_dir: Path, excel_path: Path) -> BytesIO:
    """å€‹åˆ¥ZIPãƒ•ã‚¡ã‚¤ãƒ« + æ›´æ–°æ¸ˆã¿Excelã‚’1ã¤ã®ZIPã«ã¾ã¨ã‚ã‚‹"""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # å€‹åˆ¥ZIPã‚’æ ¼ç´
        for f in sorted(zip_dir.iterdir()):
            if f.is_file() and f.suffix == '.zip':
                zf.write(f, f.name)
        # æ›´æ–°æ¸ˆã¿Excelã‚’æ ¼ç´
        zf.write(excel_path, excel_path.name)
    zip_buffer.seek(0)
    return zip_buffer


# ===== ãƒ¡ã‚¤ãƒ³å‡¦ç† =====
def process(uploaded_file, main_only):
    session = requests.Session()

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir = Path(tmp_dir)
        excel_path = tmp_dir / uploaded_file.name
        with open(excel_path, 'wb') as f:
            f.write(uploaded_file.getvalue())

        wb = load_workbook(excel_path)
        ws = wb.active

        # ç”»åƒä¿å­˜ç”¨ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒª
        img_dir = tmp_dir / "images"
        img_dir.mkdir(exist_ok=True)

        rows = [r for r in range(2, ws.max_row + 1) if ws[f'C{r}'].value or ws[f'D{r}'].value]
        total = len(rows)

        if total == 0:
            st.warning("å‡¦ç†å¯¾è±¡ã®è¡ŒãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸï¼ˆCåˆ—ã¾ãŸã¯Dåˆ—ã«ãƒ‡ãƒ¼ã‚¿ãŒå¿…è¦ã§ã™ï¼‰")
            return None

        # ãƒ•ã‚©ãƒ«ãƒ€æŒ¯ã‚Šåˆ†ã‘ç®¡ç†
        folder_mgr = FolderManager(img_dir)

        # ãƒ–ãƒ©ã‚¦ã‚¶èµ·å‹•
        with st.status("å‡¦ç†ä¸­...", expanded=True) as status:
            status.update(label="ãƒ–ãƒ©ã‚¦ã‚¶ã‚’èµ·å‹•ä¸­...")
            try:
                driver = setup_driver()
            except Exception as e:
                st.error(f"ãƒ–ãƒ©ã‚¦ã‚¶èµ·å‹•ã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
                return None

            progress_bar = st.progress(0)
            status_text = st.empty()
            log_area = st.empty()

            stats = {'total': total, 'success': 0, 'not_found': 0}
            logs = []

            try:
                for i, r in enumerate(rows, 1):
                    progress_bar.progress(i / total)
                    status.update(label=f"å‡¦ç†ä¸­: {i} / {total} ({i / total * 100:.1f}%)")
                    status_text.text(f"å‡¦ç†ä¸­: {i} / {total}")

                    seq = str(ws[f'B{r}'].value or "0").zfill(6)
                    jan = str(ws[f'C{r}'].value or "").strip()
                    asin = str(ws[f'D{r}'].value or "").strip()
                    shelf = sanitize(ws[f'K{r}'].value) or "00"
                    base_code = sanitize(ws[f'M{r}'].value) or "XX"
                    base_fname = f"{shelf}-{base_code}-{seq}"

                    product_name = str(ws[f'E{r}'].value or "").strip()
                    if len(product_name) > 30:
                        product_name = product_name[:30] + "..."

                    images = []
                    source_site = ""

                    # Amazon â†’ é§¿æ²³å±‹ â†’ ãƒ–ãƒƒã‚¯ã‚ªãƒ•
                    if asin and asin not in ["-", ""]:
                        images = get_amazon_images(driver, asin, main_only)
                        if images:
                            source_site = "Amazon"

                    if not images and jan:
                        images = get_surugaya_images(driver, jan)
                        if images:
                            source_site = "é§¿æ²³å±‹"

                    if not images and jan:
                        images = get_bookoff_images(driver, jan)
                        if images:
                            source_site = "ãƒ–ãƒƒã‚¯ã‚ªãƒ•"

                    timestamp = datetime.now().strftime("%H:%M:%S")

                    if images:
                        downloaded_count = download_and_filter_images(
                            session, images, base_fname, folder_mgr, main_only
                        )
                        if downloaded_count > 0:
                            ws[f'J{r}'].value = downloaded_count
                            stats['success'] += 1
                            log_msg = f"[{timestamp}] [{i}/{total}] âœ… {source_site} / {product_name} / ç”»åƒ{downloaded_count}æš"
                        else:
                            stats['not_found'] += 1
                            log_msg = f"[{timestamp}] [{i}/{total}] âš ï¸ {source_site} / {product_name} / æœ‰åŠ¹ãªç”»åƒãªã—"
                    else:
                        stats['not_found'] += 1
                        log_msg = f"[{timestamp}] [{i}/{total}] âŒ å–å¾—å¤±æ•— / {product_name}"

                    logs.append(log_msg)
                    log_area.code("\n".join(logs[-50:]))  # ç›´è¿‘50ä»¶è¡¨ç¤º

                    time.sleep(random.uniform(0.5, 1.0))

            finally:
                driver.quit()

            # ZIPåœ§ç¸®
            status.update(label="ZIPåœ§ç¸®ä¸­...")
            zip_dir = create_zip_files(img_dir, folder_mgr)

            status.update(label="âœ… å‡¦ç†å®Œäº†", state="complete", expanded=False)

        # Excel ä¿å­˜
        wb.save(excel_path)

        # æœ€çµ‚ZIPä½œæˆï¼ˆå€‹åˆ¥ZIP + Excelï¼‰
        zip_buffer = create_final_zip(zip_dir, excel_path)

        stats['folder_count'] = len(folder_mgr.folder_names)

        return {
            'zip': zip_buffer,
            'stats': stats,
            'logs': logs,
            'filename': Path(uploaded_file.name).stem
        }


# ===== ã²ãªå½¢Excelç”Ÿæˆ =====
@st.cache_data
def create_template_excel() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "å•†å“ãƒªã‚¹ãƒˆ"

    FONT = "ãƒ¡ã‚¤ãƒªã‚ª"
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # åˆ—å®šç¾©: (ãƒ˜ãƒƒãƒ€ãƒ¼å, å¿…é ˆãƒ¬ãƒ™ãƒ«, å¹…)
    # å¿…é ˆãƒ¬ãƒ™ãƒ«: "required"=å¿…é ˆ, "optional"=ä»»æ„, "auto"=è‡ªå‹•, "unused"=æœªä½¿ç”¨
    columns = {
        'A': ('', 'unused', 4),
        'B': ('é€£ç•ª', 'required', 8),
        'C': ('JANã‚³ãƒ¼ãƒ‰ â˜…', 'required', 18),
        'D': ('ASIN â˜…', 'required', 16),
        'E': ('å“å', 'optional', 30),
        'F': ('ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒª1', 'optional', 14),
        'G': ('ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒª2', 'optional', 14),
        'H': ('ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒª3', 'optional', 14),
        'I': ('ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒª4', 'optional', 14),
        'J': ('DLæšæ•°', 'auto', 10),
        'K': ('æ£šç•ª', 'required', 8),
        'L': ('', 'unused', 4),
        'M': ('æ‹ ç‚¹ã‚³ãƒ¼ãƒ‰', 'required', 12),
        'N': ('ãƒ•ã‚¡ã‚¤ãƒ«åï¼ˆè‡ªå‹•ï¼‰', 'auto', 28),
    }

    # ãƒ˜ãƒƒãƒ€ãƒ¼è‰²ã®å®šç¾©
    fills = {
        'required': PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
        'optional': PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
        'auto':     PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
        'unused':   PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    font_colors = {
        'required': Font(name=FONT, bold=True, size=10, color="FFFFFF"),
        'optional': Font(name=FONT, bold=True, size=10, color="1F4E79"),
        'auto':     Font(name=FONT, bold=True, size=10, color="375623"),
        'unused':   Font(name=FONT, bold=True, size=10, color="808080"),
    }

    # 1è¡Œç›®: ãƒ˜ãƒƒãƒ€ãƒ¼
    for col, (title, level, width) in columns.items():
        cell = ws[f'{col}1']
        cell.value = title
        cell.font = font_colors[level]
        cell.fill = fills[level]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[col].width = width

    # 2è¡Œç›®: è£œè¶³èª¬æ˜
    notes = {
        'A': 'â–¼ 3è¡Œç›®ã‹ã‚‰å…¥åŠ›',
        'B': 'â€»å¿…é ˆ', 'C': 'â€»C or D å¿…é ˆ', 'D': 'â€»C or D å¿…é ˆ',
        'E': 'ä»»æ„', 'F': 'ä»»æ„', 'G': 'ä»»æ„', 'H': 'ä»»æ„', 'I': 'ä»»æ„',
        'J': 'â€»è‡ªå‹•è¨˜å…¥', 'K': 'â€»å¿…é ˆ', 'M': 'â€»å¿…é ˆ', 'N': 'â€»è‡ªå‹•è¡¨ç¤º',
    }
    note_font = Font(name=FONT, size=8, color="808080")
    note_highlight = Font(name=FONT, size=8, bold=True, color="FF0000")
    for col, note in notes.items():
        cell = ws[f'{col}2']
        cell.value = note
        cell.font = note_highlight if col == 'A' else note_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 3è¡Œç›®: ã‚µãƒ³ãƒ—ãƒ«ãƒ‡ãƒ¼ã‚¿
    sample = {
        'B': 1, 'C': '4902370549560', 'D': 'B0CHY3GYW4',
        'E': 'ã‚µãƒ³ãƒ—ãƒ«å•†å“å', 'F': 'ã‚²ãƒ¼ãƒ ', 'G': 'Switch',
        'K': '999', 'M': 'TK',
    }
    data_font = Font(name=FONT, size=10)
    sample_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col_letter in columns:
        cell = ws[f'{col_letter}3']
        cell.value = sample.get(col_letter)
        cell.font = data_font
        cell.fill = sample_fill
        cell.border = thin_border

    # Nåˆ—: ãƒ•ã‚¡ã‚¤ãƒ«åãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼é–¢æ•°ï¼ˆ3è¡Œç›®ã€œ102è¡Œç›®ï¼‰
    formula_font = Font(name=FONT, size=10, color="808080")
    for row in range(3, 103):
        cell = ws[f'N{row}']
        cell.value = f'=IF(B{row}="","",LOWER(K{row}&"-"&M{row}&"-"&TEXT(B{row},"000000")&".jpg"))'
        cell.font = formula_font
        if row == 3:
            cell.fill = sample_fill
        cell.border = thin_border

    # è¡Œã®é«˜ã•
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===== UI =====
st.title("ğŸ“¦ å•†å“ç”»åƒä¸€æ‹¬ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰")
st.caption("Excelãƒªã‚¹ãƒˆã«åŸºã¥ãã€ãƒãƒƒãƒˆä¸Šã®å•†å“ç”»åƒã‚’è‡ªå‹•åé›†ãƒ»æ•´ç†ã—ã¾ã™")

with st.container(border=True):
    col_upload, col_template = st.columns([3, 1])
    with col_upload:
        uploaded_file = st.file_uploader("Excelãƒ•ã‚¡ã‚¤ãƒ«", type=["xlsx"])
    with col_template:
        st.write("")  # ã‚¹ãƒšãƒ¼ã‚µãƒ¼
        st.write("")
        st.download_button(
            label="ğŸ“„ ã²ãªå½¢ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
            data=create_template_excel(),
            file_name="å•†å“ç”»åƒDL_ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    mode = st.radio("å–å¾—ãƒ¢ãƒ¼ãƒ‰", ["å…¨ç”»åƒã‚’å–å¾—", "ãƒ¡ã‚¤ãƒ³ã®ã¿"], horizontal=True)
    st.info(
        "**å‘½åè¦å‰‡**: æ£šç•ª-æ‹ ç‚¹ã‚³ãƒ¼ãƒ‰-é€£ç•ª.jpgï¼ˆå°æ–‡å­—ï¼‰  \n"
        "**ãƒ•ã‚©ãƒ«ãƒ€æŒ¯ã‚Šåˆ†ã‘**: å„ãƒ•ã‚©ãƒ«ãƒ€23MBä»¥ä¸‹ã§è‡ªå‹•åˆ†å‰²  \n"
        "**ZIPå½¢å¼**: ãƒ•ã‚©ãƒ«ãƒ€ã”ã¨ã«ãƒ•ãƒ©ãƒƒãƒˆãªZIPï¼ˆZIPåï¼å…ˆé ­ãƒ•ã‚¡ã‚¤ãƒ«åï¼‰  \n"
        "**ãƒ•ã‚£ãƒ«ã‚¿**: 2KBä»¥ä¸‹ã®ç”»åƒã¯è‡ªå‹•ã‚¹ã‚­ãƒƒãƒ—"
    )

if uploaded_file:
    if st.button("â–¶ å®Ÿè¡Œé–‹å§‹", type="primary", use_container_width=True):
        result = process(uploaded_file, mode == "ãƒ¡ã‚¤ãƒ³ã®ã¿")

        if result:
            st.divider()
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("åˆè¨ˆ", result['stats']['total'])
            col2.metric("æˆåŠŸ", result['stats']['success'])
            col3.metric("æœªå–å¾—", result['stats']['not_found'])
            col4.metric("ZIPãƒ•ã‚¡ã‚¤ãƒ«æ•°", result['stats']['folder_count'])

            st.download_button(
                label="ğŸ“¥ çµæœã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ï¼ˆZIPï¼‰",
                data=result['zip'],
                file_name=f"{result['filename']}_images.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )
